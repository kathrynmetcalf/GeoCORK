import csv
import os
import sqlite3
import sys
import time

import qtawesome
from PyQt6 import QtCore, QtWidgets
from PyQt6.QtCore import Qt, QAbstractTableModel, QSize
from PyQt6.QtGui import QDesktopServices, QShowEvent
from PyQt6.QtSql import QSqlDatabase, QSqlQueryModel, QSqlQuery, QSqlTableModel
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QTableView,
    QLabel, QCheckBox, QSpacerItem, QComboBox,
    QSizePolicy, QTabWidget, QInputDialog, QDialog, QListWidget, QHBoxLayout, QMessageBox, QGroupBox, QScrollArea,
    QHeaderView, QAbstractItemView, QListWidgetItem
)
from PyQt6.uic import loadUi
from openpyxl import Workbook

import logger_setup
from Functions.Database_manager import update_database
from Functions import ExportDatabase, Settings_manager
from Functions import SQLUtils
from Functions.Database_manager import turn_on_foreign_keys, turn_off_foreign_keys
from Functions.Database_views import ViewQuery
from Functions.Widget_classes import (ReadableProxyModel, SQLiteTableModel, find_parent_items, show_loading_dialog,
                                      close_loading_dialog, columns_as_list, get_total_records, populate_combo_box,
                                      find_current_sub_items)
from Functions.Settings_manager import SettingsManager

settings = SettingsManager().settings
from Functions.Widget_classes import CheckableComboBox
from ui import Filters
from ui.DisplayTablesSimplified import DisplayTablesSimplified
from ui.FlowLayout import FlowLayout


class ExportWidget(QWidget):
    """
    ExportWidget is the main widget for exporting data from the database. The exporter allows users to select columns
    to export, apply filters, and save the data to various formats such as CSV, Excel, or Database.
    """

    def __init__(self, parent=None, database: QSqlDatabase = QSqlDatabase()):
        super().__init__(parent)
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        sources_ui_file = os.path.join(base_path, "ExporterUI.ui")
        loadUi(sources_ui_file, self)

        self.database = database

        self.checked_sample_list = []
        """List of SampleIDs that are currently checked to be included in the export"""

        self.checked_sample_ids_str = '()'
        """checked SampleIDs in the format (1, 2, 3) to be used in the SQL query to limit the results to only those 
        samples"""

        self.checked_filter_list = []
        """list of FilterGroupIDs that are currently checked to filter selected data by, multiple filters can be 
        selected but are OR'd together, so if Filter 1 and Filter 2 are selected, then the both filter;s data will be
        included."""

        self.checked_filter_ids_str = '()'
        """checked FilterGroupIDs in the format (1, 2, 3) to be used in the SQL query to limit the results to only 
        those filter groups"""

        self.checked_grouped_filter_list = []
        """list of FilterGroupIDs that are currently checked to add as a grouped new sample. If FilterGroupID with name 
        Modern Samples is selected, all samples matching the criteria will be added to the exporter as a distinct sample
        called 'Modern Samples'."""

        self.filtered_upb_ids = set()
        """Set of filtered UPbAnalysisIDs matching the selected samples and filters. This is used to limit the results
        to only those UPbAnalysisIDs that match the selected samples and filters. Only these UPbAnalysisIDs and dependents
        will be exported to a new database"""

        self.column_name_mappings = dict()
        """dictionary containing the column names and their mappings to be used in the SQL query. This is used to rename columns"""

        self.worksheet_tabs_dict = {}
        """main dictionary to store all worksheet tabs."""
        # in the format of
        # {
        # 'Worksheet 1': {
        #   'tableView': tableView,
        #   'model': model,
        #   'distinct': False,
        #   'pivot': False,
        #   'selected_columns': {(table, attribute): bool, },
        #   'ordered_columns': [table.attribute_name]
        #   'label': counter_label,
        #   'headers': True,
        #   'sql': ''
        #   }
        # }

        self.active_filter_sample_ids = []
        """List of SampleIDs that match the current active filter, not saved, in the Filters tab. Used as additional samples"""

        self.previous_worksheet = None
        """variable to hold the name of the previous worksheet"""

        self.exportformat_comboBox: QComboBox
        # Clear all the existing items in the combo box
        while self.exportformat_comboBox.count() > 0:
            self.exportformat_comboBox.removeItem(0)
        for export_format in SQLUtils.export_formats:
            self.exportformat_comboBox.addItem(export_format)

        for widget in QApplication.topLevelWidgets():
            if widget.inherits("QMainWindow"):
                self.db_file = widget.db_file

        self.settings = SettingsManager().settings

        self.max_rows_to_display = 1000

        self.sample_count = 0

        self.use_converted_label: QLabel
        # Make the existing label text bold
        text = self.use_converted_label.text()
        self.use_converted_label.setText(f"<b>{text}</b>")

        self.refresh_button = QPushButton("")
        self.refresh_button.setObjectName("refreshbutton")
        self.refresh_button.setToolTip('Refreshes the widget and can fix issues')
        self.refresh_button.setMaximumSize(QSize(25, 40))

        self.add_worksheet_button = QPushButton("Add Worksheet")
        self.add_worksheet_button.setObjectName("add_worksheet_button")
        self.add_worksheet_button.setToolTip('Adds a worksheet to the exporter')

        self.remove_worksheet_button = QPushButton("Delete Worksheet")
        self.remove_worksheet_button.setObjectName("remove_worksheet_button")
        self.remove_worksheet_button.setToolTip('Deletes the currently viewed worksheet from the exporter')

        self.edit_columnnames_button = QPushButton("Edit Column Names")
        self.edit_columnnames_button.setObjectName("edit_columnnames_button")
        self.edit_columnnames_button.setToolTip('Edits the column names to be exported. Column names seen below are what will be exported, not what they are called in the database.')

        self.worksheetbutton_layout.addWidget(self.refresh_button)
        self.worksheetbutton_layout.addWidget(self.add_worksheet_button)
        self.worksheetbutton_layout.addWidget(self.remove_worksheet_button)
        self.worksheetbutton_layout.addWidget(self.edit_columnnames_button)


        self.editorder_button = QPushButton("Edit Column Order")
        self.editorder_button.setObjectName("editorder_button")
        self.editorder_button.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        self.editorder_button.setToolTip('Edits the order of selected columns to be viewed in the current worksheet')

        self.export_button = QPushButton("Export")
        self.export_button.setObjectName("export_button")
        self.export_button.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        self.export_button.setToolTip('Exports the current workbook with all options selected')

        self.settings_label = QLabel()

        self.options_layout.addWidget(self.editorder_button)
        self.options_layout.addWidget(self.export_button)
        self.options_layout.addWidget(self.settings_label)

        # Connect buttons to methods
        self.refresh_button.setIcon(qtawesome.icon('fa6s.rotate-right', color='green', scale_factor=1.2))
        self.refresh_button.clicked.connect(self.refresh_widget)

        self.add_worksheet_button.clicked.connect(lambda: self.add_worksheet_tab(None, False, False, {}, [], False))
        # self.add_worksheet_button.clicked.connect(self.update_table_view)

        self.remove_worksheet_button.clicked.connect(self.remove_current_worksheet_tab)
        self.remove_worksheet_button.clicked.connect(self.update_table_view)

        self.edit_columnnames_button.clicked.connect(self.open_columnname_mapping_dialog)


        self.export_button.clicked.connect(self.export_data)

        self.editorder_button.clicked.connect(self.open_column_order_dialog)

        self.active_filter_sample_checkBox.checkStateChanged.connect(self.update_table_view)

        self.columnselection_comboBox.addItems(SQLUtils.table_attributes_dict)

        self.populate_stack()


    def sync_selected_ordered_columns(self, worksheet_name: str = None):
        """
        :param str worksheet_name: worksheet name to view, otherwise the current index will be used
        """
        # Get the current workbook
        logger_setup.get_logger().info('Updating table view with new parameters')
        show_loading_dialog('Loading', 'Updating table view with new parameters')
        start_update_table_view_time = time.time()
        if worksheet_name is None:
            current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
            if not current_worksheet_name:
                self.export_format()
                current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        else:
            current_worksheet_name = worksheet_name
        worksheet_dict = self.worksheet_tabs_dict.copy()
        selected_columns = worksheet_dict[current_worksheet_name]['selected_columns']
        if selected_columns == {}:
            self.worksheet_tabs_dict[current_worksheet_name]['ordered_columns'] = []
            return
        ordered_columns = []
        for column_name in self.worksheet_tabs_dict[current_worksheet_name]['ordered_columns']:
            table = column_name.split('.')[0]
            field = column_name.split('.')[1]
            checked = selected_columns.get((table, field), False)
            if checked and column_name not in ordered_columns:
                ordered_columns.append(column_name)
            if checked and field not in self.column_name_mappings:
                self.column_name_mappings[field] = field
            if not checked and field in self.column_name_mappings:
                # remove the key from the column name mappings if the column is no longer selected
                del self.column_name_mappings[field]
        for (table, field) in selected_columns.keys():
            column_name = f'{table}.{field}'
            checked = selected_columns.get((table, field), False)
            if not checked and field in self.column_name_mappings:
                # remove the key from the column name mappings if the column is no longer selected
                del self.column_name_mappings[field]
            if checked and column_name not in ordered_columns:
                ordered_columns.append(column_name)
            if checked and field not in self.column_name_mappings:
                self.column_name_mappings[field] = field
        if self.worksheet_tabs_dict[current_worksheet_name]['ordered_columns'] != ordered_columns:
            self.worksheet_tabs_dict[current_worksheet_name]['ordered_columns'] = ordered_columns

    def update_table_view(self, order_changed: bool = False, worksheet_name: str = None):
        """
        Main method for the ExportWidget to display data to the user. This converts user-selections of columns, formats,
        and data transformers into SQL queries.
        :param bool order_changed: if the order/amount of columns has changed, then set selected_columns to order_columns
        :param str worksheet_name: worksheet name to view, otherwise the current index will be used
        """
        # Get the current workbook
        logger_setup.get_logger().info('Updating table view with new parameters')
        show_loading_dialog('Loading', 'Updating table view with new parameters')
        start_update_table_view_time = time.time()
        if worksheet_name is None:
            current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
            if not current_worksheet_name:
                self.export_format()
                current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        else:
            current_worksheet_name = worksheet_name

        self.update_active_filters()

        # Clear the previous filtered UPb IDs
        self.filtered_upb_ids = set()

        # Get the current selected samples, filters, and grouped filters
        self.checked_sample_list = self.samplesincluded_comboBox.source_model().return_checked_ids()[0]
        self.checked_sample_ids_str = f"({', '.join(map(str, self.checked_sample_list))})"

        self.checked_filter_list = self.filterselection_comboBox.source_model().return_checked_ids()[0]
        self.checked_filter_ids_str = f"({', '.join(map(str, self.checked_filter_list))})"

        self.checked_grouped_filter_list = self.groupedfilter_comboBox.source_model().return_checked_ids()[0]


        if self.exportformat_comboBox.currentText() != 'Database':
            # Get the current TableView
            tableView: QTableView = self.worksheet_tabs_dict[current_worksheet_name]['tableView']
            # Get the selected columns for the current workbook
            self.sync_selected_ordered_columns(current_worksheet_name)
            if not self.worksheet_tabs_dict[current_worksheet_name]['ordered_columns']:
                # prevents unnecessary compute time
                # No columns selected, clear the table view
                tableView.setSortingEnabled(False)
                tableView.setModel(None)
                close_loading_dialog('Loading', 'Updating table view with new parameters')
                return False

        tables = set()
        # always ensures UPbAnalyses in the resulting query, prevents edge cases
        tables.add('UPbAnalyses')
        if self.exportformat_comboBox.currentText() == 'Database':
            # if the export format is Database, then we only need the list of analyses that pass the filters
            columns_str = 'UPbAnalyses.UPbAnalysisID'
            concat_col_str = ''
        else:
            columns_str = ''
            # creates column select string in format [SampleID], [CalculatedU/Th] AS 'RenamedColumn', etc...
            concat_col_str = ''
            # creates column select string for concatenated columns, in format REPLACE(GROUP_CONCAT(DISTINCT [field]), ',', '; ') AS 'RenamedColumn'
            for column_name in self.worksheet_tabs_dict[current_worksheet_name]['ordered_columns']:
                table = column_name.split('.')[0]
                field = column_name.split('.')[1]
                tables.add(table)
                # If the export format is detritalPy and the current worksheet is Samples and the field is not Sample ID or name,
                # we need to group_concat other fields
                if (self.exportformat_comboBox.currentText() == 'detritalPy' and
                        current_worksheet_name == 'Samples'):
                    if field not in ('SampleID', 'SampleName'):
                        concat_field_str = f"REPLACE(GROUP_CONCAT(DISTINCT [{field}]), ',', '; ')"
                    else:
                        concat_field_str = f"[{field}]"
                else:
                    concat_field_str = ''
                field_str = f"[{field}]"
                for key, values in SQLUtils.many_editable.items():
                    if field in values:
                        # If the field is in the many_editable dictionary, then there may be multiple values for it,
                        # so we need to use GROUP_CONCAT
                        field_str = f"REPLACE(GROUP_CONCAT(DISTINCT [{field}]), ',', '; ')"
                        break
                # if "GROUP_CONCAT" not in field_str:
                #     for key, values in SQLUtils.one_editable.items():
                #         if key != 'Samples':
                #             if field in values:
                #                 # If the field is in the one_editable dictionary, then we need to use GROUP_CONCAT
                #                 field_str = f"REPLACE(GROUP_CONCAT(DISTINCT [{field}]), ',', '; ')"
                #                 break
                if field in self.column_name_mappings:
                    columns_str += f"{field_str} AS '{self.column_name_mappings[field]}', "
                    if concat_field_str:
                        concat_col_str += f"{concat_field_str} AS '{self.column_name_mappings[field]}', "
                else:
                    columns_str += f'{field_str}, '
                    if concat_field_str:
                        concat_col_str += f'{concat_field_str}, '

            # remove final ", "
            columns_str = columns_str[0:-2]
            if concat_col_str:
                concat_col_str = concat_col_str[0:-2]
            # always ensures samples is included, since tables is a set only one copy will exist
        tables.add('Samples')

        # gets final join from all found tables.
        join = SQLUtils.get_join_from_table("", list(tables))

        filtered_where_clause = ''
        self.filtered_upb_ids = set()
        """Set of filtered filtered_upb_ids"""
        # Filters for filters step, so if Samples1,2,3 are selected but only want bestage<500ma this
        # section finds the UPbAnalysisID that match the criteria. Multiple filters used will be OR'd together
        # so if Filter 1 includes (1, 3, 5) and Fiter 2 includes (2, 4) items listed would be (1, 2, 3, 4, 5)
        for filter_id in self.checked_filter_list:
            json_query = QSqlQuery()
            json_query.prepare('SELECT SQLQuery FROM FilterGroups WHERE FilterGroupID = :filter_id')
            json_query.bindValue(':filter_id', filter_id)
            if not json_query.exec():
                logger_setup.get_logger().critical(f'Could not fetch SQL Query for filter')
                logger_setup.get_logger().info(f'Filter ID: {filter_id}')

            json_query.next()
            filter_json = json_query.value(0)

            # loops through each filter in the checked filter list, processes the json to sql
            filtered_where_clause, ctes = Filters.process_json_to_sql(filter_json[1:-1], scope='UPbAnalyses')
            filtered_where_clause = filtered_where_clause[0:-1]

            sql_query = f"SELECT DISTINCT UPbAnalysisID FROM ({filtered_where_clause});"
            uri = f'file:{settings.value('db_file', type=str)}?mode=ro&immutable=1'
            # Execute the query
            logger_setup.get_logger().info(f'Fetching distinct UPbAnalyisIDs from FilterID: {filter_id}')
            logger_setup.get_logger().debug(f'SQL command: {sql_query}')
            try:
                conn = sqlite3.connect(uri, uri=True)
                with conn:
                    cursor = conn.cursor()
                    cursor.execute(sql_query)
                    filtered_upb_ids = [str(row[0]) for row in cursor.fetchall()]
                conn.commit()
                conn.close()
            except sqlite3.Error as e:
                if 'no such column' in e.text():
                    sql_query = f"SELECT DISTINCT UPbAnalysisID FROM ({filtered_where_clause});"
                    try:
                        conn = sqlite3.connect(uri, uri=True)
                        with conn:
                            cursor = conn.cursor()
                            cursor.execute(sql_query)
                            filtered_upb_ids = [str(row[0]) for row in cursor.fetchall()]
                        conn.commit()
                        conn.close()
                    except sqlite3.Error as e:
                        logger_setup.get_logger().critical(f'Error fetching distinct values')
                        logger_setup.get_logger().debug(f'Error: {e}')
                        logger_setup.get_logger().debug(f'SQL query: {sql_query}')
                        close_loading_dialog('Loading', 'Updating table view with new parameters')
                        return False
                else:
                    logger_setup.get_logger().critical(f'Error fetching distinct values')
                    logger_setup.get_logger().debug(f'Error: {e}')
                    logger_setup.get_logger().debug(f'SQL query: {sql_query}')
                    close_loading_dialog('Loading', 'Updating table view with new parameters')
                    return False
                logger_setup.get_logger().critical(f"Error fetching distinct values")
                logger_setup.get_logger().debug(f"Error: {e}")
                logger_setup.get_logger().debug(f"SQL query: {sql_query}")
                close_loading_dialog('Loading', 'Updating table view with new parameters')
                return None

            logger_setup.get_logger().info(f'Fetched distinct UPbAnalysisIDs from FilterID: {filter_id} sucessfully')
            # add the filtered UPbAnalysisIDs to the set
            self.filtered_upb_ids.update(filtered_upb_ids)

        logger_setup.get_logger().info(f'Number of Filtered UPbAnalysis IDs Found: {len(self.filtered_upb_ids)}')

        if self.exportformat_comboBox.currentText() == 'Database':
            # If the export format is Database, then we only need the list of analyses that pass the filters
            # The set of analyses that pass the filters is already stored in self.filtered_upb_ids
            self.update_database_export()
            close_loading_dialog('Loading', 'Updating table view with new parameters')
            return

        # due to how the above logic is, the filters are added with an OR clause, therefore it full unions Filters 1 and 2
        filtered_upb_ids_sql = f"({', '.join(self.filtered_upb_ids)})"

        # checks for logic to see what kind of SQL query is needed.
        # self.checked_sample_ids_str defaults to '()', so length of 2,
        # if a sample is checked then len > 2, so UPbAnalysisID are needed, so we limit to LIMIT {self.max_rows_to_display} so its quicker and
        # still shows example data to be exported.
        # if filtered where clause is not blank, len > 0, then we need to filter by UPbAnalysisID
        if "AS 'Sample_ID'" in columns_str and current_worksheet_name == 'Samples':
            group_by = 'GROUP BY Sample_ID'
        else:
            group_by = 'GROUP BY UPbAnalyses.UPbAnalysisID'
        if len(self.checked_sample_ids_str) > 2:
            if len(filtered_where_clause) > 0:
                where_clause = f"WHERE Samples.SampleID IN {self.checked_sample_ids_str} AND UPbAnalyses.UPbAnalysisID IN {filtered_upb_ids_sql}"
            else:
                where_clause = f"WHERE Samples.SampleID IN {self.checked_sample_ids_str}"
        else:
            if len(filtered_where_clause) > 0:
                where_clause = f"WHERE UPbAnalyses.UPbAnalysisID IN {filtered_upb_ids_sql}"
            else:
                where_clause = f"WHERE FALSE"
        sample_query_str = f"""SELECT {'DISTINCT' if self.worksheet_tabs_dict[current_worksheet_name]['distinct'] is True else ''} 
                                {columns_str} FROM Samples {join}
                                {where_clause}
                                {group_by} LIMIT {self.max_rows_to_display}"""

        logger_setup.get_logger().debug(f'Final TableView SQL command: {sample_query_str}')

        # code to add optional grouped filters as a new Sample ID, if a filter name is 'Modern River Sand'
        # and returns UPbAnalysesIDs from multiple samples it will group them all together as
        # SampleName = 'Modern River Sand'
        if len(self.checked_grouped_filter_list) > 0 and "FALSE" in sample_query_str.split('WHERE')[1]:
            # If the WHERE clause is 'WHERE FALSE', this will return an extra row with NULL values, so we need to remove it
            query_str = ""
        else:
            query_str = sample_query_str
        for filter_id in self.checked_grouped_filter_list:
            json_query = QSqlQuery()
            json_query.prepare('SELECT FilterGroupName, SQLQuery FROM FilterGroups WHERE FilterGroupID = :filter_id')
            json_query.bindValue(':filter_id', filter_id)
            if not json_query.exec():
                logger_setup.get_logger().critical(f'Could not fetch SQL Query for filter')
                logger_setup.get_logger().info(f'Filter ID: {filter_id}')

            json_query.next()
            name = json_query.value(0)
            filter_json = json_query.value(1)

            logger_setup.get_logger().info('Fetching Filters for Grouped Filter List')
            # loops through each filter in the checked filter list, processes the json to sql
            filtered_where_clause, ctes = Filters.process_json_to_sql(filter_json[1:-1], scope='UPbAnalyses')
            filtered_where_clause = filtered_where_clause[0:-1]

            sql_query = f"SELECT DISTINCT UPbAnalysisID FROM ({filtered_where_clause});"

            uri = f'file:{settings.value('db_file', type=str)}?mode=ro&immutable=1'
            # Execute the query
            logger_setup.get_logger().info(f'Fetching distinct UPbAnalyisIDs from FilterID: {filter_id}')
            logger_setup.get_logger().debug(f'SQL command: {sql_query}')
            try:
                conn = sqlite3.connect(uri, uri=True)
                with conn:
                    cursor = conn.cursor()
                    cursor.execute(sql_query)
                    filtered_upb_ids = [str(row[0]) for row in cursor.fetchall()]
                    # add the filtered UPbAnalysisIDs to the set
                    self.filtered_upb_ids.update(filtered_upb_ids)
                    filtered_upb_ids = f"({', '.join(filtered_upb_ids)})"
                conn.commit()
                conn.close()
            except sqlite3.Error as e:
                logger_setup.get_logger().critical(f"Error fetching distinct values")
                logger_setup.get_logger().debug(f"Error: {e}")
                logger_setup.get_logger().debug(f"SQL query: {sql_query}")
                close_loading_dialog('Loading', 'Updating table view with new parameters')
                return None

            logger_setup.get_logger().info(f'Fetched distinct UPbAnalysisIDs from FilterID: {filter_id} sucessfully')

            # remove LIMIT {self.max_rows_to_display} from original query_str, can only have one of those
            query_str = query_str.replace(f'LIMIT {self.max_rows_to_display}', '')
            # take the original sample_query_str and only the content before WHERE CLAUSE
            modified_query_str = sample_query_str.split('WHERE')[0]
            if concat_col_str:
                # If there are concatenated columns, we need to add them to the modified query string
                modified_query_str = modified_query_str.replace(columns_str, concat_col_str)
            # replace SampleName with filter name AS
            modified_query_str = modified_query_str.replace('[SampleName]', f'\'{name}\'')
            modified_query_str = modified_query_str.replace('SELECT', 'SELECT DISTINCT')
            modified_query_str = modified_query_str.replace(f'LIMIT {self.max_rows_to_display}', '')
            modified_query_str = modified_query_str.replace('DISTINCT DISTINCT', 'DISTINCT')
            # If this is the detritalPy export format and the worksheet is Samples, we need to group by SampleID
            if "AS 'Sample_ID'" in modified_query_str and current_worksheet_name == 'Samples':
                group_by = 'GROUP BY Sample_ID'
            else:
                group_by = ''

            if query_str:
                query_str = f"{query_str} \nUNION ALL \n {modified_query_str} WHERE UPbAnalyses.UPbAnalysisID IN {filtered_upb_ids} {group_by} LIMIT {self.max_rows_to_display} \n"
            else:
                query_str = f"{modified_query_str} WHERE UPbAnalyses.UPbAnalysisID IN {filtered_upb_ids} {group_by} LIMIT {self.max_rows_to_display} \n"
            logger_setup.get_logger().debug(f'SQL command: {query_str}')

        # code to transform the query into a pivot table
        # SQLite doesn't have a builtin Pivot function, so it must be done manually.
        if self.worksheet_tabs_dict[current_worksheet_name]['pivot']:
            query_str = query_str.replace(f'LIMIT {self.max_rows_to_display}', '')

            # Any transactions shouldn't be present at this time, but just in case
            db = QSqlDatabase.database()
            db.commit()
            db.close()
            db.open()
            if not db.isOpen():
                logger_setup.get_logger().critical('Error opening database connection')
                close_loading_dialog('Loading', 'Updating table view with new parameters')
                self.uncheck_pivot()
                return None
            if not turn_on_foreign_keys():
                close_loading_dialog('Loading', 'Updating table view with new parameters')
                self.uncheck_pivot()
                return None

            drop_table_qry = QSqlQuery()
            logger_setup.get_logger().info('Dropping TempPivotTable')
            if not drop_table_qry.exec('DROP TABLE IF EXISTS TempPivotTable'):
                logger_setup.get_logger().critical(
                    f'Error dropping TempPivotTable: {drop_table_qry.lastError().text()}')
                self.uncheck_pivot()
                close_loading_dialog('Loading', 'Updating table view with new parameters')
                return None

            uri = f'file:{settings.value('db_file', type=str)}'
            # Execute the query
            sql_temptable_create = 'CREATE TABLE TempPivotTable AS SELECT * FROM (' + query_str + ')'
            logger_setup.get_logger().debug(f'SQL command: {sql_temptable_create}')
            try:
                conn = sqlite3.connect(uri, uri=True)
                with conn:
                    cursor = conn.cursor()
                    logger_setup.get_logger().info('Creating table TempPivotTable')
                    cursor.execute(sql_temptable_create)
                conn.commit()
                conn.close()
            except sqlite3.Error as e:
                logger_setup.get_logger().critical(f"Error creating TempPivotTable")
                logger_setup.get_logger().debug(f"Error: {e}")
                logger_setup.get_logger().debug(f"SQL query: {sql_temptable_create}")
                self.uncheck_pivot()
                close_loading_dialog('Loading', 'Updating table view with new parameters')
                return None

            logger_setup.get_logger().info('Created table TempPivotTable successfully')

            # defaults to pivot based on the first column in the exporter.
            # tuple is in format (table, attribute)
            ordered_columns = self.worksheet_tabs_dict[current_worksheet_name]['ordered_columns']
            pivot_col = ordered_columns[0].split('.')[1]
            if 'Name' not in pivot_col:
                response = QMessageBox.question(self, 'Missing name',
                                                'Pivot may not work correctly without a name field in the first column.\nDo you want to continue?',
                                                QMessageBox.StandardButton.No | QMessageBox.StandardButton.No,
                                                QMessageBox.StandardButton.Yes)
                if response == QMessageBox.StandardButton.No:
                    # Uncheck the pivot checkbox
                    self.uncheck_pivot()
                    close_loading_dialog('Loading', 'Updating table view with new parameters')
                    return None
            pivot_col = self.column_name_mappings[pivot_col]

            first_column_list = []
            uri = f'file:{settings.value('db_file', type=str)}?mode=ro&immutable=1'
            # Execute the query
            sql_distinct_first_column = f'SELECT DISTINCT {pivot_col} FROM TempPivotTable ORDER BY {pivot_col}'
            logger_setup.get_logger().debug(f'SQL command: {sql_distinct_first_column}')
            try:
                conn = sqlite3.connect(uri, uri=True)
                with conn:
                    cursor = conn.cursor()
                    logger_setup.get_logger().info('Selecting distinct values from TempPivotTable first column')
                    cursor.execute(sql_distinct_first_column)
                    first_column_list = [row[0] for row in cursor.fetchall()]
                conn.commit()
                conn.close()
            except sqlite3.Error as e:
                logger_setup.get_logger().critical(f"Error selecting distinct values from TempPivotTable")
                logger_setup.get_logger().debug(f"Error: {e}")
                logger_setup.get_logger().debug(f"SQL query: {sql_distinct_first_column}")
                self.uncheck_pivot()
                close_loading_dialog('Loading', 'Updating table view with new parameters')
                return None

            if len(first_column_list) == 0:
                # if no columns/values are found then could be an error, check if items are checked, if there are
                # then something went wrong.
                if not len(self.checked_sample_list) == 0:
                    logger_setup.get_logger().error('No rows returned. Try broadening filters, samples, or additional filters.')
                    model = SQLiteTableModel()
                    proxy_model = ReadableProxyModel()
                    proxy_model.setSourceModel(model)
                    tableView.setModel(proxy_model)
                    self.uncheck_pivot()
                    close_loading_dialog('Loading', 'Updating table view with new parameters')
                    return False
                else:
                    tableView.setModel(None)
                    close_loading_dialog('Loading', 'Updating table view with new parameters')
                    return True
            case_expressions = []

            # Creates the column names for the first col and other columns, so if SampleID, BestAge is being pivot
            # with samples in the list as S1, S2, S3, then:
            # end result should be S1_BestAge, S2_BestAge, S3_BestAge
            columns_names = []
            for name in first_column_list:
                for column_name in ordered_columns:
                    field = column_name.split('.')[1]
                    field_name = self.column_name_mappings[field]
                    if field_name not in columns_names:
                        columns_names.append(field_name)

                    if field_name == pivot_col:
                        continue
                    try:
                        if not isinstance(name, str):
                            name = str(name)
                        case_expressions.append(
                            f'MAX(CASE WHEN [{pivot_col}] = \'{name}\' THEN [{field_name}] END) AS [{name + "_" + field_name}]')
                    except Exception as e:
                        logger_setup.get_logger().critical(f"Error getting column name for {name}")
                        logger_setup.get_logger().debug(f"Error: {e}")
                        self.uncheck_pivot()
                        close_loading_dialog('Loading', 'Updating table view with new parameters')
                        return None

            case_list_sql = '\n, '.join(case_expressions)

            # final pivot string, takes the data from TempPivotTable and modifies it.
            for column in range(len(columns_names)):
                columns_names[column] = f'"{columns_names[column]}"'
            columns_names_str = ', '.join(columns_names)

            query_str = (f"""With cte AS (SELECT {columns_names_str}, ROW_NUMBER() OVER (
            PARTITION BY {pivot_col}
            ORDER BY rowid) AS RowNum
            FROM TempPivotTable)
            SELECT {case_list_sql}
            FROM cte c
            GROUP BY c.RowNum
            ORDER BY c.RowNum""")

        # At this point the final query_str is complete, either with or without pivot.
        # saves final string used for exporting, removed LIMIT, and saved model for future use.
        model = SQLiteTableModel(database = settings.value('db_file', type=str))
        model.setQuery(query_str)
        if model.last_error:
            if 'too many columns' in str(model.last_error):
                logger_setup.get_logger().critical(f"Too many rows to pivot")
                self.uncheck_pivot()
            else:
                logger_setup.get_logger().critical(f'Error updating table view')
            close_loading_dialog('Loading', 'Updating table view with new parameters')
            return None
        while model.canFetchMore():
            model.fetchMore()
        self.worksheet_tabs_dict[current_worksheet_name]['sql'] = query_str.replace(f'LIMIT {self.max_rows_to_display}', '')
        self.worksheet_tabs_dict[current_worksheet_name]['model'] = model

        # Remove LIMIT {self.max_rows_to_display} from the original query string and build the COUNT query, for the count label
        counter_sql_query = f"SELECT COUNT('UPbAnalyses') FROM ({self.worksheet_tabs_dict[current_worksheet_name]['sql']}) AS SubQuery"

        # Prepare and execute the query
        counter_query = QSqlQuery(db=self.database)
        # logger_setup.get_logger().debug(f"SQL Command: {counter_sql_query}")
        if not counter_query.exec(counter_sql_query):
            logger_setup.get_logger().critical(f'Error fetching total records')
            logger_setup.get_logger().debug(f'Error: {counter_query.lastError().text()}')
            logger_setup.get_logger().debug(f'SQL query: {counter_query.lastQuery()}')
            close_loading_dialog('Loading', 'Updating table view with new parameters')
            return
        else:
            # Move to the first record to retrieve the count
            if counter_query.next():
                count = counter_query.value(0)
                if count >= self.max_rows_to_display:
                    self.worksheet_tabs_dict[current_worksheet_name]['label'].setText(f"Showing {self.max_rows_to_display}/{count} rows")
                else:
                    self.worksheet_tabs_dict[current_worksheet_name]['label'].setText(f"Showing {count} rows")
            else:
                # Handle case where query doesn't return a result
                self.worksheet_tabs_dict[current_worksheet_name]['label'].setText(f"Number of Rows: 0")

        # while model.canFetchMore():
        #     model.fetchMore()
        proxy_model = ReadableProxyModel()
        proxy_model.setSourceModel(model)
        proxy_model.original_headers = True
        tableView.setModel(proxy_model)
        tableView.resizeColumnsToContents()
        close_loading_dialog('Loading', 'Updating table view with new parameters')
        logger_setup.get_logger().info(f'Updated table view in {time.time() - start_update_table_view_time:.2f} seconds')

    def update_active_filters(self):
        if not self.active_filter_sample_checkBox.isChecked():
            self.active_filter_sample_ids = []
            return
        tab_widget = self.parentWidget().parentWidget()
        if tab_widget is not None:
            try:
                filter_widget = tab_widget.findChild(Filters.Filters)
                self.active_filter_sample_ids = filter_widget.querybuilder.get_filtered_ids('Samples')
                if self.active_filter_sample_ids is None:
                    self.active_filter_sample_ids = []
            except Exception as e:
                logger_setup.get_logger().critical(f'Error updating active filter')
                return

        if len(self.active_filter_sample_ids) > 0 and self.active_filter_sample_checkBox.isChecked():
            # if there are active filter sample ids, and the checkbox is checked, add them to the checked sample list
            self.checked_sample_list.extend(self.active_filter_sample_ids)
            self.samplesincluded_comboBox.source_model().update_model_checks(set(self.checked_sample_list), set())
        elif not self.active_filter_sample_checkBox.isChecked():
            # if the checkbox is not checked, remove the active filter sample from the checked sample list
            checked_sample_list = [sample_id for sample_id in self.checked_sample_list if sample_id not in self.active_filter_sample_ids]
            self.checked_sample_list = checked_sample_list
            self.samplesincluded_comboBox.source_model().update_model_checks(set(self.checked_sample_list), set())

    def export_format(self):
        """
        Main method to swap between predefined export formats.
        :return:
        """

        match self.exportformat_comboBox.currentText():
            # DetritalPy requires an excel file, with multiple sheets
            # sheet 1 (Samples) is a distinct list of sample, units, basins(), age, lat, long, and source
            # sheet 2 (ZrUPb) is list of samples, grains, analysis, and upb data

            case 'detritalPy':
                self.clear_worksheet_data()
                self.fileformat_comboBox.setEnabled(True)
                # means error is in % and not sigma as required by detritalPy
                message_text = ''
                if settings.value('age_error_format_id', int) not in (1,2):
                    message_text += 'detritalPy uses absolute error for ages.\nWould you like to update the settings now?\n1% will be converted to 1sigma and 2% will be converted to 2sigma.'
                if settings.value('gps_format_id', int) == 7:  # UTM format
                    text = 'detritalPy uses Lat/Lon coordinates.\nWould you like to update the settings to DD +/- now?'
                    if message_text:
                        message_text += '\n\n' + text
                    else:
                        message_text += text
                if message_text:
                    response = QMessageBox.question(self, 'Update settings', message_text,
                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.Yes)
                    if response == QMessageBox.StandardButton.Yes:
                        if 'absolute error' in message_text:
                            if settings.value('age_error_format_id', int) == 3:  # 1 sigma %
                                settings.setValue('age_error_format_id', 1)
                                settings.setValue('age_error_format_abbreviation', '1σ abs')

                            elif settings.value('age_error_format_id', int) == 4:  # 2 sigma %
                                settings.setValue('age_error_format_id', 2)
                                settings.setValue('age_error_format_abbreviation', '2σ abs')
                        if 'Lat/Lon' in message_text:
                            if settings.value('gps_format_id', int) == 7:
                                settings.setValue('gps_format_id', 1)
                                settings.setValue('gps_format_abbreviation', 'DD +/-')

                        if not update_database():
                            logger_setup.get_logger().critical(f'Error updating and displaying database')
                            self.parent().close()
                            return
                        self.populate_stack()

                self.fileformat_comboBox.setCurrentText('Excel (.xlsx)')
                Samples_columns = {
                    ('Samples', 'SampleName'): True,
                    ('Units', 'UnitName'): True,
                    ('Regions', 'RegionName'): True,
                    ('GPSLocations', 'CalculatedLatDisplay'): True,
                    ('GPSLocations', 'CalculatedLonDisplay'): True,
                    ('References', 'ReferenceDisplay'): True
                }
                self.add_worksheet_tab('Samples', True, False, Samples_columns, [], True)
                self.previous_worksheet = 'Samples'

                ZrUPb_columns = {
                    ('Samples', 'SampleName'): True,
                    ('Grains', 'GrainName'): True,
                    ('Spots', 'SpotName'): True,

                    ('UPbAnalyses', "Calculated206Pb/204Pb"): True,
                    ('UPbAnalyses', "Uppm"): True,
                    ('UPbAnalyses', "CalculatedU/Th"): True,
                    ('UPbAnalyses', "CalculatedTh/U"): True,

                    ('UPbAnalyses', "Calculated207Pb/206Pb"): True,
                    ('UPbAnalyses', "Calculated207Pb/206PbError"): True,
                    ('UPbAnalyses', "Calculated207Pb/235U"): True,
                    ('UPbAnalyses', "Calculated207Pb/235UError"): True,
                    ('UPbAnalyses', "Calculated206Pb/238U"): True,
                    ('UPbAnalyses', "Calculated206Pb/238UError"): True,

                    ('UPbAnalyses', "ErrorCorr/Rho_68v76"): True,
                    ('UPbAnalyses', "ErrorCorr/Rho_68v75"): True,

                    ('UPbAnalyses', "Calculated207Pb/235UAge"): True,
                    ('UPbAnalyses', "Calculated207Pb/235UAgeError"): True,

                    ('UPbAnalyses', "Calculated206Pb/238UAge"): True,
                    ('UPbAnalyses', "Calculated206Pb/238UAgeError"): True,

                    ('UPbAnalyses', "Calculated207Pb/206PbAge"): True,
                    ('UPbAnalyses', "Calculated207Pb/206PbAgeError"): True,

                    ('UPbAnalyses', "CalculatedBestAgeFilled"): True,
                    ('UPbAnalyses', "CalculatedBestAgeErrorFilled"): True,

                    ('UPbAnalyses', "CalculatedConcordance_206Pb/238Uv207Pb/206Pb"): True,
                    ('UPbAnalyses', "CalculatedConcordance_206Pb/238Uv207Pb/235U"): True
                }

                self.column_name_mappings = {
                    "SampleName": "Sample_ID",
                    "UnitName": "Unit",
                    "RegionName": "Basin",
                    "CalculatedLatDisplay": "Latitude",
                    "CalculatedLonDisplay": "Longitude",
                    "ReferenceDisplay": "Source",

                    "GrainName": "Grain_ID",
                    "SpotName": "Analysis_ID",

                    "Calculated206Pb/204Pb": "206Pb_204Pb",
                    "Uppm": "U_ppm",
                    "CalculatedU/Th": "U_Th",
                    "CalculatedTh/U": "Th_U",

                    "Calculated207Pb/206Pb": "207Pb_206Pb",
                    "Calculated207Pb/206PbError": "207Pb_206Pb_err",
                    "Calculated207Pb/235U": "207Pb_235Pb",
                    "Calculated207Pb/235UError": "207Pb_235Pb_err",
                    "Calculated206Pb/238U": "206Pb_238Pb",
                    "Calculated206Pb/238UError": "206Pb_238Pb_err",

                    "ErrorCorr/Rho_68v76": "RHO_68v76",
                    "ErrorCorr/Rho_68v75": "RHO_68v76",

                    "Calculated207Pb/235UAge": "75Age",
                    "Calculated207Pb/235UAgeError": "75Age_err",

                    "Calculated206Pb/238UAge": "68Age",
                    "Calculated206Pb/238UAgeError": "68Age_err",

                    "Calculated207Pb/206PbAge": "76Age",
                    "Calculated207Pb/206PbAgeError": "76Age_err",

                    "CalculatedBestAgeFilled": "BestAge",
                    "CalculatedBestAgeErrorFilled": "BestAge_err",

                    "CalculatedConcordance_206Pb/238Uv207Pb/206Pb": "Conc_68v76",
                    "CalculatedConcordance_206Pb/238Uv207Pb/235U": "Conc_68v75",
                    "ConcordanceFormat"
                    "MinimumSegmentedDiscordance": "Min_Seg_Discordance"
                }
                self.add_worksheet_tab('ZrUPb', False, False, ZrUPb_columns, [], True)
            case 'IsoplotR - 07/35, 06/38, 04/38, 07/06, 04/07, 04/06':
                self.clear_worksheet_data()
                self.fileformat_comboBox.setEnabled(True)
                # modeled after UPb6.csv in IsoplotR
                # 207/235
                # 206/238
                # 204/238
                # 207/206
                # 204/207
                # 204/206
                self.fileformat_comboBox.setCurrentText('Comma-Separated Value (.csv)')
                UPb_columns = {
                    ('UPbAnalyses', 'Calculated207Pb/235U'): True,
                    ('UPbAnalyses', 'Calculated207Pb/235UError'): True,
                    ('UPbAnalyses', 'Calculated206Pb/238U'): True,
                    ('UPbAnalyses', 'Calculated206Pb/238UError'): True,
                    ('UPbAnalyses', 'Calculated204Pb/238U'): True,
                    ('UPbAnalyses', 'Calculated204Pb/238UError'): True,
                    ('UPbAnalyses', 'Calculated207Pb/206Pb'): True,
                    ('UPbAnalyses', 'Calculated207Pb/206PbError'): True,
                    ('UPbAnalyses', 'Calculated204Pb/207Pb'): True,
                    ('UPbAnalyses', 'Calculated204Pb/207PbError'): True,
                    ('UPbAnalyses', 'Calculated204Pb/206Pb'): True,
                    ('UPbAnalyses', 'Calculated204Pb/206PbError'): True,
                }
                self.column_name_mappings = {
                    'Calculated207Pb/235U': 'Calculated207Pb/235U',
                    'Calculated207Pb/235UError': 'Calculated207Pb/235UError',
                    'Calculated206Pb/238U': 'Calculated206Pb/238U',
                    'Calculated206Pb/238UError': 'Calculated206Pb/238UError',
                    'Calculated204Pb/238U': 'Calculated204Pb/238U',
                    'Calculated204Pb/238UError': 'Calculated204Pb/238UError',
                    'Calculated207Pb/206Pb': 'Calculated207Pb/206Pb',
                    'Calculated207Pb/206PbError': 'Calculated207Pb/206PbError',
                    'Calculated204Pb/207Pb': 'Calculated204Pb/207Pb',
                    'Calculated204Pb/207PbError': 'Calculated204Pb/207PbError',
                    'Calculated204Pb/206Pb': 'Calculated204Pb/206Pb',
                    'Calculated204Pb/206PbError': 'Calculated204Pb/206PbError'
                }
                self.add_worksheet_tab('IsoplotR', False, False, UPb_columns, [], True)

            case 'IsoplotR - 38/06, 07/06':
                self.clear_worksheet_data()
                self.fileformat_comboBox.setEnabled(True)
                # modeled after UPb2.csv in IsoplotR
                # 238/206
                # 207/206
                self.fileformat_comboBox.setCurrentText('Comma-Separated Value (.csv)')
                UPb_columns = {
                    ('UPbAnalyses', 'Calculated238U/206Pb'): True,
                    ('UPbAnalyses', 'Calculated238U/206PbError'): True,
                    ('UPbAnalyses', 'Calculated207Pb/206Pb'): True,
                    ('UPbAnalyses', 'Calculated207Pb/206PbError'): True
                }
                self.column_name_mappings = {
                    'Calculated238U/206Pb': 'Calculated238U/206Pb',
                    'Calculated238U/206PbError': 'Calculated238U/206PbError',
                    'Calculated207Pb/206Pb': 'Calculated207Pb/206Pb',
                    'Calculated207Pb/206PbError': 'Calculated207Pb/206PbError',
                }
                self.add_worksheet_tab('IsoplotR', False, False, UPb_columns, [], True)

            case 'DZstats':
                self.clear_worksheet_data()
                self.fileformat_comboBox.setEnabled(True)
                # means error is in % and not sigma as required by detritalPy
                if settings.value('age_error_format_id', int) not in (1, 2):
                    response = QMessageBox.question(self, 'Update settings',
                                                    'DZstats uses absolute error for ages.\nWould you like to update the settings now?\n1% will be converted to 1sigma and 2% will be converted to 2sigma.',
                                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                                    QMessageBox.StandardButton.Yes)
                    if response == QMessageBox.StandardButton.Yes:
                        if settings.value('age_error_format_id', int) == 3:  # 1 sigma %
                            settings.setValue('age_error_format_id', 1)
                            settings.setValue('age_error_format_abbreviation', '1σ abs')

                        elif settings.value('age_error_format_id', int) == 4:  # 2 sigma %
                            settings.setValue('age_error_format_id', 2)
                            settings.setValue('age_error_format_abbreviation', '2σ abs')

                        if not update_database():
                            logger_setup.get_logger().critical(f'Error updating and displaying database')
                            self.parent().close()
                            return
                        self.populate_stack()

                self.fileformat_comboBox.setCurrentText('Comma-Separated Value (.csv)')
                UPb_columns = {
                    ('Samples', 'SampleName'): True,
                    ('UPbAnalyses', 'CalculatedBestAgeFilled'): True,
                    ('UPbAnalyses', 'CalculatedBestAgeErrorFilled'): True
                }
                self.column_name_mappings = {
                    'SampleName': 'SampleName',
                    'CalculatedBestAgeFilled': 'CalculatedBestAgeFilled',
                    'CalculatedBestAgeErrorFilled': 'CalculatedBestAgeErrorFilled'
                }
                self.add_worksheet_tab('DZStats', False, True, UPb_columns, [], False)
            case 'DZmix, DZmds, DZnmf':
                self.clear_worksheet_data()
                self.fileformat_comboBox.setEnabled(True)
                # means error is in % and not sigma as required by detritalPy
                if settings.value('age_error_format_id', int) not in (1, 2):
                    response = QMessageBox.question(self, 'Update settings',
                                                    'DZmix, DZmds, DZnmf uses absolute error for ages.\nWould you like to update the settings now?\n1% will be converted to 1sigma and 2% will be converted to 2sigma.',
                                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                                    QMessageBox.StandardButton.Yes)
                    if response == QMessageBox.StandardButton.Yes:
                        if settings.value('age_error_format_id', int) == 3:  # 1 sigma %
                            settings.setValue('age_error_format_id', 1)
                            settings.setValue('age_error_format_abbreviation', '1σ abs')

                        elif settings.value('age_error_format_id', int) == 4:  # 2 sigma %
                            settings.setValue('age_error_format_id', 2)
                            settings.setValue('age_error_format_abbreviation', '2σ abs')

                        if not update_database():
                            logger_setup.get_logger().critical(f'Error updating and displaying database')
                            self.parent().close()
                            return
                        self.populate_stack()

                self.fileformat_comboBox.setCurrentText('Excel (.xlsx)')
                UPb_columns = {
                    ('Samples', 'SampleName'): True,
                    ('UPbAnalyses', 'CalculatedBestAgeFilled'): True,
                    ('UPbAnalyses', 'CalculatedBestAgeErrorFilled'): True
                }
                self.column_name_mappings = {
                    'SampleName': 'SampleName',
                    'CalculatedBestAgeFilled': 'CalculatedBestAgeFilled',
                    'CalculatedBestAgeErrorFilled': 'CalculatedBestAgeErrorFilled'
                }
                self.add_worksheet_tab('DZmix, DZmds, DZnmf', False, True, UPb_columns, [], True)
            case 'AgeCalcML concordia':
                self.clear_worksheet_data()
                self.fileformat_comboBox.setEnabled(True)
                self.fileformat_comboBox.setCurrentText('Comma-Separated Value (.csv)')
                UPb_columns = {
                    ('Samples', 'SampleName'): True,
                    ('UPbAnalyses', 'Calculated207Pb/235U'): True,
                    ('UPbAnalyses', 'Calculated207Pb/235UError'): True,
                    ('UPbAnalyses', 'Calculated206Pb/238U'): True,
                    ('UPbAnalyses', 'Calculated206Pb/238UError'): True,
                    ('UPbAnalyses', 'ErrorCorr/Rho_68v75'): True
                }
                self.column_name_mappings = {
                    'SampleName': 'SampleName',
                    'Calculated207Pb/235U': 'Calculated207Pb/235U',
                    'Calculated207Pb/235UError': 'Calculated207Pb/235UError',
                    'Calculated206Pb/238U': 'Calculated206Pb/238U',
                    'Calculated206Pb/238UError': 'Calculated206Pb/238UError',
                    'ErrorCorr/Rho_68v75': 'ErrorCorr/Rho_68v75'
                }
                self.add_worksheet_tab('AgeCalcML concordia', False, True, UPb_columns, [], True)
            case 'Database':
                if self.findChild(QSqlTableModel, 'database_QSqlTableModel') is not None:
                    self.findChild(QSqlTableModel, 'database_QSqlTableModel').clear()
                    self.findChild(QSqlTableModel, 'database_QSqlTableModel').setParent(None)
                    QSqlDatabase.removeDatabase('temp')
                    self.findChild(QWidget, 'database_tab').setParent(None)

                # self.update_database_export()

                """The following code creates a temporary database file to export the selected data to.
                It is time consuming for large databases, so use DataViewerWidget instead."""
                # if os.path.isfile("temp.db"):
                #     if 'temp' in QSqlDatabase().connectionNames():
                #         QSqlDatabase.database('temp').close()
                #         QSqlDatabase().removeDatabase('temp')
                #         os.remove("temp.db")
                # tgt_db_file = "temp.db"
                #
                # tgt_db = QSqlDatabase().addDatabase('QSQLITE', 'temp')
                # tgt_db.setDatabaseName(tgt_db_file)
                # tgt_db.open()
                # if not tgt_db.isOpen():
                #     logger_setup.get_logger().critical('Could not open target database')
                #     return
                # if not turn_on_foreign_keys():
                #     return
                #
                # src_db = QSqlDatabase()
                #
                # show_loading_dialog('Loading', 'Loading selected data...')
                # ExportDatabase.subset_database(src_db, tgt_db, sample_ids_to_subset)
                # close_loading_dialog('Loading', 'Loading selected data...')
                # tgt_db.commit()
                # tgt_db.close()
                #
                # # Create a new tab
                # new_tab = DisplayTablesSimplified(self, tgt_db_file)
                # tab_layout = QVBoxLayout(self)
                # new_tab.setLayout(tab_layout)


                self.worksheet_tabs_dict['Database'] = {
                    'tableView': QTableView(),
                    'model': QSqlQueryModel(),
                    'distinct': None,
                    'pivot': False,
                    'selected_columns': {},
                    'ordered_columns': [],
                    'label': '',
                    'headers': None,
                    'sql': ''
                }
            case 'Custom':
                self.clear_worksheet_data()
                self.create_first_worksheet_tab()
        self.update_table_view()

    def update_database_export(self):
        self.clear_worksheet_data()

        self.fileformat_comboBox.setEnabled(False)
        # self.selectionscope_comboBox.setCurrentText('Samples')
        # self.selectionscope_comboBox.setEnabled(False)
        self.columnattributes_stack.setEnabled(False)
        self.columnselection_comboBox.setEnabled(False)
        self.editorder_button.setEnabled(False)
        self.add_worksheet_button.setEnabled(False)
        self.remove_worksheet_button.setEnabled(False)
        self.edit_columnnames_button.setEnabled(False)
        # self.filterselection_comboBox.hide()
        self.groupedfilter_comboBox.hide()
        self.groupedfilter_label.hide()
        # self.filters_label.hide()

        if len(self.filtered_upb_ids) == 0:
            sample_ids_to_subset = []
        else:
            if len(self.filtered_upb_ids) == 1:
                where = f"= {list(self.filtered_upb_ids)[0]}"
            else:
                where = f"IN ({', '.join(self.filtered_upb_ids)})"
            sample_query = f"""SELECT DISTINCT SampleID FROM UPbAnalyses
                                {SQLUtils.upb_spot_join}
                                {SQLUtils.spot_aliquot_join}
                                WHERE UPbAnalyses.UPbAnalysisID {where}"""
            sample_ids_to_subset = columns_as_list(sample_query, [0])[0]

        all_samples = self.checked_sample_list + sample_ids_to_subset
        from ui.DataViewerWidget import DataViewerWidget
        new_tab = DataViewerWidget(self, set(all_samples), 'Samples')
        tab_layout = QVBoxLayout(self)
        new_tab.setLayout(tab_layout)
        self.workbooktabs.addTab(new_tab, 'Database')

    def clear_worksheet_data(self):
        self.delete_all_worksheet_tabs()
        self.selectionscope_comboBox.setEnabled(True)
        self.columnattributes_stack.setEnabled(True)
        self.columnselection_comboBox.setEnabled(True)
        self.editorder_button.setEnabled(True)
        self.add_worksheet_button.setEnabled(True)
        self.remove_worksheet_button.setEnabled(True)
        self.edit_columnnames_button.setEnabled(True)
        self.fileformat_comboBox.setEnabled(True)
        self.filterselection_comboBox.show()
        self.groupedfilter_comboBox.show()
        self.groupedfilter_label.show()
        self.filters_label.show()
        self.column_name_mappings.clear()

    def update_step_2_list(self):
        """Updates the CheckableComboBox model based upon selected values. Allows the user to select
         samples by either Samples or FilterGroups """
        logger_setup.get_logger().info(f'Updating Step 2 list for {self.selectionscope_comboBox.currentText()}')
        start_update_step_2_time = time.time()
        self.samplesincluded_comboBox.setEnabled(True)
        self.samplesincluded_comboBox.show()
        self.step_2_label.show()
        self.filters_label.show()
        self.filters_label.setText("Select Additional Filters (optional):")
        self.filters_label.setToolTip(
            "Additional filters to filter the samples, multiple filters are combined with OR.")
        self.samplesincluded_comboBox.clear_all_checks()
        self.filterselection_comboBox.clear_all_checks()
        self.groupedfilter_comboBox.clear_all_checks()

        if self.selectionscope_comboBox.currentText() == 'Filter Groups':
            self.step_2_label.hide()
            self.samplesincluded_comboBox.hide()
            self.samplesincluded_comboBox: CheckableComboBox
            self.checked_sample_list = []

            self.filters_label.setText("Select Filters:")
            self.filters_label.setToolTip("")

            # self.update_table_view()
        logger_setup.get_logger().info(f'Update Step 2 list took {time.time() - start_update_step_2_time:.2f} seconds')
        self.update_table_view()

    def tab_changed(self):
        """Method to update the table view when a tab/worksheet is changed or switched by the user. QTabWidgets
        emitting currentChanged signal occur AFTER the tab is changed, therefore the current index is the new tab.
        Therefore, saving the tab that was previously selected is necessary to save the checkbox states."""
        if self.workbooktabs.tabText(self.workbooktabs.currentIndex()) == 'Database':
            return
        self.save_checkbox_states(self.previous_worksheet)
        self.load_checkbox_states()
        # variable to hold the name of the previous worksheet
        self.previous_worksheet = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        self.update_table_view()

    def add_worksheet_tab(self, worksheet_name: str = None, distinct: bool = False, pivot: bool = False,
                          selected_columns: dict = None, ordered_columns: list = None, headers: bool = False):
        """
        Method to add a new worksheet tab to the workbook. This method is called when the user clicks the "Add Worksheet" button.
        :param str worksheet_name: name of the worksheet to be created
        :param bool distinct: utilize distinct rows in the SQL query
        :param bool pivot: pivot the table based on the first column
        :param dict selected_columns: dictionary of all selected columns, table then field name
        :param dict ordered_columns: list of all ordered columns, table.field name
        :param bool headers: include headers in the output files
        :return:
        """
        if ordered_columns is None:
            ordered_columns = []
        if selected_columns is None:
            selected_columns = {}

        if worksheet_name is None:
            worksheet_name, ok = QInputDialog.getText(self, "New Worksheet", "Enter worksheet name:")
            if not ok or not worksheet_name:
                return  # User canceled or didn't enter a name

            if worksheet_name in self.worksheet_tabs_dict:
                QMessageBox.warning(self, "Duplicate Name", "A worksheet with that name already exists.")
                return

            if self.exportformat_comboBox.currentText() != 'Custom':
                self.exportformat_comboBox.currentIndexChanged.disconnect()
                self.exportformat_comboBox.setCurrentText('Custom')
                self.exportformat_comboBox.currentIndexChanged.connect(self.export_format)
                self.save_checkbox_states()

        # Save previous sheet information
        self.previous_worksheet = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        if self.previous_worksheet:
            self.save_checkbox_states(self.previous_worksheet)

        # Create a new tableView
        new_tableView = QTableView()
        new_tableView.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        # Create a new data model for the new tableView
        model = QSqlQueryModel()

        # Create a new tab
        new_tab = QWidget()
        tab_layout = QVBoxLayout()
        tab_layout.setContentsMargins(0, 0, 0, 0)
        tab_layout.setSpacing(0)

        new_tab.setLayout(tab_layout)

        horizontal_layout = QHBoxLayout()
        horizontal_layout.setContentsMargins(0, 0, 0, 0)
        horizontal_layout.setSpacing(0)

        distinct_checkbox = QCheckBox("Distinct Rows")
        distinct_checkbox.setToolTip("Check this box to only show distinct or unique rows a single time")
        distinct_checkbox.setChecked(distinct)
        horizontal_layout.addWidget(distinct_checkbox)

        headers_checkbox = QCheckBox("Include Headers")
        headers_checkbox.setToolTip("Check this box include headers in output files")
        headers_checkbox.setChecked(headers)
        headers_checkbox.setFixedSize(150, 20)
        horizontal_layout.addWidget(headers_checkbox)

        pivot_checkbox = QCheckBox("Pivot Table")
        pivot_checkbox.setToolTip("Check this box to pivot the table based on first column")
        pivot_checkbox.setChecked(pivot)
        pivot_checkbox.setFixedSize(150, 20)
        horizontal_layout.addWidget(pivot_checkbox)

        counter_label = QLabel("Number of Rows: ")
        counter_label.setFixedSize(200, 20)
        horizontal_layout.addWidget(counter_label)

        tab_layout.addLayout(horizontal_layout)
        tab_layout.addWidget(new_tableView)

        # Store the tableView and model in the worksheet_tabs_dict
        self.worksheet_tabs_dict[worksheet_name] = {
            'tableView': new_tableView,
            'model': model,
            'distinct': distinct,
            'pivot': pivot,
            'selected_columns': selected_columns,
            'ordered_columns': ordered_columns,
            'label': counter_label,
            'headers': headers,
            'sql': ''
        }
        self.workbooktabs.blockSignals(True)
        self.workbooktabs.addTab(new_tab, worksheet_name)
        self.workbooktabs.blockSignals(False)

        self.load_checkbox_states(worksheet_name)

        # this blocks the tab_changed function and signal from emitting
        self.workbooktabs.blockSignals(True)
        self.workbooktabs.setCurrentWidget(new_tab)
        self.workbooktabs.blockSignals(False)

        self.previous_worksheet = self.workbooktabs.tabText(self.workbooktabs.currentIndex())

        distinct_checkbox.stateChanged.connect(self.update_distinct_checkbox)
        headers_checkbox.stateChanged.connect(self.update_header_checkbox)
        pivot_checkbox.stateChanged.connect(self.update_pivottable_checkbox)

    def create_first_worksheet_tab(self):
        """Creates the initial worksheet tab when the ExportWidget is first created. This is the default tab that is shown"""
        tab1 = QWidget()
        tab1_layout = QVBoxLayout()
        tab1_layout.setContentsMargins(0, 0, 0, 0)
        tab1_layout.setSpacing(0)
        tab1.setLayout(tab1_layout)
        tableView = QTableView()

        horizontal_layout = QHBoxLayout()
        horizontal_layout.setContentsMargins(0, 0, 0, 0)
        horizontal_layout.setSpacing(0)

        # Adds a distinct clause to the SQL query
        distinct_checkbox = QCheckBox("Distinct Rows")
        distinct_checkbox.setToolTip("Check this box to only show distinct or unique rows a single time")
        distinct_checkbox.setChecked(False)
        distinct_checkbox.setFixedSize(150, 20)
        horizontal_layout.addWidget(distinct_checkbox)

        # Adds headers to the exported Excel or CSV files when checked
        headers_checkbox = QCheckBox("Include Headers")
        headers_checkbox.setToolTip("Check this box include headers in output files")
        headers_checkbox.setChecked(True)
        headers_checkbox.setFixedSize(150, 20)
        horizontal_layout.addWidget(headers_checkbox)

        # Pivots the SQL query based on the first column when checked, SQLite3 does not support pivoting and a custom
        # TempPivotTable is used.
        pivot_checkbox = QCheckBox("Pivot Table")
        pivot_checkbox.setToolTip("Check this box to pivot the table based on first column")
        pivot_checkbox.setChecked(False)
        pivot_checkbox.setFixedSize(150, 20)
        horizontal_layout.addWidget(pivot_checkbox)

        # label to hold the number of rows returned from the SQL query, the tableview only shows the first {self.max_rows_to_display} rows
        # so {self.max_rows_to_display}/#### is common
        counter_label = QLabel("Number of Rows: ")
        counter_label.setFixedSize(200, 20)
        horizontal_layout.addWidget(counter_label)

        tab1_layout.addLayout(horizontal_layout)
        tab1_layout.addWidget(tableView)

        # Create a data model for this tableView
        model = QSqlQueryModel()

        self.worksheet_tabs_dict["Worksheet 1"] = {
            'tableView': tableView,
            'model': model,
            'distinct': False,
            'pivot': False,
            'selected_columns': {},
            'ordered_columns': [],
            'label': counter_label,
            'headers': True,
            'sql': ''
        }

        self.workbooktabs.blockSignals(True)
        self.workbooktabs.addTab(tab1, "Worksheet 1")
        self.workbooktabs.blockSignals(False)

        self.load_checkbox_states('Worksheet 1')

        distinct_checkbox.stateChanged.connect(self.update_distinct_checkbox)
        headers_checkbox.stateChanged.connect(self.update_header_checkbox)
        pivot_checkbox.stateChanged.connect(self.update_pivottable_checkbox)
        # self.update_table_view()

    def delete_all_worksheet_tabs(self):
        """Delete all worksheet tabs and their associated data. This is used when the ExportWidget has a change
        in selected export format."""
        self.workbooktabs.setParent(None)
        self.workbook_layout.removeWidget(self.workbooktabs)
        self.workbooktabs.deleteLater()

        self.workbooktabs = QTabWidget()

        self.workbooktabs.currentChanged.connect(self.tab_changed)
        self.workbooktabs.tabBarDoubleClicked.connect(self.rename_worksheet_tab)
        self.previous_worksheet = self.workbooktabs.tabText(self.workbooktabs.currentIndex())

        self.workbook_layout.addWidget(self.workbooktabs)

        self.worksheet_tabs_dict = {}
        self.previous_worksheet = None

    def remove_current_worksheet_tab(self):
        """Method to remove the current worksheet tab from the dictionary. Also removes the tab from the tabWidget."""

        if self.workbooktabs.count() <= 1:
            QMessageBox.warning(self, "Cannot Remove Worksheet", "At least one worksheet must remain.")
            return

        # Get the current workbook name
        current_index = self.workbooktabs.currentIndex()
        current_worksheet_name = self.workbooktabs.tabText(current_index)

        reply = QMessageBox.question(self, 'Remove Worksheet',
                                     f"Are you sure you want to remove '{current_worksheet_name}'?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return

        # Remove the tab from the tabWidget
        self.workbooktabs.removeTab(current_index)

        # Remove the workbook from the dictionary
        del self.worksheet_tabs_dict[current_worksheet_name]

        if self.exportformat_comboBox.currentText() != 'Custom':
            self.exportformat_comboBox.currentIndexChanged.disconnect()
            self.exportformat_comboBox.setCurrentText('Custom')
            self.exportformat_comboBox.currentIndexChanged.connect(self.export_format)
            self.save_checkbox_states()

    def populate_stack(self):
        """
        Method to populate the stacked widget with the column attributes for each table. This is used to show the
        available columns to export. Each page stack widget has checkboxes for all attributes in a single table.
        """

        # Clear the existing widgets in the columnattributes_stack and all widgets inside
        while self.columnattributes_stack.count():
            widget = self.columnattributes_stack.widget(0)
            self.columnattributes_stack.removeWidget(widget)
            widget.deleteLater()

        # loop over all tables and their attributes.
        for table_name, field_items in SQLUtils.table_attributes_dict.items():
            if table_name == "GPSLocations":
                # sets the GPSLocations table to have a different set of fields based on user-selection
                if SettingsManager().settings.value('gps_format_id', 1) == 7:  # UTM Selected
                    field_items = ['GPSLocationConverted', 'GPSLocationDisplay', 'CalculatedZoneDisplay',
                                   'CalculatedEastingDisplay', 'CalculatedNorthingDisplay', 'CalculatedGPSElev',
                                   'CalculatedGPSElevError']
                else:
                    field_items = ['GPSLocationConverted', 'GPSLocationDisplay', 'CalculatedLatDisplay', 'CalculatedLonDisplay',
                                   'CalculatedGPSElev', 'CalculatedGPSElevError']

            # Create container widget with QVBoxLayout
            container_widget = QWidget()
            container_layout = QVBoxLayout(container_widget)
            container_layout.setContentsMargins(0, 0, 0, 0)
            container_layout.setSpacing(0)

            # Create a scroll area that expands vertically within vertical_layout9
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)  # Ensures the content resizes dynamically
            scroll_area.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

            # Create a widget to hold the FlowLayout
            table_widget = QWidget()
            flow_layout = FlowLayout()
            flow_layout.setSpacing(0)
            flow_layout.setContentsMargins(0, 0, 0, 0)

            # loop over all the fields/attributes for this given table and create checkboxes for them
            for field in field_items:
                checkbox = QCheckBox(field)

                # Prevent expanding
                checkbox.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)

                # Group label and checkbox
                group_widget = QGroupBox()
                hbox_layout = QHBoxLayout()
                hbox_layout.setContentsMargins(0, 0, 0, 0)
                hbox_layout.setSpacing(8)
                hbox_layout.addWidget(checkbox)
                group_widget.setLayout(hbox_layout)

                flow_layout.addWidget(group_widget)

                # Save field metadata
                checkbox.setProperty("field_name", field)
                checkbox.setProperty("table_name", table_name)
                checkbox.checkStateChanged.connect(self.field_check_state_changed)

            # Add a vertical spacer to push content up
            vertical_spacer = QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding)
            flow_layout.addItem(vertical_spacer)

            # Set the layout to the table_widget and add it to the scroll area
            table_widget.setLayout(flow_layout)
            scroll_area.setWidget(table_widget)  # Attach scroll area to table_widget

            # Add the scroll area to the container layout
            container_layout.addWidget(scroll_area)

            # Add container widget to the main layout stack
            self.columnattributes_stack.addWidget(container_widget)
        self.columnselection_comboBox.blockSignals(True)
        self.columnselection_comboBox.setCurrentIndex(0)
        self.columnselection_comboBox.blockSignals(False)

        self.settings_label.setText(
            f"""GPS Format: {self.settings.value('gps_format_abbreviation')}
Elevation unit: {self.settings.value('elevation_unit_abbreviation')}
Column height/depth unit: {self.settings.value('heightdepth_unit_abbreviation')}
Spot size unit: {self.settings.value('spotsize_unit_abbreviation')}
Age unit: {self.settings.value('age_unit_abbreviation')}
Age error format: {self.settings.value('age_error_format_abbreviation')}
U-Pb ratio error format: {self.settings.value('ratio_error_format_abbreviation')}
U-Pb concordance format: {self.settings.value('concordance_format_abbreviation')}""")

    def switch_table_layout(self):
        """Method to switch the stack widget to show the layout corresponding to the selected table"""

        selected_table_index = self.columnselection_comboBox.currentIndex()
        self.columnattributes_stack.setCurrentIndex(selected_table_index)

        # Save and load checkbox states for each table
        self.save_checkbox_states()
        self.load_checkbox_states()

        self.update_table_view()

    def field_check_state_changed(self, previous_worksheet=None):
        """
        Called when a field is checked or unchecked.
        :param previous_worksheet: The name of the previous worksheet, if any. If None, use the current worksheet.
        """
        if self.exportformat_comboBox.currentText() != 'Custom':
            self.exportformat_comboBox.currentIndexChanged.disconnect()
            self.exportformat_comboBox.setCurrentText('Custom')
            self.exportformat_comboBox.currentIndexChanged.connect(self.export_format)
        self.save_checkbox_states()
        self.load_checkbox_states()
        self.update_table_view()

    def save_checkbox_states(self, previous_worksheet=None):
        """
        Method to loop over all checkboxes in the columnattributes_stack and save their states to the current
        worksheet in the workbook_tabs_dict. This is used to save the state of the checkboxes when switching between
        multiple sheets. When switching sheets, the signal is emitted AFTER changing so the current tab index in the
        tab widget is the new tab.
        :param previous_worksheet: The name of the previous worksheet, if any. If None, use the current worksheet.
        """
        if self.exportformat_comboBox.currentText() != 'Custom':
            return
        if self.workbooktabs.currentIndex() == -1:
            return  # No tabs available, skip saving states
        current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        checkbox_states = self.get_selected_values()

        # Store checkbox_states in the workbook's data
        if previous_worksheet is None:
            self.worksheet_tabs_dict[current_worksheet_name]['selected_columns'] = checkbox_states
        else:
            self.worksheet_tabs_dict[previous_worksheet]['selected_columns'] = checkbox_states

    def load_checkbox_states(self, worksheet_name=None):
        """
        Method to loop over all checkboxes in the columnattributes_stack and set their states based on
         the workbook_tabs_dict. This is used to load the state of the checkboxes when switching between
         multiple sheets. When switching sheets, the signal is emitted AFTER changing so the current tab index in the
         tab widget is the new tab. Therefore, no need to save the previous worksheet name.
        :param worksheet_name: The name of the worksheet to load checkbox states from. If None, use the current worksheet.
        """
        if self.workbooktabs.currentIndex() == -1:
            return  # No tabs available, skip saving states
        current_worksheet_name = worksheet_name if worksheet_name else self.workbooktabs.tabText(
            self.workbooktabs.currentIndex())

        checkbox_states = self.worksheet_tabs_dict[current_worksheet_name].get('selected_columns', {})

        # loop over all tables in the columnattributes_stack
        for index in range(self.columnattributes_stack.count()):
            # get the table_widget which contains the checkboxes
            table_widget = self.columnattributes_stack.widget(index)

            # if table_widget exists, loop over all checkboxes in the table_widget
            if table_widget:
                for widget in table_widget.findChildren(QCheckBox,
                                                        options=QtCore.Qt.FindChildOption.FindChildrenRecursively):
                    if widget is None:
                        continue
                    if isinstance(widget, QCheckBox):
                        field_name = widget.property('field_name')
                        table_name = widget.property('table_name')
                        # get the current checkbox state from the checkbox_states dictionary, default to False
                        checked = checkbox_states.get((table_name, field_name), False)

                        widget.blockSignals(True)  # Prevent signals during state change
                        widget.setChecked(checked)
                        widget.blockSignals(False)

    def get_selected_values(self):
        """Method to get all selected values from the columnattributes_stack. """

        # todo: this could be probably better to change the checkboxes in the stack to emit a custom signal when their
        #  state changes and dynamically update the dictionary, rather than constantly looping over all widgets
        #  in the stack.

        selected_columns = {}

        # loop over all tables in the columnattributes_stack
        for index in range(self.columnattributes_stack.count()):
            # get the table_widget which contains the checkboxes
            table_widget = self.columnattributes_stack.widget(index)
            table_name = self.columnselection_comboBox.itemText(index)

            # if table_widget exists, loop over all checkboxes in the table_widget
            if table_widget:
                for widget in table_widget.findChildren(QCheckBox,
                                                        options=QtCore.Qt.FindChildOption.FindChildrenRecursively):
                    if isinstance(widget, QCheckBox) and widget.isChecked():
                        field_name = widget.property('field_name')
                        table_name = widget.property('table_name')

                        # set the selected columns dict[tuple(), bool] to true
                        selected_columns[(table_name, field_name)] = True

        # Store selected_columns in the current workbook
        current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        self.worksheet_tabs_dict[current_worksheet_name]['selected_columns'] = selected_columns

        # update the column_name_mappings to ensure the values are stored
        for column in selected_columns:
            if column[1] not in self.column_name_mappings:
                self.column_name_mappings[column[1]] = column[1]

        return selected_columns

    def select_checkboxes(self, values: tuple[str, str]):
        """
        Method to select a given checkbox based on a table_name and field_name. Currently not used.
        :param tuple[str, str] values: tuple containing the table_name and field_name
        """
        # Values should be tuple format ('table_name', 'field_name')
        for index in range(self.columnattributes_stack.count()):
            table_widget = self.columnattributes_stack.widget(index)
            if table_widget:
                for widget in table_widget.findChildren(QCheckBox,
                                                        options=QtCore.Qt.FindChildOption.FindChildrenRecursively):
                    if isinstance(widget, QCheckBox):
                        table_name = widget.property('table_name')
                        field_name = widget.property('field_name')

                        if (table_name, field_name) in values:
                            widget.setChecked(True)
                        else:
                            widget.setChecked(False)
        self.update_table_view()

    def update_distinct_checkbox(self):
        """
        Helper method to update the worksheet tabs dictionary when the checkbox state is changed.
        """

        current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        distinct_checkbox = self.worksheet_tabs_dict[current_worksheet_name]['distinct']
        self.worksheet_tabs_dict[current_worksheet_name]['distinct'] = not distinct_checkbox
        self.update_table_view()

    def update_header_checkbox(self):
        """
        Helper method to update the worksheet tabs dictionary when the checkbox state is changed.
        """

        current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        headers_checkbox = self.worksheet_tabs_dict[current_worksheet_name]['headers']
        self.worksheet_tabs_dict[current_worksheet_name]['headers'] = not headers_checkbox

    def uncheck_pivot(self):
        """
        Find and uncheck the pivot checkbox
        :return:
        """
        for child in self.findChildren(QCheckBox):
            if isinstance(child, QCheckBox) and child.text() == "Pivot Table":
                child.setChecked(False)
                current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
                self.worksheet_tabs_dict[current_worksheet_name]['pivot'] = False
                break

    def update_pivottable_checkbox(self):
        """
        Helper method to update the worksheet tabs dictionary when the checkbox state is changed.
        """
        pivot_checkbox = self.sender()
        current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        if not pivot_checkbox or pivot_checkbox.checkState() == Qt.CheckState.Unchecked:
            self.worksheet_tabs_dict[current_worksheet_name]['pivot'] = False
        else:
            self.worksheet_tabs_dict[current_worksheet_name]['pivot'] = True
        self.update_table_view()

    def refresh_widget(self):
        """
        Method to force a refresh of the dropdowns, stacked checkboxes, and table view.
        """
        logger_setup.get_logger().info('Refresh Button Clicked')

        if self.exportformat_comboBox.currentText() != 'Database':
            # Recreate column stacks and reload checkbox states
            for tab_index in range(self.workbooktabs.count()):
                sheet_name = self.workbooktabs.tabText(tab_index)
                self.save_checkbox_states(sheet_name)
            self.populate_stack()
            for tab_index in range(self.workbooktabs.count()):
                sheet_name = self.workbooktabs.tabText(tab_index)
                self.load_checkbox_states(sheet_name)
            for tab_index in range(self.workbooktabs.count()):
                sheet_name = self.workbooktabs.tabText(tab_index)
                self.save_checkbox_states(sheet_name)

        self.showEvent(None)

        # Recheck items
        self.samplesincluded_comboBox.source_model().update_model_checks(set(self.checked_sample_list), set())
        self.filterselection_comboBox.source_model().update_model_checks(set(self.checked_filter_list), set())
        self.groupedfilter_comboBox.source_model().update_model_checks(set(self.checked_grouped_filter_list), set())

        self.update_table_view()

    def showEvent(self, a0):
        """Overridden showEvent to repopulate the table models when the widget is shown. This occurs mainly when
        the tabs are switched so if samples, filters are modified, the models are updated."""
        logger_setup.get_logger().info('Populating ExportWidget with data from the database')
        start_show_time = time.time()
        show_loading_dialog('Loading', 'Loading data for export...')

        sample_count = get_total_records('Samples')
        if self.sample_count != sample_count:
            self.sample_count = sample_count
            logger_setup.get_logger().info(f'Found {self.sample_count} samples in the database')
            show_loading_dialog('Loading', f'Loading {self.sample_count} Samples...')
            self.samplesincluded_comboBox.enable_context_menu(show_context_menu=True, only_select_deselect=True)
            view_query = ViewQuery('Samples', True, **{'show_columns': settings.value('sample_edit_columns')[0:4]})
            populate_combo_box(self.samplesincluded_comboBox, **{'table': 'Samples', 'live': False,
                                                                 'query': view_query.table_query, 'view_query': view_query})
            self.samplesincluded_comboBox.source_model().update_model_checks(set(self.checked_sample_list), set())
            close_loading_dialog('Loading', f'Loading {self.sample_count} Samples...')

        filter_count = get_total_records('FilterGroups')
        logger_setup.get_logger().info(f'Found {filter_count} filters in the database')

        show_loading_dialog('Loading', f'Loading {filter_count} Filter Groups...')
        self.filterselection_comboBox.enable_context_menu(show_context_menu=True, only_select_deselect=True)
        populate_combo_box(self.filterselection_comboBox, **{'table': 'FilterGroups', 'live': False})
        self.filterselection_comboBox.source_model().update_model_checks(set(self.checked_filter_list), set())
        self.groupedfilter_comboBox.enable_context_menu(show_context_menu=True, only_select_deselect=True)
        populate_combo_box(self.groupedfilter_comboBox, **{'table': 'FilterGroups', 'live': False})
        self.groupedfilter_comboBox.source_model().update_model_checks(set(self.checked_grouped_filter_list), set())
        close_loading_dialog('Loading', f'Loading {filter_count} Filter Groups...')

        if self.exportformat_comboBox.currentText() != 'Custom' or len(self.worksheet_tabs_dict.keys()) == 0:
            # Only update the format if the format is not custom or if the format is custom but no sheet has been created yet
            self.export_format()
        # self.update_step_2_list()
        try:
            self.exportformat_comboBox.currentIndexChanged.disconnect()
        except TypeError: pass

        self.exportformat_comboBox.currentIndexChanged.connect(self.export_format)
        # self.exportformat_comboBox.currentIndexChanged.connect(self.update_table_view)

        try:
            self.selectionscope_comboBox.currentIndexChanged.disconnect()
        except TypeError: pass
        self.selectionscope_comboBox.currentIndexChanged.connect(self.update_step_2_list)

        try:
            self.samplesincluded_comboBox.closing.disconnect()
        except TypeError:
            pass
        try:
            self.filterselection_comboBox.closing.disconnect()
        except TypeError:
            pass
        try:
            self.groupedfilter_comboBox.closing.disconnect()
        except TypeError:
            pass
        try:
            self.columnselection_comboBox.currentIndexChanged.disconnect()
        except TypeError:
            pass

        self.samplesincluded_comboBox.closing.connect(self.update_table_view)
        self.filterselection_comboBox.closing.connect(self.update_table_view)
        self.groupedfilter_comboBox.closing.connect(self.update_table_view)
        self.columnselection_comboBox.currentIndexChanged.connect(self.switch_table_layout)
        self.workbooktabs.currentChanged.connect(self.tab_changed)
        self.workbooktabs.tabBarDoubleClicked.connect(self.rename_worksheet_tab)

        self.samplesincluded_comboBox.clearEditText()

        # self.update_table_view()

        close_loading_dialog('Loading', 'Loading data for export...')
        logger_setup.get_logger().info(f'ExportWidget populated in {time.time() - start_show_time:.2f} seconds')
        super().showEvent(a0)

    def export_data(self):
        """Method to export the generated tableView and SQL code to a given format. Based on the exportformat_comboBox's
         current index. This method is called when the export_puhsbutton is clicked."""
        show_loading_dialog('Export', 'Exporting data...')
        if self.exportformat_comboBox.currentText() == 'Database':
            self.export_to_database()
        else:
            if self.fileformat_comboBox.currentText() == 'Excel (.xlsx)':
                self.export_to_excel()
            elif self.fileformat_comboBox.currentText() == 'Comma-Separated Value (.csv)':
                self.export_to_csv()
        close_loading_dialog('Export', 'Exporting data...')


    def export_to_database(self):
        """Exports the selected samples, either by selection or filter, to a new database file containing all related
        data to the samples, including but not limited to, Samples, Aliquots, Spots, UPbAnalyses, RockTypes,
        Units, and References... This will only include data that is absolutely related, no unnecessary data will be
        included."""
        fileName, _ = QFileDialog.getSaveFileName(
            None,
            "Save Database File",
            "",
            "Database Files (*.db)"
        )

        if not fileName:
            return

        logger_setup.get_logger().info(f'Exporting selected samples to database {fileName}')
        if not fileName.lower().endswith(".db"):
            fileName += ".db"

        # sanity checks and removes the target_connection if it exists
        if 'target_connection' in QSqlDatabase().connectionNames():
            QSqlDatabase.database('target_connection').close()
            QSqlDatabase().removeDatabase('target_connection')
            os.remove("temp.db")

        # If file already exists, delete it and create a new one
        if os.path.exists(fileName):
            try:
                os.remove(fileName)
            except OSError as e:
                logger_setup.get_logger().critical(f'Could not clear existing database file: {e}')
                return

        tgt_db = QSqlDatabase().addDatabase('QSQLITE', 'target_connection')
        tgt_db.setDatabaseName(fileName)
        tgt_db.open()
        if not tgt_db.isOpen() or not turn_off_foreign_keys(tgt_db):
            logger_setup.get_logger().critical('Could not open target database')
            return
        # Ensure the database is created and static tables are set up
        if not update_database(tgt_db):
            logger_setup.get_logger().critical('Could not create target database')
            return

        # src_db is the current default database in use
        src_db = QSqlDatabase()

        if self.checked_filter_list or self.checked_filter_list:
            # Warn the user the that the filters will be applied at the sample level. Any unwanted aliquots, spots, and analyses
            # can be removed in the new database.
            response = QMessageBox.question(self, 'Export Samples',
                                 'Exporting samples will apply the filters at the sample level.\nAny unwanted aliquots, spots, and analyses can be removed in the new database.',
                                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
            if response != QMessageBox.StandardButton.Ok:
                logger_setup.get_logger().info('User canceled the database export')
                return
            filtered_sample_ids, filtered_aliquot_ids, filtered_grain_ids, filtered_spot_ids = find_parent_items(
                list(self.filtered_upb_ids), 'UPbAnalyses')
        else:
            filtered_sample_ids = []
        all_sample_ids = list(set(filtered_sample_ids + self.checked_sample_list))
        show_loading_dialog('Exporting', 'Exporting to database...')


        # subsets and copies over samples and related-data from src_db to tgt_db
        if not ExportDatabase.subset_database(src_db, tgt_db, all_sample_ids):
            logger_setup.get_logger().critical('Could not subset database')
            close_loading_dialog('Exporting', 'Exporting to database...')
            return
        tgt_db.commit()
        tgt_db.close()

        QSqlDatabase().removeDatabase('target_connection')
        close_loading_dialog('Exporting', 'Exporting to database...')

        # show completion message
        msg = QMessageBox.information(QMessageBox(), "Database Export", "Database has exported successfully",
                                      buttons=QMessageBox.StandardButton.Ok)
        # msg.exec()
        QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(os.path.dirname(fileName)))

    def export_to_excel(self) -> bool:
        """
        Method to export the current workbook and worksheets to excel. A workbook will save to a single Excel file,
        worksheets will be distinct sheets within the Excel file.
        :return: True for success, False for failure
        """
        # Prompt user for where to save the Excel file
        fileName, _ = QFileDialog.getSaveFileName(
            None,
            "Save Excel File",
            "",
            "Excel Files (*.xlsx)"
        )

        if not fileName:
            return False

        # Ensure the filename ends with .xlsx
        if not fileName.lower().endswith(".xlsx"):
            fileName += ".xlsx"

        # Create a new Excel workbook
        wb = Workbook()

        # The first sheet is created by default. We'll rename or replace it as we go.
        first_sheet = True

        for sheet_name, info in self.worksheet_tabs_dict.items():
            self.update_table_view(worksheet_name=sheet_name)
            sql = info['sql']
            if not sql:
                logger_setup.get_logger().warning(f'No SQL query found for sheet "{sheet_name}". Select columns to include in each sheet before exporting.')
                return False
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_name)

            query = QSqlQuery(db=self.database)
            query.prepare(sql)

            if not query.exec():
                logger_setup.get_logger().critical(
                    f'Error exporting SQL query to excel: {query.lastError().text()}')
                logger_setup.get_logger().debug(f'SQL command: {sql}')
                return False

            # Retrieve column names
            column_count = query.record().count()
            row_idx = 1
            # check to include headers, sometimes a string would be passed not boolean, so a typecast is needed
            if bool(self.worksheet_tabs_dict[sheet_name]['headers']):
                headers = [query.record().fieldName(i) for i in range(column_count)]

                # Write headers to the first row
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
                row_idx = 2

            # Write data rows
            while query.next():
                for col in range(column_count):
                    ws.cell(row=row_idx, column=col + 1, value=query.value(col))
                row_idx += 1

        # Attempt to save the workbook
        try:
            wb.save(fileName)
        except Exception as e:
            logger_setup.get_logger().critical(
                f'Error saving the Excel file')
            logger_setup.get_logger().debug(f'Exception: {e}')
            return False

        # Open the file using the system's default application
        QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(fileName))

        return True

    def export_to_csv(self, one_file=True) -> bool:
        """
        Method to export the current workbook and worksheets to csv. A single worksheet will save to a single csv file,
        multiple worksheets will be distinct csvs within the selected directory with their sheet name as filename.

        :param one_file: export to multiple csv files or a single one.
        :return: True for success, False for failure
        """
        # Prompt user for where to save the CSV file
        if one_file:
            fileName, _ = QFileDialog.getSaveFileName(
                None,
                "Save CSV File",
                "",
                "Comma-Separated Values Files (*.csv)"
            )

            if not fileName:
                return False

            # Ensure the filename ends with .csv
            if not fileName.lower().endswith(".csv"):
                fileName += ".csv"
        else:
            directory = QFileDialog.getExistingDirectory(None, "Select Directory to Save CSV Files", "")

            if not directory:
                return False

        for sheet_name, info in self.worksheet_tabs_dict.items():
            if not one_file:
                fileName = os.path.join(directory, f"{sheet_name}.csv")
            try:
                with open(fileName, mode='w', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)

                    self.update_table_view(worksheet_name=sheet_name)
                    sql = info['sql']

                    query = QSqlQuery(db=self.database)
                    query.prepare(sql)

                    if not query.exec():
                        logger_setup.get_logger().critical(
                            f'Query execution failed for {sheet_name}: {query.lastError().text()}')
                        logger_setup.get_logger().debug(f'SQL command: {sql}')
                        continue

                    column_count = query.record().count()

                    if self.worksheet_tabs_dict[sheet_name]['headers']:
                        headers = [query.record().fieldName(i) for i in range(column_count)]
                        # Write headers
                        writer.writerow(headers)

                    # Write data rows
                    while query.next():
                        row = [query.value(col) for col in range(column_count)]
                        writer.writerow(row)


            except Exception as e:
                logger_setup.get_logger().critical(f'Error saving the Excel file')
                logger_setup.get_logger().debug(f'Exception: {e}')
                return False

        # Open the file using the system's default application
        QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(fileName))

        return True

    def set_table(self, model, table: str):
        model.blockSignals(True)
        model.setTable(table)
        model.select()
        while model.canFetchMore():
            model.fetchMore()
        model.blockSignals(False)
        return model

    def rename_worksheet_tab(self, index):
        """Method to rename the worksheet tab when double-clicked. Updates the tab name and dictionary value"""
        if index == -1:
            return  # No tab was double-clicked

        current_worksheet_name = self.workbooktabs.tabText(index)

        # Prompt the user for a new name
        new_name, ok = QInputDialog.getText(self, "Rename Worksheet", "Enter new worksheet name:",
                                            text=current_worksheet_name)
        if not ok or not new_name:
            return  # User canceled or didn't enter a name

        if new_name in self.worksheet_tabs_dict:
            QMessageBox.warning(self, "Duplicate Name", "A worksheet with that name already exists.")
            return

        # Update the workbook_tabs dictionary
        self.worksheet_tabs_dict[new_name] = self.worksheet_tabs_dict.pop(current_worksheet_name)

        # Update the tab text
        self.workbooktabs.setTabText(index, new_name)
        self.previous_worksheet = new_name

        # Change the format type to custom
        if self.exportformat_comboBox.currentText() != 'Custom':
            self.exportformat_comboBox.currentIndexChanged.disconnect()
            self.exportformat_comboBox.setCurrentText('Custom')
            self.exportformat_comboBox.currentIndexChanged.connect(self.export_format)
            self.save_checkbox_states()

    def rename_column(self, column_index, model):
        """Show an input dialog to rename a column header."""
        current_name = str(model.headerData(column_index, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole))

        new_name, ok = QInputDialog.getText(self, "Rename Column",
                                            f"Enter new name for '{current_name}':",
                                            text=current_name)
        if ok and new_name.strip():
            model.setHeaderData(column_index, Qt.Orientation.Horizontal, new_name)

        if self.exportformat_comboBox.currentText() != 'Custom':
            self.exportformat_comboBox.currentIndexChanged.disconnect()
            self.exportformat_comboBox.setCurrentText('Custom')
            self.exportformat_comboBox.currentIndexChanged.connect(self.export_format)
            self.save_checkbox_states()

    def open_columnname_mapping_dialog(self):
        """
        Helper method to opens a column name mapping dialog when the signal is called.
        """
        # Get current selected columns
        current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        ordered_columns = self.worksheet_tabs_dict[current_worksheet_name].get('ordered_columns', [])

        if not ordered_columns:
            QMessageBox.warning(self, "No Columns Selected", "Please select columns before editing their name.")
            return

        # Open the dialog
        dialog = ColumnNamesDialog(self.column_name_mappings, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Get adjusted columns
            self.column_name_mappings = dialog.get_adjusted_columns()
            if self.exportformat_comboBox.currentText() != 'Custom':
                self.exportformat_comboBox.currentIndexChanged.disconnect()
                self.exportformat_comboBox.setCurrentText('Custom')
                self.exportformat_comboBox.currentIndexChanged.connect(self.export_format)
                self.save_checkbox_states()
            self.update_table_view()

    def open_column_order_dialog(self):
        """
        Helper method to opens a column order mapping dialog when the signal is called.
        """
        # Get current selected columns
        current_worksheet_name = self.workbooktabs.tabText(self.workbooktabs.currentIndex())
        ordered_columns = self.worksheet_tabs_dict[current_worksheet_name].get('ordered_columns', [])

        if not ordered_columns:
            QMessageBox.warning(self, "No Columns Selected", "Please select columns before editing their order.")
            return

        # Open the dialog
        dialog = ColumnOrderDialog(ordered_columns, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Get adjusted columns
            adjusted_columns = dialog.get_adjusted_columns()
            # Update the selected columns
            self.worksheet_tabs_dict[current_worksheet_name]['ordered_columns'] = adjusted_columns
            # Update the table view with the new column order
            if self.exportformat_comboBox.currentText() != 'Custom':
                self.exportformat_comboBox.currentIndexChanged.disconnect()
                self.exportformat_comboBox.setCurrentText('Custom')
                self.exportformat_comboBox.currentIndexChanged.connect(self.export_format)
                self.save_checkbox_states()
            self.update_table_view(order_changed=True)


class ColumnOrderDialog(QDialog):
    def __init__(self, ordered_columns, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Column Order")
        self.resize(300, 400)
        self.ordered_columns = ordered_columns

        # Create widgets
        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QListWidget.SelectionMode.SingleSelection)
        self.list_widget.setDragDropMode(QListWidget.DragDropMode.InternalMove)

        for column_name in self.ordered_columns:
            self.list_widget.addItem(column_name)

        self.delete_button = QPushButton("Delete Selected")
        self.ok_button = QPushButton("OK")
        self.cancel_button = QPushButton("Cancel")

        # Connect signals
        self.delete_button.clicked.connect(self.delete_selected_item)
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

        # Layouts
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.delete_button)
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.ok_button)

        main_layout = QVBoxLayout()
        main_layout.addWidget(self.list_widget)
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)

    def delete_selected_item(self):
        current_row = self.list_widget.currentRow()
        if current_row >= 0:
            self.list_widget.takeItem(current_row)
        else:
            QMessageBox.warning(self, "No Selection", "Please select an item to delete.")

    def get_adjusted_columns(self):
        adjusted_columns = []
        for index in range(self.list_widget.count()):
            item_text = self.list_widget.item(index).text()
            adjusted_columns.append(item_text)
        return adjusted_columns


class ColumnNamesDialog(QDialog):
    def __init__(self, mapped_columns, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Column Names")
        self.resize(350, 450)
        self.mapped_columns = mapped_columns

        # Create widgets
        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.list_widget.setDragDropMode(QListWidget.DragDropMode.InternalMove)
        self.list_widget.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)

        # Populate list with columns
        for original, field_name in self.mapped_columns.items():
            item = QListWidgetItem(f"{original}: {field_name}")
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)  # Make items editable
            self.list_widget.addItem(item)

        # Buttons
        self.rename_button = QPushButton("Rename Selected")
        self.ok_button = QPushButton("OK")
        self.cancel_button = QPushButton("Cancel")

        # Connect signals
        self.rename_button.clicked.connect(self.rename_selected_item)
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

        # Layouts
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.rename_button)
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.ok_button)

        main_layout = QVBoxLayout()
        main_layout.addWidget(self.list_widget)
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)

    def rename_selected_item(self):
        current_item = self.list_widget.currentItem()
        if current_item:
            self.list_widget.editItem(current_item)
        else:
            QMessageBox.warning(self, "No Selection", "Please select an item to rename.")

    def delete_selected_item(self):
        current_row = self.list_widget.currentRow()
        if current_row >= 0:
            self.list_widget.takeItem(current_row)
        else:
            QMessageBox.warning(self, "No Selection", "Please select an item to delete.")

    def get_adjusted_columns(self):
        """Returns values to be set in the column_name_mappings dictionary based on the current list widget items. The keys are the original field names, the values are the new names."""
        adjusted_columns = {}
        for index, (original_name, field_name) in enumerate(self.mapped_columns.items()):
            item_text = self.list_widget.item(index).text()

            # Extract the new name from the list item text (assuming format: "original: new_name")
            if ": " in item_text:
                _, new_name = item_text.split(": ", 1)
            else:
                new_name = item_text  # Fallback if formatting isn't as expected

            # Store in dictionary
            adjusted_columns[original_name] = new_name  # Mapping original field to new name

        return adjusted_columns
