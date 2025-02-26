import sys
import os
import json
import sqlite3
from dataclasses import field

import pandas as pd
import qtawesome
from difflib import get_close_matches

from PyQt6 import QtCore
from PyQt6.QtSql import QSqlDatabase, QSqlQuery, QSqlRecord, QSqlTableModel
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, PatternFill

from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel,
    QComboBox, QTableWidget, QTableWidgetItem, QMessageBox, QHBoxLayout,
    QLineEdit, QInputDialog, QMenu, QDialog, QFormLayout, QSplitter, QAbstractItemView, QTableView, QCheckBox,
    QProgressDialog, QListWidget, QAbstractButton
)
from PyQt6.QtCore import Qt, QPoint, QSize, QEventLoop
from PyQt6.QtGui import QBrush, QColor, QFont, QAction

from Functions import SQLUtils, Savepoint_manager
from Functions.Database_manager import update_database
from Functions.Savepoint_manager import SavepointManager, create_savepoint, rollback_savepoint, release_savepoint

from Functions.Settings_manager import settings
from Functions.Widget_classes import (
    get_selected_tree_ids, CheckableComboBox, CheckableSqlTableModel, SearchableComboBox, set_table, CheckableTreeModel,
    CheckableTreeCombobox, save_expanded_state, get_name_column)
from ui.EditTable import EditTable
from ui.EditTree import EditTree
from ui.AddTags import AddTags
from ui.AddTreeTags import AddTreeTags
from Functions.Widget_classes import CompleterInputDialog

CONFIG_FILE = 'column_mappings.json'

class ColumnMapDialog(QDialog):
    """
    Dialog that creates one ComboBox per dictionary key, ensuring only
    one ComboBox can be non-'None' at a time.
    """
    def __init__(self, original_header, current_field, parent=None):
        """
        :param fields_dict: Dict[str, List[str]] of possible values for each field.
                            e.g. {"Sample Name": ["opt1","opt2"], "Aliquot Name": ["optA","optB"]}
        """
        super().__init__(parent)
        self.setWindowTitle(f"Column Mapper {original_header}")

        # Keep track of combo boxes so we can manipulate them easily
        self.combos = []
        self._is_updating = False

        form_layout = QFormLayout()

        # Create a combo box for each dictionary entry
        for field_label, possible_values in SQLUtils.upb_possible_user_input_fields.items():
            # combo = QComboBox()
            combo = SearchableComboBox()
            # We'll prepend a 'None' option. You could also use an empty string, etc.
            combo.addItem("None")
            combo.addItems(possible_values)
            # Now keep the user from adding items from the completer
            combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)

            # Connect signal so that if this combo changes,
            # we reset all others back to 'None'.
            combo.selection_changed.connect(self.on_combo_changed)

            # if current_field is not None and current_field in possible_values:
            #     combo.setCurrentText(current_field)

            form_layout.addRow(field_label + ":", combo)
            self.combos.append(combo)

        if current_field is not None and current_field != "None":
            self.on_combo_changed()

        # Add an OK button for closing
        self.btn_ok = QPushButton("OK")
        self.btn_ok.clicked.connect(self.accept)

        # Wrap everything in a layout
        main_layout = QVBoxLayout()
        main_layout.addLayout(form_layout)
        main_layout.addWidget(self.btn_ok)
        self.setLayout(main_layout)



    def on_combo_changed(self):
        """
        Triggered whenever the current index of any combo changes.
        Ensures only one combo has a non-'None' value at a time.
        """
        if self._is_updating:
            return

        # Avoid recursion / repeated signals while we're updating
        self._is_updating = True

        # Figure out which combo box triggered
        triggered_combo = self.sender()
        if not isinstance(triggered_combo, QComboBox):
            # Should never happen in this example, but just a guard
            self._is_updating = False
            return

        # If the user selected something other than 'None',
        # reset all other combos to 'None'
        if triggered_combo.currentIndex() != 0:  # i.e., not the "None" entry
            for combo in self.combos:
                if combo is not triggered_combo:
                    combo.setCurrentIndex(0)

        self._is_updating = False

    def get_selected_value(self):
        """
        Returns which combo was selected (if any), or 'None' if all are 'None'.
        You can customize how you want this data returned.
        """
        for combo in self.combos:
            if combo.currentIndex() != 0 and combo.currentIndex() != -1:  # i.e., not "None" or no selection
                return combo.currentText()
        return "None"

class ImportWizardDialog(QWidget):
    """
    Main window of the application:
      - Left table (pinned): Sample ID, Aliquot ID, Spot ID (editable).
      - Right table (main): Excel data + 4 optional columns appended
        for Lab Facilities, Source, Analysis Method, Instrument (all editable).
      - Context menus in both tables to set selected cells to a user-defined value.
    """
    data_imported = QtCore.pyqtSignal(list)
    def __init__(self):
        super().__init__()

        self.setWindowTitle("UPb Import Wizard")
        self.loadWindowState()

        main_layout = QVBoxLayout(self)

        # Top bar: file selection, sheet, etc.
        top_layout = QHBoxLayout()
        self.btn_select = QPushButton("Select Excel File")
        self.btn_select.setFixedWidth(150)
        self.btn_select.clicked.connect(self.select_file)
        top_layout.addWidget(self.btn_select)

        self.label_file = QLabel("No file selected.")
        top_layout.addWidget(self.label_file)

        self.sheet_instructions = QLabel("Select sheet with U-Pb data:")
        self.sheet_instructions.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        top_layout.addWidget(self.sheet_instructions)

        self.combo_sheets = QComboBox()
        self.combo_sheets.setFixedWidth(150)
        top_layout.addWidget(self.combo_sheets)
        self.combo_sheets.currentIndexChanged.connect(self.load_sheet)

        main_layout.addLayout(top_layout)

        combo_box_layout = QHBoxLayout()
        combo_box_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        # Delimiter label + line edit
        delimiter_label = QLabel("Delimiter:")
        delimiter_label.setFixedWidth(50)
        self.delimiter_edit = QLineEdit()
        self.delimiter_edit.setPlaceholderText("e.g., -, etc.")
        self.delimiter_edit.setFixedSize(QSize(100, 25))
        self.delimiter_edit.textChanged.connect(self.update_left_table_on_delimiter_change)  # Connect signal

        combo_box_layout.addWidget(delimiter_label)
        combo_box_layout.addWidget(self.delimiter_edit)

        combo_box_layout.addStretch(1)

        combo_box_layout.addWidget(QLabel("Notice: These dropdowns will overwrite all data in the tables.   "), 1, Qt.AlignmentFlag.AlignLeft)

        # ComboBox for setting Reference
        self.combo_reference_comboBox = CheckableComboBox()
        self.combo_reference_comboBox.setFixedWidth(150)
        self.combo_reference_comboBox.set_single_click(True)

        self.combo_reference = CheckableSqlTableModel()
        self.combo_reference = set_table(self.combo_reference, 'References')
        self.combo_reference_comboBox.setModel(self.combo_reference)
        self.combo_reference_comboBox.closing.connect(
            lambda: self.set_all_rows("Reference Display", self.combo_reference))
        combo_box_layout.addWidget(QLabel("Reference"))
        combo_box_layout.addWidget(self.combo_reference_comboBox)
        self.combo_reference_comboBox.set_line_edit_text(None)

        # ComboBox for setting Instrument
        self.combo_instrument_comboBox = CheckableComboBox()
        self.combo_instrument_comboBox.setFixedWidth(150)
        self.combo_instrument_comboBox.set_single_click(True)

        self.combo_instrument = CheckableSqlTableModel()
        self.combo_instrument = set_table(self.combo_instrument, "Instruments")
        self.combo_instrument_comboBox.setModel(self.combo_instrument)
        self.combo_instrument_comboBox.closing.connect(
            lambda: self.set_all_rows("Instrument Name", self.combo_instrument))
        combo_box_layout.addWidget(QLabel("Instrument"))
        combo_box_layout.addWidget(self.combo_instrument_comboBox)
        self.combo_instrument_comboBox.set_line_edit_text(None)

        # ComboBox for setting LabFacility
        self.combo_lab_facility_comboBox = CheckableComboBox()
        self.combo_lab_facility_comboBox.setFixedWidth(150)
        self.combo_lab_facility_comboBox.set_single_click(True)

        self.combo_lab_facility = CheckableSqlTableModel()
        self.combo_lab_facility = set_table(self.combo_lab_facility, "LabFacilities")
        self.combo_lab_facility_comboBox.setModel(self.combo_lab_facility)
        self.combo_lab_facility_comboBox.closing.connect(
            lambda: self.set_all_rows("Lab Facility Name", self.combo_lab_facility))
        combo_box_layout.addWidget(QLabel("Lab Facility"))
        combo_box_layout.addWidget(self.combo_lab_facility_comboBox)
        self.combo_lab_facility_comboBox.set_line_edit_text(None)

        # ComboBox for setting UPbAnalysisMethod
        self.combo_upb_analysis_method_comboBox = CheckableTreeCombobox()
        self.combo_upb_analysis_method_comboBox.setFixedWidth(150)
        self.combo_upb_analysis_method_comboBox.set_single_click(True)

        self.upb_analysis_method = QSqlTableModel()
        self.upb_analysis_method = set_table(self.upb_analysis_method, "UPbAnalysisMethods")
        self.combo_upb_analysis_method = CheckableTreeModel()
        self.combo_upb_analysis_method.setSourceModel(self.upb_analysis_method)
        self.combo_upb_analysis_method_comboBox.setModel(self.combo_upb_analysis_method)
        self.combo_upb_analysis_method_comboBox.set_single_click(True)
        self.combo_upb_analysis_method_comboBox.closing.connect(
            lambda: self.set_all_rows("UPb Analysis Method Name", self.combo_upb_analysis_method))
        combo_box_layout.addWidget(QLabel("UPb Analysis Method"))
        combo_box_layout.addWidget(self.combo_upb_analysis_method_comboBox)
        self.combo_upb_analysis_method_comboBox.set_line_edit_text(None)

        main_layout.addLayout(combo_box_layout)

        formats_layout = QHBoxLayout()
        formats_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        self.delimiter_checkbox = QCheckBox('Enable Delimiter?')
        self.delimiter_checkbox.checkStateChanged.connect(self.update_left_table_on_delimiter_change)
        formats_layout.addWidget(self.delimiter_checkbox, Qt.AlignmentFlag.AlignLeft)

        formats_layout.addStretch(4)

        self.btn_add_column = QPushButton("Add Column")
        self.btn_add_column.setFixedWidth(150)
        self.btn_add_column.clicked.connect(lambda: self.add_column(None, False))
        formats_layout.addWidget(self.btn_add_column)


        self.get_valid_unit_formats()

        self.ratio_error_combobox = QComboBox()
        # self.ratio_error_combobox.setFixedWidth(100)
        for display_text, backend_id in self.error_formats:
            self.ratio_error_combobox.addItem(display_text, backend_id)
        formats_layout.addWidget(QLabel("Ratio Error"))
        formats_layout.addWidget(self.ratio_error_combobox)
        self.ratio_error_combobox.setCurrentText(settings.value('ratio_error_format_abbreviation'))

        self.age_error_combobox = QComboBox()
        # self.age_error_combobox.setFixedWidth(100)
        for display_text, backend_id in self.error_formats:
            self.age_error_combobox.addItem(display_text, backend_id)
        formats_layout.addWidget(QLabel("Age Error"))
        formats_layout.addWidget(self.age_error_combobox)
        self.age_error_combobox.setCurrentText(settings.value('age_error_format_abbreviation'))

        self.age_unit_combobox = QComboBox()
        # self.age_unit_combobox.setFixedWidth(100)
        for display_text, backend_id in self.age_formats:
            self.age_unit_combobox.addItem(display_text, backend_id)
        formats_layout.addWidget(QLabel("Age Unit"))
        formats_layout.addWidget(self.age_unit_combobox)
        self.age_unit_combobox.setCurrentText(settings.value('age_unit_abbreviation'))

        self.spot_size_unit_combobox = QComboBox()
        # self.spot_size_combobox.setFixedWidth(100)
        for display_text, backend_id in self.distance_units:
            self.spot_size_unit_combobox.addItem(display_text, backend_id)
        formats_layout.addWidget(QLabel("Spot Size Unit"))
        formats_layout.addWidget(self.spot_size_unit_combobox)
        self.spot_size_unit_combobox.setCurrentText(settings.value('spotsize_unit_abbreviation'))

        self.conc_error_combobox = QComboBox()
        # self.conc_error_combobox.setFixedWidth(150)
        for display_text, backend_id in self.concordance_formats:
            self.conc_error_combobox.addItem(display_text, backend_id)
        formats_layout.addWidget(QLabel("Concordance Error"))
        formats_layout.addWidget(self.conc_error_combobox)
        self.conc_error_combobox.setCurrentText(settings.value('concordance_format_abbreviation'))

        main_layout.addLayout(formats_layout)


        # Splitter for left (pinned) vs right (main) tables
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Left pinned table: 3 columns for Sample Name, Aliquot Name, Spot Name
        self.left_table = QTableWidget()
        self.left_table.setColumnCount(3)
        self.left_table.setHorizontalHeaderLabels(["Sample Name", "Aliquot Name", "Spot Name"])
        self.left_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.left_table.customContextMenuRequested.connect(self.show_left_table_context_menu)

        # Right table for the actual Excel data
        self.right_table = QTableWidget()
        self.right_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.right_table.customContextMenuRequested.connect(self.show_right_table_context_menu)

        # Enable the context menu for the checkable combo boxes and connect the signals
        self.combo_reference_comboBox.enable_context_menu(True)
        self.combo_reference_comboBox.edit_triggered.connect(self.edit_combo_box)
        self.combo_reference_comboBox.add_triggered.connect(self.add_combo_box)
        self.combo_instrument_comboBox.enable_context_menu(True)
        self.combo_instrument_comboBox.edit_triggered.connect(self.edit_combo_box)
        self.combo_instrument_comboBox.add_triggered.connect(self.add_combo_box)
        self.combo_lab_facility_comboBox.enable_context_menu(True)
        self.combo_lab_facility_comboBox.edit_triggered.connect(self.edit_combo_box)
        self.combo_lab_facility_comboBox.add_triggered.connect(self.add_combo_box)
        self.combo_upb_analysis_method_comboBox.enable_context_menu(True)
        self.combo_upb_analysis_method_comboBox.edit_triggered.connect(self.edit_combo_box)
        self.combo_upb_analysis_method_comboBox.add_triggered.connect(self.add_combo_box)

        # Connect the header double-click signal to the handler
        header = self.right_table.horizontalHeader()
        header.sectionDoubleClicked.connect(self.handle_header_double_clicked)

        # Scroll synchronization (vertical)
        self.left_table.verticalScrollBar().valueChanged.connect(
            self.right_table.verticalScrollBar().setValue
        )
        self.right_table.verticalScrollBar().valueChanged.connect(
            self.left_table.verticalScrollBar().setValue
        )

        splitter.addWidget(self.left_table)
        splitter.addWidget(self.right_table)
        splitter.setStretchFactor(0, 1)  # left narrower
        splitter.setStretchFactor(1, 3)  # right expands

        main_layout.addWidget(splitter)

        # Bottom bar: mapping + import
        bottom_layout = QHBoxLayout()
        self.btn_save_mapping = QPushButton("Save Mapping")
        self.btn_save_mapping.clicked.connect(self.save_mapping)
        bottom_layout.addWidget(self.btn_save_mapping)

        self.btn_load_mapping = QPushButton("Load Mapping")
        self.btn_load_mapping.clicked.connect(self.load_mapping)
        bottom_layout.addWidget(self.btn_load_mapping)

        self.validate_button = QPushButton("Validate Sample Names")
        self.validate_button.clicked.connect(self.validate_ids)
        bottom_layout.addWidget(self.validate_button)

        self.btn_import = QPushButton("Import to Database")
        self.btn_import.clicked.connect(self.check_and_import)
        bottom_layout.addWidget(self.btn_import)

        main_layout.addLayout(bottom_layout)

        self.setLayout(main_layout)

        # DataFrame for the right table
        self.df = None
        # Mappings for right table columns
        self.column_mappings = {}
        # Rejected rows
        self.rejected_rows = set()

        # openpyxl workbook
        self.wb = None
        self.current_sheet_name = None

        # Icons for accepted/rejected
        self.rejected_icon = qtawesome.icon('fa5s.minus-circle', color='red', scale_factor=1.0)
        self.accepted_icon = qtawesome.icon('fa5s.check', color='green', scale_factor=1.0)

        # Sample IDs added or updated during import
        self.sample_ids = []

        # Flash fill connections
        self.left_table.cellChanged.connect(self.handle_cell_change)
        self.right_table.cellChanged.connect(self.handle_cell_change)

        self.right_table.cellClicked.connect(self.on_cell_clicked)

        # self.right_table.cellClicked.connect(self.handle_cell_click)

        self.right_table.verticalHeader().sectionDoubleClicked.connect(self.handle_vertical_header_double_click)


        # todo fix these context menus and methods to allow for multi-column set values

        self.right_table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.right_table.horizontalHeader().customContextMenuRequested.connect(self.show_right_header_context_menu)

        self.right_table.verticalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.right_table.verticalHeader().customContextMenuRequested.connect(self.show_right_table_vertical_header_context_menu)

        self.left_table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.left_table.horizontalHeader().customContextMenuRequested.connect(self.show_left_header_context_menu)

        self.conflict_mode = "skip"

        self.deactivate_widgets()

    def closeEvent(self, a0):
        self.combo_reference_comboBox.disconnect()
        self.combo_instrument_comboBox.disconnect()
        self.combo_lab_facility_comboBox.disconnect()
        self.combo_upb_analysis_method_comboBox.disconnect()
        super().closeEvent(a0)

    def deactivate_widgets(self):
        self.btn_save_mapping.setEnabled(False)
        self.btn_load_mapping.setEnabled(False)
        self.btn_import.setEnabled(False)
        self.validate_button.setEnabled(False)
        self.btn_add_column.setEnabled(False)
        self.delimiter_edit.setEnabled(False)
        self.delimiter_checkbox.setEnabled(False)
        self.combo_reference_comboBox.setEnabled(False)
        self.combo_instrument_comboBox.setEnabled(False)
        self.combo_lab_facility_comboBox.setEnabled(False)
        self.combo_upb_analysis_method_comboBox.setEnabled(False)
        self.ratio_error_combobox.setEnabled(False)
        self.age_error_combobox.setEnabled(False)
        self.age_unit_combobox.setEnabled(False)
        self.spot_size_unit_combobox.setEnabled(False)
        self.conc_error_combobox.setEnabled(False)

    def activate_widgets(self):
        self.btn_save_mapping.setEnabled(True)
        self.btn_load_mapping.setEnabled(True)
        self.btn_import.setEnabled(True)
        self.validate_button.setEnabled(True)
        self.btn_add_column.setEnabled(True)
        self.delimiter_edit.setEnabled(True)
        self.delimiter_checkbox.setEnabled(True)
        self.combo_reference_comboBox.setEnabled(True)
        self.combo_instrument_comboBox.setEnabled(True)
        self.combo_lab_facility_comboBox.setEnabled(True)
        self.combo_upb_analysis_method_comboBox.setEnabled(True)
        self.ratio_error_combobox.setEnabled(True)
        self.age_error_combobox.setEnabled(True)
        self.age_unit_combobox.setEnabled(True)
        self.spot_size_unit_combobox.setEnabled(True)
        self.conc_error_combobox.setEnabled(True)

    def on_cell_clicked(self, row, column):
        header_name = self.right_table.horizontalHeaderItem(column).text()

        table_name_map = {
            "Reference Display": '"References"',
            "Instrument Name": "Instruments",
            "Lab Facility Name": "LabFacilities",
            "UPb Analysis Method Name": "UPbAnalysisMethods"
        }

        if header_name in table_name_map:
            self.show_listwidget_popup(row, column, header_name, table_name_map[header_name])

    def show_listwidget_popup(self, row, column, header_name, table_name):
        popup = QDialog(self)
        popup.setWindowTitle("Select a valid {} value".format(header_name))

        list_widget = QListWidget(popup)

        # Query database for ID and name
        query = QSqlQuery(f"SELECT * FROM {table_name}")
        data_map = {}  # Dictionary to store ID-Name mapping

        while query.next():
            item_id = query.value(0)
            item_name = query.value(get_name_column(table_name))
            data_map[item_name] = item_id  # Store in map
            list_widget.addItem(item_name)  # Display only name in list

        # Handle item selection
        def on_item_clicked(item):
            selected_name = item.text()
            selected_id = data_map[selected_name]  # Get the associated ID
            self.create_table_item(row, column, selected_name, selected_id, header_name, table_name)
            popup.accept()

        list_widget.itemClicked.connect(on_item_clicked)

        # Set layout
        layout = QVBoxLayout()
        layout.addWidget(list_widget)
        popup.setLayout(layout)

        popup.exec()

    def create_table_item(self, row, column, name, id, header_name, table_name):
        """Creates a QTableWidgetItem with stored data."""
        field_to_column = {
            "ReferenceID": self.get_column_index("ReferenceID"),
            "Reference Display": self.get_column_index("Reference Display"),
            "InstrumentID": self.get_column_index("InstrumentID"),
            "Instrument Name": self.get_column_index("Instrument Name"),
            "LabFacilityID": self.get_column_index("LabFacilityID"),
            "Lab Facility Name": self.get_column_index("Lab Facility Name"),
            "UPbAnalysisMethodID": self.get_column_index("UPbAnalysisMethodID"),
            "UPb Analysis Method Name": self.get_column_index("UPb Analysis Method Name")
        }
        # item = QTableWidgetItem(name)

        self.right_table.setItem(row, field_to_column[header_name], QTableWidgetItem(name))
        self.right_table.setItem(row, field_to_column[header_name]+1, QTableWidgetItem(str(id)))

    def get_valid_unit_formats(self):
        age_unit_query = QSqlQuery()
        age_unit_query.prepare(
            'SELECT AgeUnitAbbreviation, AgeUnitID From AgeUnits')
        self.age_formats = []

        if age_unit_query.exec():
            while age_unit_query.next():
                self.age_formats.append((age_unit_query.value(0), age_unit_query.value(1)))
        else:
            print("Failed to execute query:", age_unit_query.lastError().text())



        distance_units_query = QSqlQuery()
        distance_units_query.prepare(
            'SELECT DistanceUnitAbbreviation, DistanceUnitID From DistanceUnits')
        self.distance_units = []

        if distance_units_query.exec():
            while distance_units_query.next():
                self.distance_units.append((distance_units_query.value(0), distance_units_query.value(1)))
        else:
            print("Failed to execute query:", distance_units_query.lastError().text())



        concordance_units_query = QSqlQuery()
        concordance_units_query.prepare(
            'SELECT ConcordanceFormatAbbreviation, ConcordanceFormatID From ConcordanceFormats')
        self.concordance_formats = []

        if concordance_units_query.exec():
            while concordance_units_query.next():
                self.concordance_formats.append((concordance_units_query.value(0), concordance_units_query.value(1)))
        else:
            print("Failed to execute query:", concordance_units_query.lastError().text())



        error_type_format_query = QSqlQuery()
        error_type_format_query.prepare(
            'SELECT ErrorFormatAbbreviation, ErrorFormatID From ErrorFormats')
        self.error_formats = []

        if error_type_format_query.exec():
            while error_type_format_query.next():
                self.error_formats.append((error_type_format_query.value(0), error_type_format_query.value(1)))
        else:
            print("Failed to execute query:", error_type_format_query.lastError().text())

    def validate_ids(self):
        """
        Validate Sample Name, Aliquot Name, and Spot Name in the left_table against the database.
        Flag rows that have matching entries in the database.
        """

        row_count = self.left_table.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "No Data", "There are no rows to import.")
            return

        # Step 1: Check for empty cells in the left table
        empty_cells = self.check_empty_cells_in_left_table()

        if empty_cells:
            # Step 2: Show dialog to ask if the user wants to use default values
            use_defaults = self.ask_to_use_default_values(empty_cells)

            if use_defaults:
                # Step 3: Fill empty cells with default values
                self.fill_empty_cells_with_defaults(empty_cells)

                # Optional: Give user a chance to review and adjust the values
                QMessageBox.information(self, "Review",
                                        "The empty cells have been filled with default values. Please review before clicking import again.")

        # Prepare SQL queries for validation
        # find values where SampleName in the database matches listed value
        sample_query = QSqlQuery()
        sample_query.prepare("SELECT SampleID FROM Samples WHERE SampleName = :sample_name COLLATE NOCASE")

        # find values where AliquotName and SampleID match in the database.
        aliquot_query = QSqlQuery()
        aliquot_query.prepare("SELECT AliquotID FROM Aliquots WHERE AliquotName = :aliquot_name COLLATE NOCASE AND SampleID = :sample_id")

        # find values where SpotName and AliquotID match in the database.
        spot_query = QSqlQuery()
        spot_query.prepare("SELECT SpotID FROM Spots WHERE SpotName = :spot_name COLLATE NOCASE AND AliquotID = :aliquot_id COLLATE NOCASE")

        # Iterate through rows in the left_table
        for row in range(self.left_table.rowCount()):
            sample_name = self.left_table.item(row, 0).text() if self.left_table.item(row, 0) else None
            aliquot_name = self.left_table.item(row, 1).text() if self.left_table.item(row, 1) else None
            spot_name = self.left_table.item(row, 2).text() if self.left_table.item(row, 2) else None

            # Check Sample Name
            sample_match = False
            if sample_name:
                sample_query.bindValue(":sample_name", sample_name)
                sample_match = sample_query.exec() and sample_query.next()
                sample_id = sample_query.value(0) if sample_match else None

            # Check Aliquot Name
            aliquot_match = False
            if aliquot_name:
                aliquot_query.bindValue(":aliquot_name", aliquot_name)
                aliquot_query.bindValue(":sample_id", sample_id)
                aliquot_match = aliquot_query.exec() and aliquot_query.next()
                aliquot_id = aliquot_query.value(0) if aliquot_match else None

            # Check Spot Name
            spot_match = False
            if spot_name:
                spot_query.bindValue(":spot_name", spot_name)
                spot_query.bindValue(":aliquot_id", aliquot_id)
                spot_match = spot_query.exec() and spot_query.next()
                spot_id = spot_query.value(0) if spot_match else None

            self.left_table.blockSignals(True)
            # Highlight the row if any match is found
            # Highlight matching cells
            if sample_match:
                item = self.left_table.item(row, 0)
                if item:
                    item.setBackground(QColor('#FCAE1E'))
            else:
                item = self.left_table.item(row, 0)
                if item:
                    item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            if aliquot_match:
                item = self.left_table.item(row, 1)
                if item:
                    item.setBackground(QColor('#FCAE1E'))
            else:
                item = self.left_table.item(row, 1)
                if item:
                    item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            if spot_match:
                item = self.left_table.item(row, 2)
                if item:
                    item.setBackground(QColor('#FCAE1E'))
            else:
                item = self.left_table.item(row, 2)
                if item:
                    item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            self.left_table.blockSignals(False)

        QMessageBox.information(self, "Validation Complete", "Validation of IDs is complete.")

    def edit_combo_box(self, pos):
        """
        Selected 'Edit' from the context menu for the combo box.
        Open the table edit dialog.
        Args:
            pos (QPoint): Position of the context menu request.
        """
        combo = self.sender()
        if combo == self.combo_reference_comboBox:
            table = "References"
        elif combo == self.combo_instrument_comboBox:
            table = "Instruments"
        elif combo == self.combo_lab_facility_comboBox:
            table = "LabFacilities"
        elif combo == self.combo_upb_analysis_method_comboBox:
            table = "UPbAnalysisMethods"
        else:
            return
        if table in SQLUtils.user_viewable_trees:
            dlg = EditTree(table)
        else:
            dlg = EditTable(table)
        dlg.exec()

    def add_combo_box(self, pos, action: QAction | None = None):
        """
        Selected an add action from the context menu for the combo box.
        :param pos: QPoint of the context menu request.
        :param action: QAction that was triggered.
        :return:
        """
        combo = self.sender()
        if combo == self.combo_reference_comboBox:
            table = "References"
        elif combo == self.combo_instrument_comboBox:
            table = "Instruments"
        elif combo == self.combo_lab_facility_comboBox:
            table = "LabFacilities"
        elif combo == self.combo_upb_analysis_method_comboBox:
            table = "UPbAnalysisMethods"
        else:
            return

        dlg = None
        dlg_args = None
        if table in SQLUtils.user_viewable_trees:
            save_expanded_state(table, combo.model(), combo.treeView())
            indexes = combo.treeView().selectedIndexes()
            item_ids, parent_ids, parent_rows = get_selected_tree_ids(combo.model(), indexes)
            if action:
                if action.text() == 'Insert above':
                    row = parent_rows[0]
                    parent_id = parent_ids[0]
                    dlg_args = (table, parent_id, row)
                elif action.text() == 'Insert below':
                    row = parent_rows[0] + 1
                    parent_id = parent_ids[0]
                    dlg_args = (None, parent_id, row)
                elif action.text() == 'Add child':
                    parent_id = item_ids[0]
                    dlg_args = (None, parent_id)
                elif action.text() == 'Add parent':
                    dlg_args = (item_ids, parent_ids, parent_rows)
                elif action.text() == 'Add to end':
                    dlg_args = (None, None)
            if dlg_args:
                dlg = AddTreeTags(table, *dlg_args)
        else:
            dlg = AddTags(table)
        dlg.exec()
        # combo.model().select()

    def show_left_header_context_menu(self, pos):
        """
        Show a context menu on the horizontal header to set column values or insert columns.
        Args:
            pos (QPoint): Position of the context menu request.
        """
        # Determine the column index under the cursor
        column = self.left_table.horizontalHeader().logicalIndexAt(pos)
        if column < 0:
            return

        menu = QMenu(self)

        # Add options to the menu
        set_value_action = menu.addAction("Set Entire Column to Value")
        set_blank_action = menu.addAction("Set Entire Column to Blank")

        # Execute the menu and get the selected action
        action = menu.exec(self.left_table.horizontalHeader().mapToGlobal(pos))
        if action == set_value_action:
            self.set_column_to_value(column, self.left_table)
        elif action == set_blank_action:
            self.set_column_to_blank(column, self.left_table)


    def show_right_header_context_menu(self, pos):
        """
        Show a context menu on the horizontal header to set column values or insert columns.
        Args:
            pos (QPoint): Position of the context menu request.
        """
        # Determine the column index under the cursor
        column_index = self.right_table.horizontalHeader().logicalIndexAt(pos)
        if column_index < 0:
            return

        menu = QMenu(self)

        # Add options to the menu
        set_value_action = menu.addAction("Set Entire Column to Value")
        set_blank_action = menu.addAction("Set Entire Column to Blank")
        insert_before_action = menu.addAction("Insert Column Before")
        insert_after_action = menu.addAction("Insert Column After")

        # Execute the menu and get the selected action
        action = menu.exec(self.right_table.horizontalHeader().mapToGlobal(pos))
        if action == set_value_action:
            self.set_column_to_value(column_index, self.right_table)
        elif action == set_blank_action:
            self.set_column_to_blank(column_index, self.right_table)
        elif action == insert_before_action:
            self.add_column(column_index, before=True)
        elif action == insert_after_action:
            self.add_column(column_index, before=False)

    def set_column_to_value(self, column, table: QTableWidget):
        """
        Set all cells in the specified column to a user-provided value.
        Args:
            column (int): The column index to update.
        """
        # Prompt the user for a value
        value, ok = QInputDialog.getText(self, "Set Column Value", f"Enter value for column {column + 1}:")
        if not ok or value is None:
            return  # User canceled

        # Update all rows in the column
        table.blockSignals(True)
        for row in range(table.rowCount()):
            item = table.item(row, column)
            if item is None:
                item = QTableWidgetItem()
                table.setItem(row, column, item)
            item.setText(value)
        table.blockSignals(False)

    def set_column_to_blank(self, column, table: QTableWidget):
        """
        Set all cells in the specified column to blank.
        Args:
            column (int): The column index to update.
        """
        # Update all rows in the column
        for row in range(table.rowCount()):
            item = table.item(row, column)
            if item is None:
                item = QTableWidgetItem()
                table.setItem(row, column, item)
            item.setText("")


    def handle_vertical_header_double_click(self, logical_index):
        """
        Handle double-clicks on vertical headers to mark rows as rejected.
        Args:
            logical_index (int): The row index corresponding to the double-clicked header.
        """
        item = self.right_table.item(logical_index, 0)
        if logical_index in self.rejected_rows:
            self.mark_selected_rows_rejected([item], False)
        else:
            self.mark_selected_rows_rejected([item], True)


    # ---------------------------
    #    Context Menu Methods
    # ---------------------------

    # def handle_cell_click(self, row, column):
    #     """
    #     Handle cell clicks for specific columns and show a popup with a QSqlTableModel.
    #     Args:
    #         row (int): Row index of the clicked cell.
    #         column (int): Column index of the clicked cell.
    #     """
    #     # Determine the clicked column's header
    #     header_item = self.right_table.horizontalHeaderItem(column)
    #     if not header_item:
    #         return
    #
    #     column_name = header_item.text()
    #
    #     # Map column names to database tables
    #     column_to_table = {
    #         "Reference Display": "References",
    #         "Instrument Name": "Instruments",
    #         "Lab Facility Name": "LabFacilities",
    #         "UPb Analysis Method Name": "UPbAnalysisMethods"
    #     }
    #
    #     if column_name in column_to_table:
    #         table_name = column_to_table[column_name]
    #         # self.show_table_popup(table_name, row, column)

    # def show_table_popup(self, table_name, row, column_index):
    #     """
    #     Show a popup with a QSqlTableModel for the specified table.
    #     Args:
    #         @param table_name: Name of the database table to display.
    #         @param column_index:
    #         @param row:
    #     """
    #     if table_name == "Reference Display":
    #         table_name = '"References"'
    #     elif table_name == "Instrument Name":
    #         table_name = "Instruments"
    #     elif table_name == "Lab Facility Name":
    #         table_name = "LabFacilities"
    #     elif table_name == "UPb Analysis Method Name":
    #         table_name = "UPbAnalysisMethods"
    #
    #     # Create a QSqlTableModel and set the table
    #     model = CheckableSqlTableModel()
    #     model = set_table(model, table_name)
    #
    #     # Create a QTableView to display the model
    #     combobox = QComboBox()
    #     # combobox.set_single_click(True)
    #     # combobox.set_line_edit_text(None)
    #     # combobox.setModel(model)
    #     combobox.addItems(["Edit", "Add"])
    #
    #     # combobox.closing.connect(lambda: self.set_cell_combobox(model, row, column_index))
    #
    #     self.right_table.setCellWidget(row, column_index, combobox)
    #     self.right_table.setColumnWidth(column_index, 200)

    # def set_cell_combobox(self, model, row, column_index):
    #     name_col = WC.name_column(model.tableName())
    #
    #     field = self.right_table.horizontalHeaderItem(column_index).text().strip()
    #
    #     for temp_row in range(model.rowCount()):
    #         name_index = model.index(temp_row, name_col)
    #         id_index = model.index(temp_row, 0)
    #
    #         if model.data(name_index, QtCore.Qt.ItemDataRole.CheckStateRole) == Qt.CheckState.Checked:
    #             checked_item_name = model.data(name_index, Qt.ItemDataRole.DisplayRole)
    #             checked_item_id = model.data(id_index, Qt.ItemDataRole.DisplayRole)
    #
    #     field_to_column = {
    #         "ReferenceID": self.get_column_index("ReferenceID"),
    #         "Reference Display": self.get_column_index("Reference Display"),
    #         "InstrumentID": self.get_column_index("InstrumentID"),
    #         "Instrument Name": self.get_column_index("Instrument Name"),
    #         "LabFacilityID": self.get_column_index("LabFacilityID"),
    #         "Lab Facility Name": self.get_column_index("Lab Facility Name"),
    #         "UPbAnalysisMethodID": self.get_column_index("UPbAnalysisMethodID"),
    #         "UPb Analysis Method Name": self.get_column_index("UPb Analysis Method Name"),
    #     }
    #
    #     # Update all rows in the column
    #     self.right_table.blockSignals(True)
    #
    #     if field == "Reference Display":
    #         id_column = field_to_column.get("ReferenceID")
    #     elif field == "Instrument Name":
    #         id_column = field_to_column.get("InstrumentID")
    #     elif field == "Lab Facility Name":
    #         id_column = field_to_column.get("LabFacilityID")
    #     elif field == "UPb Analysis Method Name":
    #         id_column = field_to_column.get("UPbAnalysisMethodID")
    #     else:
    #         id_column = None
    #
    #     name_column = field_to_column.get(field)
    #
    #     item = self.right_table.item(row, id_column)
    #     if item is None:
    #         item = QTableWidgetItem()
    #         self.right_table.setItem(row, id_column, item)
    #     item.setText(str(checked_item_id))
    #
    #     item = self.right_table.item(row, name_column)
    #     if item is None:
    #         item = QTableWidgetItem()
    #         self.right_table.setItem(row, name_column, item)
    #     item.setText(str(checked_item_name))
    #     self.right_table.blockSignals(False)


    def add_column(self, column_index=None, before=False, field=None):
        """
        Adds a column to the right QTableWidget
        @param column_index: Int index of the column to add.
        @param before:
        """
        # Open the Column Map Dialog to let the user select a column name and data type

        if field is None:
            dialog = ColumnMapDialog("New Column", "None", self)
            if dialog.exec():
                selected_field = dialog.get_selected_value()
            else:
                return
        else:
            selected_field = field

        # Ensure the user selected a valid field
        if selected_field == "None":
            QMessageBox.warning(self, "Invalid Selection", "Please select a valid column field.")
            return

        for test_idx in range(self.right_table.columnCount()):
            header = self.right_table.horizontalHeaderItem(test_idx)
            if header and header.text().startswith(selected_field):
                QMessageBox.warning(self, "Duplicate Column", f"Column '{selected_field}' already exists.")
                # todo this duplicate is not working
                return

        if column_index is None:
            # Insert the new column at the end of the table
            column_index = self.right_table.columnCount()
        else:
            if not before:
                column_index += 1

        self.right_table.insertColumn(column_index)

        # Set the column header
        header_text = f"{selected_field}"
        header_item = QTableWidgetItem(header_text)
        header_item.setBackground(QBrush(QColor("#ffffcc")))  # Light yellow background for new column
        self.right_table.setHorizontalHeaderItem(column_index, header_item)

        # Add the new column to the column mappings
        self.column_mappings[column_index] = (selected_field)

        self.right_table.blockSignals(True)
        # Initialize the column cells with empty values

        if selected_field == "Reference Display":
            field = "ReferenceID"
        elif selected_field == "Instrument Name":
            field = "InstrumentID"
        elif selected_field == "Lab Facility Name":
            field = "LabFacilityID"
        elif selected_field == "UPb Analysis Method Name":
            field = "UPbAnalysisMethodID"

        for row in range(self.right_table.rowCount()):
            self.right_table.setItem(row, column_index, QTableWidgetItem(""))
        self.right_table.blockSignals(False)

        #add additional ID column if column is References, Instruments, Analysis Methods, or Lab Facilities
        if selected_field in ["Reference Display", "Instrument Name", "Lab Facility Name", "UPb Analysis Method Name"]:
            # ADD ID Column to the tablewidget
            # Insert the new column at the end of the table
            column_index = self.right_table.columnCount()
            self.right_table.insertColumn(column_index)

            # Set the column header
            header_text = f"{field}"
            header_item = QTableWidgetItem(header_text)
            self.right_table.setHorizontalHeaderItem(column_index, header_item)

            # Add the new column to the column mappings
            # Add the new column to the column mappings
            self.column_mappings[column_index] = (field)

            self.right_table.blockSignals(True)
            # Initialize the column cells with empty values
            for row in range(self.right_table.rowCount()):
                self.right_table.setItem(row, column_index, QTableWidgetItem(""))
            self.right_table.blockSignals(False)
            # self.right_table.hideColumn(col_index)
            self.right_table.resizeColumnsToContents()

        # Notify the user only when method is initated from the user
        if field is None:
            QMessageBox.information(self, "Column Added", f"Column '{header_text}' added successfully.")

    def show_left_table_context_menu(self, pos: QPoint):
        """
        Context menu for the left table (Sample ID, Aliquot ID, Spot ID).
        Allows setting all selected cells to a user-defined value.
        """
        menu = QMenu(self)
        set_value_action = menu.addAction("Set Selected Cells to Value...")

        action = menu.exec(self.left_table.mapToGlobal(pos))
        if action == set_value_action:
            new_value, ok = QInputDialog.getText(self, "Set Value", "Enter new value:")
            if ok:
                for item in self.left_table.selectedItems():
                    item.setText(new_value)

    def show_right_table_context_menu(self, pos: QPoint):
        """
        Context menu for the right table.
        Includes remove rows, mark rows rejected, unmark,
        and set selected cells to a user-defined value.
        """
        menu = QMenu(self)
        remove_action = menu.addAction("Remove Selected Rows")
        reject_action = menu.addAction("Mark Selected Rows as Rejected")
        accept_action = menu.addAction("Mark Selected Rows as Accepted")
        set_value_action = menu.addAction("Set Selected Cells to Value...")
        remove_column = menu.addAction('Remove Selected Column')

        action = menu.exec(self.right_table.mapToGlobal(pos))
        if action == remove_action:
            self.remove_selected_rows()
        elif action == reject_action:
            self.mark_selected_rows_rejected(self.right_table.selectedItems(), True)
        elif action == accept_action:
            self.mark_selected_rows_rejected(self.right_table.selectedItems(), False)
        elif action == set_value_action:
            new_value, ok = QInputDialog.getText(self, "Set Value", "Enter new value:")
            if ok:
                for item in self.right_table.selectedItems():
                    item.setText(new_value)
        elif action == remove_column:
            self.remove_selected_columns()

        self.repaint()

    def show_right_table_vertical_header_context_menu(self, pos: QPoint):
        """
        Context menu for the right table.
        Includes remove rows, mark rows rejected, unmark,
        and set selected cells to a user-defined value.
        """
        row = self.right_table.verticalHeader().logicalIndexAt(pos)
        if row == -1:  # Ensure a valid header was clicked
            return

        menu = QMenu(self)
        remove_action = menu.addAction("Remove Selected Rows")
        reject_action = menu.addAction("Mark Selected Rows as Rejected")
        accept_action = menu.addAction("Mark Selected Rows as Accepted")

        action = menu.exec(self.right_table.mapToGlobal(pos))
        items = self.right_table.selectedItems()
        if action == remove_action:
            self.remove_selected_rows()
        elif action == reject_action:
            self.mark_selected_rows_rejected(items, True)
        elif action == accept_action:
            self.mark_selected_rows_rejected(items, False)

        self.repaint()

    # ---------------------------
    #     File & Sheet Loading
    # ---------------------------

    def handle_cell_change(self, row, column):
        """
        Handle cell value changes. Ask the user if they want to flash fill downward.
        """

        if self.sender() == self.left_table:
            target_table = self.left_table
        elif self.sender() == self.right_table:
            target_table = self.right_table
        else:
            return

        # Get the current value of the cell
        current_value = target_table.item(row, column).text().strip()
        if target_table.item(row + 1, column) is None:
            return

        next_value = target_table.item(row + 1, column).text().strip()
        # If the value is empty or invalid, ignore
        if (not current_value or
                next_value == current_value or
                len(next_value) > 0 or
                len(target_table.selectionModel().selectedIndexes()) > 1):
            return

        # Check if the user wants to flash fill
        reply = QMessageBox.question(
            self, "Flash Fill Downward",
            "Do you want to auto-fill downward with this value for blank cells?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            self.flash_fill_downward(target_table, row, column, current_value)

    def flash_fill_downward(self, target_table, start_row, column, value):
        """
        Flash fill a column downward starting from a given row.
        """

        target_table.blockSignals(True)

        # Start from the next row and go downward
        for row in range(start_row + 1, target_table.rowCount()):
            item = target_table.item(row, column)

            # If the cell is blank, fill it with the given value
            if item is None or not item.text().strip():
                # If the cell doesn't exist, create it
                if item is None:
                    item = QTableWidgetItem()
                    target_table.setItem(row, column, item)

                # Set the value
                item.setText(value)
            else:
                # Stop when a non-blank value is encountered
                break
        target_table.blockSignals(False)

    def select_file(self):
        dlg = QFileDialog(self)
        path, _ = dlg.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
        if path:
            self.selected_file_path = path
            self.label_file.setText(f"Selected File: {os.path.basename(path)}")
            try:
                self.wb = load_workbook(path, data_only=True, rich_text=True)
                self.combo_sheets.clear()
                self.combo_sheets.addItems(self.wb.sheetnames)
                # self.load_sheet(bypass=True)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to read Excel file:\n{e}")
                return
        self.activate_widgets()

    def load_sheet(self, bypass=False):

        if bypass:
            if QMessageBox.question(self, "Confirmation", "Loading this sheet will clear all existing data. Continue?",
                                 QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                 QMessageBox.StandardButton.Yes) == QMessageBox.StandardButton.No:
                return


        if not hasattr(self, 'selected_file_path') or not self.selected_file_path:
            QMessageBox.warning(self, "No File", "Please select an Excel file first.")
            return
        sheet_name = self.combo_sheets.currentText()
        if not sheet_name:
            QMessageBox.warning(self, "No Sheet", "Please select a sheet.")
            return

        try:
            self.df = pd.read_excel(self.selected_file_path, header=None, sheet_name=sheet_name, engine="openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to parse sheet with pandas:\n{e}")
            return

        # Remove initial blank rows
        while not self.df.empty and self.df.iloc[0].isna().all():
            self.df = self.df.iloc[1:].reset_index(drop=True)

        # Reset mapping & rejections
        self.column_mappings.clear()
        self.rejected_rows.clear()

        # Display data on the right table
        self.display_right_table_with_styles(sheet_name)

        # Build the left table rows
        self.sync_left_table_rows()

        # Auto-guess column names
        # self.auto_guess_column_names()

    def display_right_table_with_styles(self, sheet_name):
        """
        Display the right table with openpyxl-based formatting
        + add 4 extra columns for Lab Facilities, Source, Analysis Method, Instrument (editable).
        """
        sheet = self.wb[sheet_name]
        self.right_table.clear()
        self.right_table.setRowCount(0)
        self.right_table.setColumnCount(0)

        rows, cols = self.df.shape
        # We'll add 4 extra columns for Lab Facilities, Source, Analysis Method, Instrument
        # extra_cols = 4
        self.right_table.setRowCount(rows)
        self.right_table.setColumnCount(cols)

        # Set column headers for the loaded data columns
        for c in range(cols):
            col_name = str(self.df.columns[c])
            hdr_item = QTableWidgetItem(col_name)
            self.right_table.setHorizontalHeaderItem(c, hdr_item)

        # Populate cells for the loaded data
        self.left_table.blockSignals(True)
        self.right_table.blockSignals(True)
        for r in range(rows):
            row_rejected = False
            for c in range(cols):
                cell = sheet.cell(row=r+1, column=c+1)
                value = self.df.iat[r, c]
                disp_val = "NULL" if pd.isna(value) or value == "" else str(value)

                item = QTableWidgetItem(disp_val)
                font = cell.font
                fill = cell.fill

                # Foreground
                if font.color and hasattr(font.color, "rgb") and isinstance(font.color.rgb, str):
                    hex_col = "#" + font.color.rgb[-6:]
                    item.setForeground(QBrush(QColor(hex_col)))
                    if hex_col.lower() == "#ff0000":
                        row_rejected = True
                # else:
                #     item.setForeground(QBrush(Qt.GlobalColor.black))

                qfont = QFont()
                qfont.setBold(font.bold if font.bold else False)
                qfont.setItalic(font.italic if font.italic else False)
                qfont.setStrikeOut(font.strike if font.strike else False)
                item.setFont(qfont)

                if font.strike:
                    row_rejected = True

                # Background
                if isinstance(fill, PatternFill) and fill.fgColor and fill.fgColor.rgb:
                    bg_hex = "#" + fill.fgColor.rgb[-6:]
                    if fill.fill_type and fill.fill_type != "none":
                        item.setBackground(QBrush(QColor(bg_hex)))

                self.right_table.setItem(r, c, item)

            if row_rejected:
                self.rejected_rows.add(r)

        self.left_table.blockSignals(False)
        self.right_table.blockSignals(False)
        # Setup vertical header icons
        for r in range(rows):
            self.update_row_icon(r, (r in self.rejected_rows))

        self.right_table.resizeColumnsToContents()

    def sync_left_table_rows(self):
        """
        Make the left table have the same row count as the right table
        and add editable cells for Sample ID, Aliquot ID, Spot ID.
        """
        self.left_table.blockSignals(True)
        row_count = self.right_table.rowCount()
        self.left_table.setRowCount(row_count)
        for r in range(row_count):
            for c in range(3):
                if not self.left_table.item(r, c):
                    self.left_table.setItem(r, c, QTableWidgetItem(""))
        self.left_table.blockSignals(False)

        self.left_table.resizeColumnsToContents()

    # def auto_guess_column_names(self):
    #     """
    #     Use difflib to guess the best match from SQLUtils.upb_possible_input_fields for the right table columns
    #     (excluding the 4 appended columns).
    #     """
    #     import difflib
    #     cutoff = 0.5
    #
    #     total_cols = self.right_table.columnCount()
    #     # Exclude appended columns for Lab, Source, Method, Instrument
    #     # main_cols = total_cols - 4 if (total_cols > 4 and self.lab_col is not None) else total_cols
    #     main_cols = total_cols
    #     for col_idx in range(main_cols):
    #         original_header = self.right_table.horizontalHeaderItem(col_idx).text()
    #         best = difflib.get_close_matches(original_header, SQLUtils.upb_possible_user_input_fields, n=1, cutoff=cutoff)
    #         if best:
    #             field = best[0]
    #             self.column_mappings[col_idx] = (field)
    #             item = self.right_table.horizontalHeaderItem(col_idx)
    #             item.setText(f"{field}")
    #             item.setBackground(QBrush(QColor("#ffffcc")))
    #         else:
    #             self.column_mappings[col_idx] = ("None")

    # ---------------------------
    #     Context Menu Logic
    # ---------------------------

    def set_all_rows(self, field, model):
        """
        Set all rows in the specified column to the given value.
        Args:
            field (str): The field name (e.g., 'Reference', 'Instrument').
            value (str): The value to set.
        """
        # todo crashing on closing of drop down
        if isinstance(model.tableName, str):
            table = model.tableName
        else:
            table = model.tableName()
        name_column = get_name_column(table)
        # Determine the column index for the field
        source_checked_row = None
        checked_item_name = None
        checked_item_id = None
        if table != '':
            for row in range(model.rowCount()):
                name_index = model.index(row, name_column)
                id_index = model.index(row, 0)

                if model.data(name_index, QtCore.Qt.ItemDataRole.CheckStateRole) == Qt.CheckState.Checked:
                    checked_item_name = model.data(name_index, Qt.ItemDataRole.DisplayRole)
                    checked_item_id = model.data(id_index, Qt.ItemDataRole.DisplayRole)
                    source_checked_row = row
            if checked_item_name is None or checked_item_id is None:
                return
            print(checked_item_name, checked_item_id)
        else:
            # todo: add logic to get checked item from tree
            return


        field_to_column = {
            "ReferenceID": self.get_column_index("ReferenceID"),
            "Reference Display": self.get_column_index("Reference Display"),
            "InstrumentID": self.get_column_index("InstrumentID"),
            "Instrument Name": self.get_column_index("Instrument Name"),
            "LabFacilityID": self.get_column_index("LabFacilityID"),
            "Lab Facility Name": self.get_column_index("Lab Facility Name"),
            "UPbAnalysisMethodID": self.get_column_index("UPbAnalysisMethodID"),
            "UPb Analysis Method Name": self.get_column_index("UPb Analysis Method Name"),
        }

        column = field_to_column.get(field)
        if column is None:
            self.add_column(field=field)

        # Update all rows in the column
        self.right_table.blockSignals(True)

        field_to_column = {
            "ReferenceID": self.get_column_index("ReferenceID"),
            "Reference Display": self.get_column_index("Reference Display"),
            "InstrumentID": self.get_column_index("InstrumentID"),
            "Instrument Name": self.get_column_index("Instrument Name"),
            "LabFacilityID": self.get_column_index("LabFacilityID"),
            "Lab Facility Name": self.get_column_index("Lab Facility Name"),
            "UPbAnalysisMethodID": self.get_column_index("UPbAnalysisMethodID"),
            "UPb Analysis Method Name": self.get_column_index("UPb Analysis Method Name"),
        }

        if field == "Reference Display":
            id_name = "ReferenceID"
        elif field == "Instrument Name":
            id_name = "InstrumentID"
        elif field == "Lab Facility Name":
            id_name = "LabFacilityID"
        elif field == "UPb Analysis Method Name":
            id_name = "UPbAnalysisMethodID"
        else:
            id_name = None

        id_column = field_to_column.get(id_name)
        name_column = field_to_column.get(field)

        self.right_table.blockSignals(True)
        for row in range(self.right_table.rowCount()):
            id_item = self.right_table.item(row, id_column)
            if id_item is None:
                self.right_table.setItem(row, id_column, QTableWidgetItem(str(checked_item_id)))
            else:
                id_item.setText(str(checked_item_id))

        for row in range(self.right_table.rowCount()):
            id_item = self.right_table.item(row, name_column)
            if id_item is None:
                self.right_table.setItem(row, name_column, QTableWidgetItem(str(checked_item_name)))
            else:
                id_item.setText(str(checked_item_name))

        self.right_table.blockSignals(False)
        # self.right_table.hideColumn(id_column)
        QMessageBox.information(self, "Success", f"All rows updated with '{str(checked_item_name)}' for {field}.")

    def get_column_index(self, header_name):
        """
        Get the column index of a header by its name.
        Args:
            header_name (str): The name of the header.
        Returns:
            int: The column index, or None if not found.
        """
        for col in range(self.right_table.columnCount()):
            header_item = self.right_table.horizontalHeaderItem(col)
            if header_item and header_item.text().startswith(header_name):
                return col
        return None

    def get_column_name(self, column_index):
        return self.right_table.horizontalHeaderItem(column_index)

    def update_row_icon(self, row_idx, rejected):
        header_item = QTableWidgetItem()
        header_item.setText(str(row_idx + 1))
        if rejected:
            header_item.setIcon(self.rejected_icon)
        else:
            header_item.setIcon(self.accepted_icon)
        self.right_table.setVerticalHeaderItem(row_idx, header_item)

    # Existing logic for removing rows, marking them rejected, etc.

    def remove_selected_rows(self, row=None):
        if row is None:
            selected_rows = {i.row() for i in self.right_table.selectedItems()}
            if not selected_rows:
                return
        else:
            selected_rows = {row}

        sr = sorted(selected_rows, reverse=True)
        if self.df is not None and len(self.df) > 0:
            self.df.drop(self.df.index[sr], inplace=True)
            self.df.reset_index(drop=True, inplace=True)

        for r in sr:
            self.right_table.removeRow(r)
            self.left_table.removeRow(r)
            self.rejected_rows.discard(r)
        self.update_vertical_headers()

    def remove_selected_columns(self):
        """
        Remove selected columns from the right table, the column mappings,
        and associated data structures.
        """
        selected_columns = {i.column() for i in self.right_table.selectedItems()}
        if not selected_columns:
            return

        # Sort selected columns in descending order to remove from the rightmost column
        sc = sorted(selected_columns, reverse=True)

        for col in sc:
            # Remove the column from the right table
            self.right_table.removeColumn(col)

            # Update column mappings to reflect the removed column
            if col in self.column_mappings:
                del self.column_mappings[col]

            # Shift column mappings for columns after the removed one
            self.column_mappings = {
                (idx - 1 if idx > col else idx): value
                for idx, value in self.column_mappings.items()
            }

        # Notify the user
        QMessageBox.information(self, "Columns Removed", "Selected columns have been successfully removed.")

    def update_vertical_headers(self):
        """
        Update the vertical headers to ensure they match the current row indices.
        """
        row_count = self.right_table.rowCount()
        for row_idx in range(row_count):
            header_item = QTableWidgetItem(str(row_idx + 1))  # Update row numbers
            # Check if the row is rejected and set the appropriate icon
            if row_idx in self.rejected_rows:
                header_item.setIcon(self.rejected_icon)
            else:
                header_item.setIcon(self.accepted_icon)
            self.right_table.setVerticalHeaderItem(row_idx, header_item)

    def mark_selected_rows_rejected(self, rows: list[QTableWidgetItem],  rejected: bool):

        selected_rows = {i.row() for i in rows}
        if not selected_rows:
            return

        for r in selected_rows:
            if rejected:
                self.rejected_rows.add(r)
            else:
                self.rejected_rows.discard(r)
            self.update_row_icon(r, rejected)

    # ---------------------------
    #     Header Double Click
    # ---------------------------

    def handle_header_double_clicked(self, logical_index):
        """
        Double-click on a right table header => open mapping dialog.
        """
        item = self.right_table.horizontalHeaderItem(logical_index)
        if not item:
            return
        original_header_text = item.text()
        curr_map = self.column_mappings.get(logical_index, ("None"))
        dialog = ColumnMapDialog(original_header_text, curr_map, self)
        if dialog.exec():
            new_field = dialog.get_selected_value()
            if new_field == "None" or not new_field:
                if logical_index in self.column_mappings:
                    del self.column_mappings[logical_index]
                item.setText(original_header_text)
                item.setBackground(QBrush(Qt.GlobalColor.transparent))
            else:
                self.column_mappings[logical_index] = (new_field)
                item.setText(f"{new_field}")
                item.setBackground(QBrush(Qt.GlobalColor.green))

                # If it’s Sample Name / Aliquot Name / Spot Name, auto-populate left table
                if new_field in ["Sample Name", "Aliquot Name", "Spot Name"]:
                    self.update_left_table_on_header_change(new_field, logical_index)

    def update_left_table_on_header_change(self, field, logical_index):
        if field == "Sample Name":
            for r in range(self.right_table.rowCount()):
                cell_item = self.right_table.item(r, logical_index)
                if not cell_item:
                    continue
                sample_id_value = cell_item.text().strip()

                # Update the left table
                self.left_table.blockSignals(True)
                self.left_table.setItem(r, 0, QTableWidgetItem(sample_id_value))  # Sample ID
                self.left_table.blockSignals(False)
            self.left_table.resizeColumnsToContents()
        elif field == "Aliquot Name":
            for r in range(self.right_table.rowCount()):
                cell_item = self.right_table.item(r, logical_index)
                if not cell_item:
                    continue
                aliquot_id_value = cell_item.text().strip()

                # Update the left table
                self.left_table.blockSignals(True)
                self.left_table.setItem(r, 1, QTableWidgetItem(aliquot_id_value))  # Aliquot ID
                self.left_table.blockSignals(False)
            self.left_table.resizeColumnsToContents()

        elif field == "Spot Name":
            self.auto_split_sample_spot(logical_index)

    def update_left_table_on_delimiter_change(self):
        """
        Update the left table's Sample ID and Spot ID columns whenever the delimiter value changes.
        """
        # Find the right table column mapped to "Spot ID"
        # todo this is breaking when sample name and spot name are both set

        if self.delimiter_checkbox.isChecked():
            spot_id_column = None
            for col_idx, (field_name) in self.column_mappings.items():
                if field_name == "Spot Name":
                    spot_id_column = col_idx
                    break

            if spot_id_column is not None:
                self.auto_split_sample_spot(spot_id_column)
        else:
            spot_id_column = None
            for col_idx, (field_name) in self.column_mappings.items():
                if field_name == "Spot Name":
                    spot_id_column = col_idx
                    break

            if spot_id_column is not None:
                row_count = self.right_table.rowCount()
                for r in range(row_count):
                    cell_item = self.right_table.item(r, col_idx)
                    if not cell_item:
                        continue

                    spot_id_value = cell_item.text().strip()

                    # Update the left table
                    self.left_table.blockSignals(True)
                    self.left_table.setItem(r, 0, QTableWidgetItem(""))
                    self.left_table.setItem(r, 2, QTableWidgetItem(spot_id_value))  # Spot ID
                    self.left_table.blockSignals(False)
        self.left_table.resizeColumnsToContents()

    def auto_split_sample_spot(self, col_idx):
        """
        Split the right table's Spot ID column values into Sample ID and Spot ID
        using the delimiter, and populate the left table accordingly.
        """
        delimiter = self.delimiter_edit.text().strip()
        row_count = self.right_table.rowCount()

        for r in range(row_count):
            cell_item = self.right_table.item(r, col_idx)
            if not cell_item:
                continue

            spot_id_value = cell_item.text().strip()

            if delimiter in spot_id_value and delimiter:
                # Split based on the delimiter
                sample_id, spot_id = spot_id_value.split(delimiter, 1)
            else:
                # No delimiter found, treat the entire value as Spot ID
                sample_id = ""
                spot_id = spot_id_value

            # Update the left table
            self.left_table.blockSignals(True)
            self.left_table.setItem(r, 0, QTableWidgetItem(sample_id))  # Sample ID
            self.left_table.setItem(r, 2, QTableWidgetItem(spot_id))  # Spot ID
            self.left_table.blockSignals(False)
        self.left_table.resizeColumnsToContents()

    def save_mapping(self):
        if not self.column_mappings:
            QMessageBox.warning(self, "No Mapping", "No columns have been mapped yet.")
            return

        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    try:
                        configs = json.load(f)
                    except json.JSONDecodeError:
                        # If the file is empty, it cannot load
                        configs = {}
            else:
                configs = {}
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save mapping:\n{e}")

        if configs == {}:
            name, ok = QInputDialog.getText(self, "Save Mapping", "Enter a name for this mapping:")
        else:
            items = list(configs.keys())
            if not items:
                name, ok = QInputDialog.getText(self, "Save Mapping", "Enter a name for this mapping:")
            else:
                dlg = CompleterInputDialog(self, "Save Mapping", "Enter or select a name for this mapping:", items, True)
                if dlg.exec() == QDialog.DialogCode.Accepted:
                    name = dlg.get_input()
                    if not name:
                        return
                    else:
                        if name in items:
                            ok = True
                            reply = QMessageBox.question(self, "Overwrite Mapping",
                                                         f"Mapping '{name}' already exists. Overwrite?",
                                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                                         QMessageBox.StandardButton.No)
                            if reply != QMessageBox.StandardButton.Yes:
                                return
                else:
                    return

        if ok and name:
            jmap = {str(k): {"field": v} for k, v in self.column_mappings.items()}
            configs[name] = jmap
            with open(CONFIG_FILE, 'w') as f:
                json.dump(configs, f, indent=4)
            QMessageBox.information(self, "Saved", f"Mapping '{name}' saved successfully.")


    def load_mapping(self):
        if not os.path.exists(CONFIG_FILE):
            QMessageBox.warning(self, "No Config", "No configuration file found.")
            return

        try:
            with open(CONFIG_FILE, 'r') as f:
                configs = json.load(f)
            if not configs:
                QMessageBox.warning(self, "No Mappings", "No mappings found in configuration.")
                return

            items = list(configs.keys())
            name, ok = QInputDialog.getItem(self, "Load Mapping", "Select a mapping to load:", items, 0, False)
            if ok and name:
                loaded = configs[name]
                self.column_mappings.clear()
                for k_str, v in loaded.items():
                    idx = int(k_str)
                    self.column_mappings[idx] = (v["field"])

                total_cols = self.right_table.columnCount()
                for col_idx in range(total_cols):
                    hdr_item = self.right_table.horizontalHeaderItem(col_idx)
                    if not hdr_item:
                        continue
                    if col_idx in self.column_mappings:
                        f_name = self.column_mappings[col_idx]
                        hdr_item.setText(f"{f_name}")
                        hdr_item.setBackground(QBrush(QColor("#ffffcc")))
                        # If it’s Sample Name / Aliquot Name / Spot Name, auto-populate left table
                        if f_name in ["Sample Name", "Aliquot Name", "Spot Name"]:
                            self.update_left_table_on_header_change(f_name, col_idx)
                    else:
                        hdr_item.setBackground(QBrush(Qt.GlobalColor.transparent))

                QMessageBox.information(self, "Loaded", f"Mapping '{name}' loaded successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load mapping:\n{e}")
            return
        self.right_table.resizeColumnsToContents()

    # todo: allow deleting mappings or changing name

    def check_empty_cells_in_left_table(self):
        empty_cells = []
        for row in range(self.left_table.rowCount()):
            for col in range(3):  # Columns for Sample ID, Aliquot ID, Spot ID
                cell = self.left_table.item(row, col)
                if cell is None or cell.text().strip() == "":
                    empty_cells.append((row, col))
        return empty_cells

    def ask_to_use_default_values(self, empty_cells):
        """
        Prompt the user to confirm whether to use autogenerated default values for empty cells.
        """
        missing_count = len(empty_cells)
        msg = f"{missing_count} cells are empty. Would you like to use autogenerated default values?"
        reply = QMessageBox.question(self, 'Missing Values', msg,
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.Yes)

        if reply == QMessageBox.StandardButton.Yes:
            return True
        else:
            return False

    def fill_empty_cells_with_defaults(self, empty_cells):
        """
        Fill the empty cells in the left table with default values.
        """
        self.left_table.blockSignals(True)

        # Initialize variables for tracking SampleID and counter
        current_aliquot_id = None
        spot_counter = 0
        for row, col in empty_cells:
            if col == 0:  # Sample Name
                # If Sample ID is missing, set a default value
                if not self.left_table.item(row, col) or not self.left_table.item(row, col).text().strip():
                    self.left_table.setItem(row, col, QTableWidgetItem("DefaultSample"))

            elif col == 1:  # Aliquot Name
                # If Aliquot Name is missing, set equal to Sample Name value
                if not self.left_table.item(row, col) or not self.left_table.item(row, col).text().strip():
                    self.left_table.setItem(row, col, QTableWidgetItem(self.left_table.item(row, col - 1).text().strip()))

            elif col == 2:  # Spot ID
                # If Aliquot Name exists, create Spot Name with the counter
                aliquot_id_item = self.left_table.item(row, col - 1)
                if aliquot_id_item and aliquot_id_item.text().strip():
                    aliquot_id = aliquot_id_item.text().strip()
                    if aliquot_id != current_aliquot_id:
                        current_aliquot_id = aliquot_id
                        spot_counter = 0  # Reset counter for new Aliquot Name
                    spot_counter += 1
                    self.left_table.setItem(row, col, QTableWidgetItem(f"{aliquot_id}-{spot_counter}"))

            # Highlight the updated cell
            self.left_table.item(row, col).setBackground(Qt.GlobalColor.yellow)

        self.left_table.blockSignals(False)
        self.left_table.resizeColumnsToContents()

    def check_and_import(self):
        # Step 1: Check for empty cells in the left table
        empty_cells = self.check_empty_cells_in_left_table()

        if empty_cells:
            # Step 2: Show dialog to ask if the user wants to use default values
            use_defaults = self.ask_to_use_default_values(empty_cells)

            if use_defaults:
                # Step 3: Fill empty cells with default values
                self.fill_empty_cells_with_defaults(empty_cells)

                # Optional: Give user a chance to review and adjust the values
                QMessageBox.information(self, "Review",
                                        "The empty cells have been filled with default values. Please review before clicking import again.")
        else:
            # Step 4: Look for any existing analyses associated with existing spots and ask user how to handle conflicts
            if self.check_for_conflicts():
                # Step 5: Proceed with import
                self.import_to_db()

    def check_for_conflicts(self):
        upb_matches = []
        for row_idx in range(self.left_table.rowCount()):
            sample_name = self.left_table.item(row_idx, 0).data(Qt.ItemDataRole.DisplayRole)
            existing_sample = self.find_matching_id('Samples', 'SampleName', sample_name)
            if existing_sample:
                aliquot_name = self.left_table.item(row_idx, 1).data(Qt.ItemDataRole.DisplayRole)
                existing_aliquot = self.find_matching_id('Aliquots', 'AliquotName', aliquot_name)
                if existing_aliquot:
                    spot_name = self.left_table.item(row_idx, 2).data(Qt.ItemDataRole.DisplayRole)
                    existing_spot = self.find_matching_id('Spots', 'SpotName', spot_name)
                    if existing_spot:
                        query = QSqlQuery()
                        query.prepare('SELECT * FROM UPbAnalyses WHERE SpotID=:spot_id')
                        query.bindValue(":spot_id", existing_spot)
                        if query.exec():
                            if query.next():
                                upb_matches.append((sample_name, aliquot_name, spot_name))
        if upb_matches:
            msg = f"{len(upb_matches)}/{self.left_table.rowCount()} existing UPb analyses found for spots being imported."
            msg_box = QMessageBox()
            msg_box.setWindowTitle("Existing Values")
            msg_box.setText(msg)
            msg_box.addButton(QMessageBox.StandardButton.Cancel)
            skip_button = QPushButton("Skip duplicates")
            msg_box.addButton(skip_button, QMessageBox.ButtonRole.NoRole)
            overwrite_button = QPushButton("Overwrite duplicates")
            msg_box.addButton(overwrite_button, QMessageBox.ButtonRole.YesRole)
            msg_box.setDefaultButton(QMessageBox.StandardButton.Cancel)
            msg_box.exec()
            if msg_box.clickedButton() == QMessageBox.StandardButton.Cancel:
                return False
            elif msg_box.clickedButton() == skip_button:
                self.conflict_mode = 'skip'
                return True
            elif msg_box.clickedButton() == overwrite_button:
                self.conflict_mode = 'overwrite'
                return True

    # ---------------------------
    #      Import to DB
    # ---------------------------

    def import_to_db(self):
        row_count = self.right_table.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "No Data", "There are no rows to import.")
            return
        else:
            # Create a modal progress dialog
            progress_dialog = QProgressDialog(
                "Importing data...", "Cancel", 0, row_count, self
            )

        create_savepoint('before_upb_import')

        inserted_count = 0
        try:
            for row_idx in range(row_count):
                progress_dialog.setValue(row_idx + 1)
                # Let the event loop process the dialog's updates
                QApplication.processEvents()
                # If the user clicked "Cancel", we can break out
                if progress_dialog.wasCanceled():
                    rollback_savepoint('before_upb_import')
                    break

                # Build a record dict with every key initialized to None
                record = {field: None for field in SQLUtils.upb_possible_database_input_fields}
                if row_idx in self.rejected_rows:
                    record['Rejected'] = True
                else:
                    record['Rejected'] = False

                record['RatioErrorFormatID'] = self.ratio_error_combobox.itemData(self.ratio_error_combobox.currentIndex())
                record['AgeErrorFormatID'] = self.age_error_combobox.itemData(self.age_error_combobox.currentIndex())
                record['ConcordanceFormatID'] = self.conc_error_combobox.itemData(self.conc_error_combobox.currentIndex())
                record['AgeUnitID'] = self.age_unit_combobox.itemData(self.age_unit_combobox.currentIndex())
                record['SpotSizeUnitID'] = self.spot_size_unit_combobox.itemData(self.spot_size_unit_combobox.currentIndex())


                # Populate the left-table items (sample_id, aliquot_id, spot_id)
                sample_id_item = self.left_table.item(row_idx, 0)
                aliquot_id_item = self.left_table.item(row_idx, 1)
                spot_id_item = self.left_table.item(row_idx, 2)


                record["Sample Name"] = sample_id_item.text().strip() if sample_id_item else None
                record["Aliquot Name"] = aliquot_id_item.text().strip() if aliquot_id_item else None
                record["Spot Name"] = spot_id_item.text().strip() if spot_id_item else None

                # Find matching SampleID or create new
                sample_query = QSqlQuery()
                sample_query.prepare(f"SELECT SampleID FROM Samples WHERE SampleName=:name COLLATE NOCASE")
                sample_query.bindValue(":name", record["Sample Name"])

                if sample_query.exec():
                    if sample_query.next():
                        # found matching samplename in database, will use that sample ID
                        record["SampleID"] = sample_query.value(0)
                        self.sample_ids.append(record["SampleID"])
                    else:
                        # no matching samplename in database, will create new one.
                        create_sample = QSqlQuery()
                        create_sample.prepare('INSERT INTO Samples (SampleName) VALUES (:name)')
                        create_sample.bindValue(":name", record["Sample Name"])

                        if not create_sample.exec():
                            print("Failed to execute query:", create_sample.lastError().text())
                        else:
                            record["SampleID"] = create_sample.lastInsertId()
                            self.sample_ids.append(record["SampleID"])
                else:
                    print("Failed to execute query:", sample_query.lastError().text())


                # Find matching Aliquot Name or create new
                aliquot_query = QSqlQuery()
                aliquot_query.prepare('SELECT AliquotID FROM Aliquots WHERE AliquotName=:name COLLATE NOCASE AND SampleID=:sample_id')
                aliquot_query.bindValue(":name", record["Aliquot Name"])
                aliquot_query.bindValue(":sample_id", record["SampleID"])

                if aliquot_query.exec():
                    if aliquot_query.next():
                        # found matching aliquot name in database, will use that aliquot ID
                        record["AliquotID"] = aliquot_query.value(0)
                    else:
                        # no matching samplename in database, will create new one.
                        create_aliquot = QSqlQuery()
                        create_aliquot.prepare('INSERT INTO Aliquots (AliquotName, SampleID) VALUES (:name, :sample_id)')
                        create_aliquot.bindValue(":name", record["Aliquot Name"])
                        create_aliquot.bindValue(":sample_id", record["SampleID"])

                        if not create_aliquot.exec():
                            print("Failed to create_aliquot execute query:", create_aliquot.lastError().text())
                        else:
                            record["AliquotID"] = create_aliquot.lastInsertId()
                else:
                    print("Failed to select_aliquot execute query:", aliquot_query.lastError().text())

                # Find matching SpotID or create new
                spot_query = QSqlQuery()
                spot_query.prepare(
                    'SELECT SpotID FROM Spots WHERE SpotName=:name COLLATE NOCASE AND AliquotID=:aliquot_id')
                spot_query.bindValue(":name", record["Spot Name"])
                spot_query.bindValue(":aliquot_id", record["AliquotID"])

                if spot_query.exec():
                    if spot_query.next():
                        # found matching spot name in database, will use that spot ID
                        record["SpotID"] = spot_query.value(0)
                    else:
                        # no matching spot name in database, will create new one.
                        create_spot = QSqlQuery()

                        create_spot.prepare(
                            'INSERT INTO Spots (SpotName, AliquotID) VALUES (:name, :aliquot_id)')
                        create_spot.bindValue(":name", record["Spot Name"])
                        create_spot.bindValue(":aliquot_id", record["AliquotID"])

                        if not create_spot.exec():
                            print("Failed to execute query:", create_spot.lastError().text())
                        else:
                            record["SpotID"] = create_spot.lastInsertId()
                else:
                    print("Failed to execute query:", spot_query.lastError().text())




                # by this point a valid Sample, Aliquot, and Spot should be created.
                # Check if the spot already has a UPbAnalysis, if so, skip or overwrite
                upb_match = False
                upb_query = QSqlQuery()
                upb_query.prepare('SELECT UPbAnalysisID FROM UPbAnalyses WHERE SpotID=:spot_id')
                upb_query.bindValue(":spot_id", record["SpotID"])
                if upb_query.exec():
                    if upb_query.next():
                        upb_match = True

                if upb_match and self.conflict_mode == 'skip':
                    continue
                elif upb_match and self.conflict_mode == 'overwrite':
                    # delete existing UPbAnalysis
                    delete_query = QSqlQuery()
                    delete_query.prepare('DELETE FROM UPbAnalyses WHERE SpotID=:spot_id')
                    delete_query.bindValue(":spot_id", record["SpotID"])
                    if not delete_query.exec():
                        print("Failed to execute query:", delete_query.lastError().text())


                field_names = ", ".join([f'[{field}]' for field in SQLUtils.upb_possible_database_input_fields])

                placeholders = ', '.join(
                    [f':{field.replace('/', '').replace('*', '').replace(' ', '_')}' for field in
                     SQLUtils.upb_possible_database_input_fields])
                insert_sql = f"""
                                            INSERT INTO UPbAnalyses (
                                                {field_names}
                                            )
                                            VALUES (
                                                {placeholders}
                                            )
                                        """

                print(insert_sql)

                # Process the main columns from the mapping.
                # In your code, you might reduce main_cols by 4 if these appended columns
                # are always the last 4 columns. Adjust logic as needed.
                main_cols = self.right_table.columnCount()
                for col_idx in range(main_cols):
                    if col_idx not in self.column_mappings:
                        continue

                    field_name = self.column_mappings[col_idx]
                    if field_name == "None" or field_name in ('Sample Name', 'Aliquot Name', 'Spot Name'):
                        continue

                    cell_item = self.right_table.item(row_idx, col_idx)
                    if not cell_item:
                        continue

                    cell_text = cell_item.text().strip()
                    if cell_text.upper() == "NULL":
                        record[field_name] = None
                    else:
                        record[field_name] = cell_text

                # Finally insert the row
                insert_query = QSqlQuery()

                if not insert_query.prepare(insert_sql):
                    print("Failed to execute prepare:", insert_query.lastError().text())
                record_count = 0
                for key, value in record.items():
                    if key == "Sample Name" or key == "SampleID" or key == "Aliquot Name" or key == "AliquotID" or key == "Spot Name":
                        continue
                    insert_query.bindValue(f":{key.replace('/', '').replace('*', '').replace(' ', '_')}", value)
                    record_count += 1
                if not insert_query.exec():
                    print(f"Error executing query: {insert_query.lastError().text()}")
                    print(insert_query.executedQuery())
                    for value in insert_query.boundValues():
                        print(value)

                # Find matching Rejection Reasons, unique collate no case rejection reasons table
                # if found utilize that ID, else create one, then create association in many-to-many table

                record['UPbAnalysisID'] = insert_query.lastInsertId()
                if 'Rejection Reason' in record:
                    if record['Rejection Reason'] is not None:
                        # Find matching Rejection Reasons, unique collate no case rejection reasons table
                        # if found utilize that ID, else create one, then create assoication in many-to-many table
                        # Find matching SpotID or create new
                        rejection_query = QSqlQuery()
                        rejection_query.prepare(
                            'SELECT RejectionReasonID FROM RejectionReasons WHERE RejectionReasonName=:name COLLATE NOCASE')
                        rejection_query.bindValue(":name", record["Rejection Reason"])

                        if rejection_query.exec():
                            if rejection_query.next():
                                # found matching spot name in database, will use that spot ID
                                record["RejectionReasonID"] = rejection_query.value(---0)
                            else:
                                # no matching samplename in database, will create new one.
                                create_rejection = QSqlQuery()
                                create_rejection.prepare(
                                    'INSERT INTO RejectionReasons (RejectionReasonName) VALUES (:name)')
                                create_rejection.bindValue(":name", record["Rejection Reason"])

                                if not create_rejection.exec():
                                    print("Failed to execute query:", create_rejection.lastError().text())
                                else:
                                    record["RejectionReasonID"] = create_rejection.lastInsertId()

                                # no matching samplename in database, will create new one.
                                create_upb_rejection_assoc = QSqlQuery()
                                create_upb_rejection_assoc.prepare(
                                    'INSERT INTO UPbAnalyses_RejectionReasons (UPbAnalysisID, RejectionReasonID) VALUES (:upb_analysis_id, :rejection_reason_id)')
                                create_upb_rejection_assoc.bindValue(":upb_analysis_id", record["UPbAnalysisID"])
                                create_upb_rejection_assoc.bindValue(":rejection_reason_id", record["RejectionReasonID"])

                                if not create_upb_rejection_assoc.exec():
                                    print("Failed to execute query:", create_upb_rejection_assoc.lastError().text())

                        else:
                            print("Failed to execute query:", rejection_query.lastError().text())

                inserted_count += 1

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to import data:\n{e}")
            rollback_savepoint('before_upb_import')
        QSqlDatabase().commit()
        release_savepoint('before_upb_import')
        QMessageBox.information(self, "Success", f"Imported {inserted_count} rows into the database.")
        update_database()
        self.data_imported.emit(self.sample_ids)
        self.close()

    def find_matching_id(self, table, field_name, value):
        query = QSqlQuery()
        id_field = f"{table.strip('s')}ID"
        query.prepare(f"SELECT {id_field} FROM {table} WHERE {field_name}=:value COLLATE NOCASE")
        query.bindValue(":value", value)
        if query.exec():
            if query.next():
                return query.value(0)
        return None

    def close(self):
        self.saveWindowState()
        return super().close()

    def saveWindowState(self):
        settings.setValue("ui/ImportWizard/pos", self.pos())
        settings.setValue("ui/ImportWizard/size", self.size())

    def loadWindowState(self):
        self.move(settings.value("ui/ImportWizard/pos", defaultValue=QPoint(410, 241)))
        self.resize(settings.value("ui/ImportWizard/size", defaultValue=QSize(810, 569)))


if __name__ == "__main__":

    app = QApplication(sys.argv)
    window = ImportWizardDialog()
    window.show()
    sys.exit(app.exec())