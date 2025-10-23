import collections
import json
import os
import time

import pandas as pd
import qtawesome
from PyQt6 import QtCore
from PyQt6.QtCore import Qt, QPoint, QSize, QStringListModel, QRect, QVariant, QModelIndex, QAbstractTableModel
from PyQt6.QtGui import QBrush, QColor, QFont, QAction, QPalette, QIcon
from PyQt6.QtSql import QSqlDatabase, QSqlQuery, QSqlTableModel
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel, QTableView,
    QComboBox, QTableWidget, QTableWidgetItem, QMessageBox, QHBoxLayout, QComboBox,
    QLineEdit, QInputDialog, QMenu, QDialog, QFormLayout, QSplitter, QAbstractItemView, QCheckBox,
    QProgressDialog, QListWidget, QListView, QDialogButtonBox, QTabWidget, QSpacerItem, QSizePolicy
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import logger_setup
from Functions import SQLUtils
from Functions.Check_triggers import validate_insert, validate_update
from Functions.Database_manager import update_database
from Functions.LoadingDialog_manager import LoadingDialogManager
from Functions.Savepoint_manager import create_savepoint, rollback_savepoint, release_savepoint
from Functions.Settings_manager import SettingsManager
from ui.EditView import EditView

settings = SettingsManager().settings
from Functions.Widget_classes import (
    CheckableComboBox, CheckableSqlTableModel, SearchableComboBox, set_table, CheckableTreeModel,
    CheckableTreeCombobox, save_expanded_state, get_name_column, add_tree_popup, get_id_from_name, get_headers,
    CheckableSqlQueryModel, SQLiteTableModel, CompleterInputDialog, get_table_from_view, get_view_from_table,
    search_dictionary, ImportSheetModel, loading_manager)
from Functions.Database_views import ViewQuery
from ui.AddTags import AddTags
from ui.AddTreeTags import AddTreeTags
from ui.EditTable import EditTable
from ui.EditTree import EditTree
from ui.New_reference import NewReference

CONFIG_FILE = 'column_mappings.json'


class ColumnMapDialog(QDialog):
    """
    Class to load a helper dialog to assist the user in selecting pre-defined values of columns to a known database column.
    The list of available categories and columns is defined by SQLUtils.upb_possible_user_input_fields.
    """

    def __init__(self, original_header: str, current_field:str, parent: QWidget):
        """
        Creates a ColumnMapDialog instance with the original text of the header, typically a number, the current field,
        if defined, and a parent widget.

        :param str original_header:
        :param current_field:
        :param parent:
        """
        super().__init__(parent)
        self.setWindowTitle(f"Column Mapper {original_header}")

        # Keep track of combo boxes so we can manipulate them easily
        self.combos = []
        self._is_updating = False

        form_layout = QFormLayout()
        tab_widget = QTabWidget()
        form_layout.addRow(tab_widget)

        # Add a new tab for each category of fields
        self.field_dictionaries = [SQLUtils.sample_possible_user_input_fields,
                              SQLUtils.gps_possible_user_input_fields,
                              SQLUtils.column_possible_user_input_fields,
                              SQLUtils.aliquot_grain_spot_possible_user_input_fields,
                              SQLUtils.reference_possible_user_input_fields,
                              SQLUtils.upb_possible_user_input_fields]

        tab_names = ["Sample Info", "GPS Info", "Column Info", "Aliquot/Grain/Spot Info", "Reference", "U-Pb Data"]
        for tab_name, field_dict in zip(tab_names, self.field_dictionaries):
            tab = QWidget()
            tab_layout = QFormLayout()
            tab.setLayout(tab_layout)
            tab_widget.addTab(tab, tab_name)

            for field_label, possible_values in field_dict.items():
                combo = SearchableComboBox()
                combo.addItem("None")
                combo.addItems(possible_values.keys())
                combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
                combo.selection_changed.connect(self.on_combo_changed)
                combo.currentIndexChanged.connect(self.on_combo_changed)
                tab_layout.addRow(field_label + ":", combo)
                self.combos.append(combo)

        if current_field is not None and current_field != "None":
            self.on_combo_changed()

        # Add an OK button for closing
        self.btn_ok = QPushButton("OK")
        self.btn_ok.clicked.connect(self.accept)

        # Wrap everything in a layout
        main_layout = QVBoxLayout()
        main_layout.addLayout(form_layout)
        main_layout.addWidget(self.btn_ok)
        self.setLayout(main_layout)

        # Set the current tab to U-Pb Data
        tab_widget.setCurrentIndex(len(tab_names) - 1)


    def on_combo_changed(self):
        """
        Triggered whenever the current index of any ComboBox changes.
        Ensures only one combo has a non-'None' value at a time.
        """
        if self._is_updating:
            return

        # Avoid recursion / repeated signals while we're updating
        self._is_updating = True

        # Figure out which combo box triggered
        triggered_combo = self.sender()
        if not isinstance(triggered_combo, QComboBox):
            # Should never happen in this example, but just a guard
            self._is_updating = False
            return

        # If the user selected something other than 'None',
        # reset all other combos to 'None'
        if triggered_combo.currentIndex() != 0:  # i.e., not the "None" entry
            for combo in self.combos:
                if combo is not triggered_combo:
                    combo.setCurrentIndex(0)

        self._is_updating = False

    def get_selected_value(self):
        """
        Returns which combo was selected (if any), or 'None' if all are 'None'.
        You can customize how you want this data returned.
        """
        for combo in self.combos:
            if combo.currentIndex() != 0 and combo.currentIndex() != -1:  # i.e., not "None" or no selection
                return combo.currentText()
        return "None"

    def saveWindowState(self):
        settings.setValue("ui/ImportWizard/ColumnMappingDialog/pos", self.pos())
        settings.setValue("ui/ImportWizard/ColumnMappingDialog/size", self.size())

    def loadWindowState(self):
        self.move(settings.value("ui/ImportWizard/ColumnMappingDialog/pos", defaultValue=QPoint(410, 241)))
        self.resize(settings.value("ui/ImportWizard/ColumnMappingDialog/size", defaultValue=QSize(810, 569)))


class LoadMappingDialog(QDialog):
    """
    Class to load a helper dialog to assist the user in selecting previously saved column mappings to the column map dictionary.
    """
    def __init__(self, parent: QWidget) -> None:
        """
        Creates a LoadMappingDialog instance with a defined parent widget
        :param parent:
        """
        super().__init__(parent)

        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    try:
                        configs = json.load(f)
                    except json.JSONDecodeError:
                        # If the file is empty, it cannot load
                        configs = {}
                        return
            else:
                configs = {}
                return
        except Exception as e:
            logger_setup.get_logger().critical(f'Failed to save mapping: {e}')
            return

        self.configs = configs
        self.selected_name = ''
        self.list_model = QStringListModel()
        self.setModal(True)
        self.setWindowTitle("Load Mapping")
        self.layout = QVBoxLayout(self)
        self.list_view = QListView()
        self.list_view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.list_view.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.list_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.text_label = QLabel()
        self.text_label.setText("Select a mapping to load:")
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.layout.addWidget(self.text_label)
        self.layout.addWidget(self.list_view)
        self.layout.addWidget(self.button_box)
        self.setLayout(self.layout)

        self.display_mapping_list()

        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.list_view.customContextMenuRequested.connect(self.show_context_menu)
        self.list_view.selectionModel().selectionChanged.connect(self.update_selected_name)

    def display_mapping_list(self) -> None:
        """
        Gathers and displays a list of recent mappings from SettingsManager to the QListView.
        """
        items = list(self.configs.keys())
        recent_mappings = settings.value("recent_mappings", [])
        items_sorted = []
        if recent_mappings:
            for name in recent_mappings:
                if name in items:
                    items_sorted.append(name)
        for name in items:
            if name not in items_sorted:
                items_sorted.append(name)
        self.list_model.setStringList(items_sorted)
        self.list_view.setModel(self.list_model)

    def update_selected_name(self):
        """
        Updates selected_name to the currently selected value in the QListView
        """
        index = self.list_view.currentIndex()
        if not index.isValid():
            self.selected_name = ''
            return
        self.selected_name = index.data()

    def show_context_menu(self, pos):
        """
        Shows the context menu at a given position. Pos is automatically passed when called from a signal/slot combo.
        :param pos: Position of the context menu to be displayed at.
        """
        menu = QMenu()
        menu.addAction("Rename", self.rename_mapping)
        menu.addAction("Delete", self.delete_mapping)
        menu.exec(self.list_view.mapToGlobal(pos))

    def rename_mapping(self):
        """
        Helper method to open dialogs to rename a given mapping.
        """
        index = self.list_view.currentIndex()
        if not index.isValid():
            return
        name = index.data()
        new_name, ok = QInputDialog.getText(self, "Rename", "Enter a new name for the mapping:", text=name)
        if ok:
            mapping = self.configs[name]
            del self.configs[name]
            self.configs[new_name] = mapping
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.configs, f, indent=4)
            items = settings.value("recent_mappings", [])
            items.remove(name)
            items.insert(index, new_name)
            settings.setValue("recent_mappings", items)
            QMessageBox.information(self, "Edited", f"Mapping '{name}' renamed successfully.")
            self.display_mapping_list()

    def delete_mapping(self):
        """
        Helper method to open dialogs to delete a given mapping.
        """
        index = self.list_view.currentIndex()
        if not index.isValid():
            return
        name = index.data()
        response = QMessageBox.question(self, "Delete Mapping",
                                        f"Are you sure you want to delete the mapping '{name}'?",
                                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if response == QMessageBox.StandardButton.Yes:
            del self.configs[name]
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.configs, f, indent=4)
            items = settings.value("recent_mappings", [])
            if name in items:
                items.remove(name)
            settings.setValue("recent_mappings", items)
            QMessageBox.information(self, "Deleted", f"Mapping '{name}' deleted successfully.")
            self.display_mapping_list()

    def accept(self):
        index = self.list_view.currentIndex()
        if not index.isValid():
            logger_setup.get_logger().error("Must select a mapping")
            return
        super().accept()


class ImportWizardDialog(QWidget):
    """
    Main dialog used for assisting the user in importing UPbAnalyses into the database.
      - Left table (pinned): Sample ID, Aliquot ID, Spot ID (editable).
      - Right table (main): Excel data + 4 optional columns appended
        for Lab Facilities, References, UPb Analysis Method, Instrument (all editable).
      - Context menus in both tables to set selected cells to a user-defined value.
    """
    data_imported = QtCore.pyqtSignal(list)

    def __init__(self, parent: QWidget):
        super().__init__(parent=None)
        self.loading_manager = LoadingDialogManager.get_instance()

        self.setWindowTitle("UPb Import Wizard")
        self.loadWindowState()
        self.loading_manager = LoadingDialogManager.get_instance()

        main_layout = QVBoxLayout(self)

        # Top bar: file selection, sheet, etc.
        top_layout = QHBoxLayout()
        self.btn_select = QPushButton("Select Excel File")
        self.btn_select.setFixedWidth(150)
        self.btn_select.clicked.connect(self.select_file)
        top_layout.addWidget(self.btn_select)

        self.label_file = QLabel("No file selected.")
        top_layout.addWidget(self.label_file)

        self.identify_rejected = QCheckBox()
        self.identify_rejected.setText("Auto identify rejected analyses")
        self.identify_rejected.setToolTip(
            'If checked, will automatically identify rows with red or strikethrough text as rejected')
        top_layout.addWidget(self.identify_rejected)

        self.sheet_instructions = QLabel("Select sheet with U-Pb data:")
        self.sheet_instructions.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        top_layout.addWidget(self.sheet_instructions)

        self.combo_sheets = QComboBox()
        self.combo_sheets.setFixedWidth(150)
        top_layout.addWidget(self.combo_sheets)
        self.combo_sheets.currentIndexChanged.connect(self.update_upb_sheet)


        main_layout.addLayout(top_layout)

        combo_box_layout = QHBoxLayout()
        combo_box_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        # # Delimiter label + line edit
        # delimiter_label = QLabel("Delimiter:")
        # delimiter_label.setFixedWidth(50)
        # self.delimiter_edit = QLineEdit()
        # self.delimiter_edit.setPlaceholderText("e.g., -, etc.")
        # self.delimiter_edit.setFixedSize(QSize(100, 25))
        # self.delimiter_edit.textChanged.connect(self.update_left_table_on_delimiter_change)  # Connect signal

        combo_box_layout.addWidget(QLabel("Notice: These dropdowns will overwrite all data in the tables.   "), 1,
                                   Qt.AlignmentFlag.AlignLeft)

        # ComboBox for setting Reference
        # does this need to be single click
        self.combo_reference_comboBox = CheckableComboBox()
        self.combo_reference = CheckableSqlQueryModel()
        self.combo_reference_comboBox.setToolTip('Applies the selected reference to all rows')

        # ComboBox for setting Instrument
        self.combo_instrument_comboBox = CheckableComboBox()
        self.combo_instrument = CheckableSqlTableModel()
        self.combo_instrument_comboBox.setToolTip('Applies the selected instrument to all rows')

        # ComboBox for setting LabFacility
        self.combo_lab_facility_comboBox = CheckableComboBox()
        self.combo_lab_facility = CheckableSqlTableModel()
        self.combo_lab_facility_comboBox.setToolTip('Applies the selected lab facility to all rows')

        # ComboBox for setting UPbAnalysisMethod
        self.combo_upb_analysis_method_comboBox = CheckableTreeCombobox()
        self.upb_analysis_method = QSqlTableModel()
        self.combo_upb_analysis_method = CheckableTreeModel()
        self.combo_upb_analysis_method_comboBox.setToolTip('Applies the selected U-Pb analysis method to all rows')

        self.populate_comboBoxes()

        combo_box_layout.addWidget(QLabel("Reference:"))
        combo_box_layout.addWidget(self.combo_reference_comboBox)

        combo_box_layout.addWidget(QLabel("Instrument:"))
        combo_box_layout.addWidget(self.combo_instrument_comboBox)

        combo_box_layout.addWidget(QLabel("Lab Facility:"))
        combo_box_layout.addWidget(self.combo_lab_facility_comboBox)

        combo_box_layout.addWidget(QLabel("UPb Analysis Method:"))
        combo_box_layout.addWidget(self.combo_upb_analysis_method_comboBox)

        main_layout.addLayout(combo_box_layout)

        formats_layout1 = QHBoxLayout()
        formats_layout2 = QHBoxLayout()
        formats_layout1.setAlignment(Qt.AlignmentFlag.AlignRight)
        formats_layout2.setAlignment(Qt.AlignmentFlag.AlignRight)

        # formats_layout1.addWidget(delimiter_label)
        # formats_layout1.addWidget(self.delimiter_edit)

        formats_layout1.addStretch(1)

        # self.delimiter_checkbox = QCheckBox('Enable Delimiter?')
        # self.delimiter_checkbox.checkStateChanged.connect(self.update_left_table_on_delimiter_change)
        # formats_layout2.addWidget(self.delimiter_checkbox, Qt.AlignmentFlag.AlignLeft)
        # self.delimiter_checkbox.setToolTip('Will split the Spot Name column into Sample, Aliquot, (Grain), and Spot Names based on the delimiter')

        formats_layout1.addStretch(4)
        formats_layout2.addStretch(4)

        self.get_valid_unit_formats()

        self.elevation_unit_combobox = QComboBox()
        # self.elevation_unit_combobox.setFixedWidth(100)
        for display_text, backend_id in self.distance_units:
            self.elevation_unit_combobox.addItem(display_text, backend_id)
        formats_layout1.addWidget(QLabel("Elevation Unit:"))
        formats_layout1.addWidget(self.elevation_unit_combobox)
        self.elevation_unit_combobox.setCurrentText(settings.value('elevation_unit_abbreviation'))
        self.elevation_unit_combobox.setToolTip('Applies the selected elevation unit to all rows')

        self.heightdepth_unit_combobox = QComboBox()
        # self.heightdepth_unit_combobox.setFixedWidth(100)
        for display_text, backend_id in self.distance_units:
            self.heightdepth_unit_combobox.addItem(display_text, backend_id)
        formats_layout1.addWidget(QLabel("Height/Depth Unit:"))
        formats_layout1.addWidget(self.heightdepth_unit_combobox)
        self.heightdepth_unit_combobox.setCurrentText(settings.value('heightdepth_unit_abbreviation'))
        self.heightdepth_unit_combobox.setToolTip('Applies the selected height/depth unit to all rows')

        self.sample_age_error_combobox = QComboBox()
        # self.sample_age_error_combobox.setFixedWidth(100)
        for display_text, backend_id in self.error_formats:
            self.sample_age_error_combobox.addItem(display_text, backend_id)
        formats_layout1.addWidget(QLabel("Sample Age Error:"))
        formats_layout1.addWidget(self.sample_age_error_combobox)
        self.sample_age_error_combobox.setCurrentText(settings.value('age_error_format_abbreviation'))
        self.sample_age_error_combobox.setToolTip('Applies the selected sample age error format to all rows')

        self.btn_add_column = QPushButton("Add Column")
        self.btn_add_column.setFixedWidth(150)
        self.btn_add_column.clicked.connect(lambda: self.add_column(None, False))
        formats_layout2.addWidget(self.btn_add_column)

        self.age_unit_combobox = QComboBox()
        # self.age_unit_combobox.setFixedWidth(100)
        for display_text, backend_id in self.age_formats:
            self.age_unit_combobox.addItem(display_text, backend_id)
        formats_layout2.addWidget(QLabel("Age Unit:"))
        formats_layout2.addWidget(self.age_unit_combobox)
        self.age_unit_combobox.setCurrentText(settings.value('age_unit_abbreviation'))
        self.age_unit_combobox.setToolTip('Applies the selected age unit to all rows')

        self.upb_age_error_combobox = QComboBox()
        # self.upb_age_error_combobox.setFixedWidth(100)
        for display_text, backend_id in self.error_formats:
            self.upb_age_error_combobox.addItem(display_text, backend_id)
        formats_layout2.addWidget(QLabel("U-Pb Age Error:"))
        formats_layout2.addWidget(self.upb_age_error_combobox)
        self.upb_age_error_combobox.setCurrentText(settings.value('age_error_format_abbreviation'))
        self.upb_age_error_combobox.setToolTip('Applies the selected U-Pb age error format to all rows')

        self.ratio_error_combobox = QComboBox()
        # self.ratio_error_combobox.setFixedWidth(100)
        for display_text, backend_id in self.error_formats:
            self.ratio_error_combobox.addItem(display_text, backend_id)
        formats_layout2.addWidget(QLabel("Ratio Error"))
        formats_layout2.addWidget(self.ratio_error_combobox)
        self.ratio_error_combobox.setCurrentText(settings.value('ratio_error_format_abbreviation'))
        self.ratio_error_combobox.setToolTip('Applies the selected ratio error format to all rows')

        self.spot_size_unit_combobox = QComboBox()
        # self.spot_size_combobox.setFixedWidth(100)
        for display_text, backend_id in self.distance_units:
            self.spot_size_unit_combobox.addItem(display_text, backend_id)
        formats_layout2.addWidget(QLabel("Spot Size Unit"))
        formats_layout2.addWidget(self.spot_size_unit_combobox)
        self.spot_size_unit_combobox.setCurrentText(settings.value('spotsize_unit_abbreviation'))
        self.spot_size_unit_combobox.setToolTip('Applies the selected spot size unit to all rows')

        self.conc_format_combobox = QComboBox()
        # self.conc_format_combobox.setFixedWidth(150)
        for display_text, backend_id in self.concordance_formats:
            self.conc_format_combobox.addItem(display_text, backend_id)
        formats_layout2.addWidget(QLabel("Concordance Format"))
        formats_layout2.addWidget(self.conc_format_combobox)
        self.conc_format_combobox.setCurrentText(settings.value('concordance_format_abbreviation'))
        self.conc_format_combobox.setToolTip('Applies the selected concordance format to all rows')

        main_layout.addLayout(formats_layout1)
        main_layout.addLayout(formats_layout2)

        self.combos = {"Reference": self.combo_reference_comboBox,
                       "Instrument": self.combo_instrument_comboBox,
                       "Lab Facility": self.combo_lab_facility_comboBox,
                       "UPb Analysis Method": self.combo_upb_analysis_method_comboBox
                       }
        self.static_combos = {
                       "Elevation Unit": self.elevation_unit_combobox,
                       "Height/Depth Unit": self.heightdepth_unit_combobox,
                       "Sample Age Error": self.sample_age_error_combobox,
                       "Age Unit": self.age_unit_combobox,
                       "U-Pb Age Error": self.upb_age_error_combobox,
                       "Ratio Error": self.ratio_error_combobox,
                       "Spot Size Unit": self.spot_size_unit_combobox,
                       "Concordance Format": self.conc_format_combobox
                       }

        # Splitter for left (pinned) vs right (main) tables
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Left pinned table: 3 columns for Sample Name, Aliquot Name, Spot Name
        self.left_tables = {}
        self.left_table = QTableWidget()
        self.left_table.setColumnCount(4)
        self.left_table.setHorizontalHeaderLabels(["Sample Name", "Aliquot Name", "Spot Name", "UPb Analysis Name"])
        self.left_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.left_table.customContextMenuRequested.connect(self.show_left_table_context_menu)

        # Right table for the actual Excel data
        self.right_tables = {}
        self.workbook_tabs = QTabWidget()
        self.right_table = QTableView()
        self.workbook_tabs.addTab(self.right_table, None)

        # Enable the context menu for the checkable combo boxes and connect the signals
        self.combo_reference_comboBox.enable_context_menu(True)
        self.combo_reference_comboBox.edit_triggered.connect(self.edit_combo_box)
        self.combo_reference_comboBox.add_triggered.connect(self.add_combo_box)
        self.combo_instrument_comboBox.enable_context_menu(True)
        self.combo_instrument_comboBox.edit_triggered.connect(self.edit_combo_box)
        self.combo_instrument_comboBox.add_triggered.connect(self.add_combo_box)
        self.combo_lab_facility_comboBox.enable_context_menu(True)
        self.combo_lab_facility_comboBox.edit_triggered.connect(self.edit_combo_box)
        self.combo_lab_facility_comboBox.add_triggered.connect(self.add_combo_box)
        self.combo_upb_analysis_method_comboBox.enable_context_menu(True)
        self.combo_upb_analysis_method_comboBox.edit_triggered.connect(self.edit_combo_box)
        self.combo_upb_analysis_method_comboBox.add_triggered.connect(self.add_combo_box)

        self.workbook_tabs.currentChanged.connect(self.on_tab_changed)

        self.left_layout = QVBoxLayout()
        # Add a vertical spacer to give space at the top
        self.left_top_spacer = QSpacerItem(0, self.workbook_tabs.tabBar().size().height(), QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.left_layout.addItem(self.left_top_spacer)
        self.left_layout.addWidget(self.left_table)

        self.left_widget = QWidget()
        self.left_widget.setLayout(self.left_layout)

        splitter.addWidget(self.left_widget)
        splitter.addWidget(self.workbook_tabs)
        splitter.setStretchFactor(0, 1)  # left narrower
        splitter.setStretchFactor(1, 3)  # right expands

        main_layout.addWidget(splitter)

        # Bottom bar: mapping + import
        bottom_layout = QHBoxLayout()
        self.btn_save_mapping = QPushButton("Save Mapping")
        self.btn_save_mapping.clicked.connect(self.save_mapping)
        bottom_layout.addWidget(self.btn_save_mapping)
        self.btn_save_mapping.setToolTip('Saves the current column mappings, U-Pb metadata selections, and various units/formats to a user-defined name for future use.')

        self.btn_load_mapping = QPushButton("Load Mapping")
        self.btn_load_mapping.clicked.connect(self.load_mapping)
        bottom_layout.addWidget(self.btn_load_mapping)
        self.btn_load_mapping.setToolTip('Loads a previously saved mapping, overwriting the current column mappings, U-Pb metadata selections, and various units/formats.')

        self.conflict_combo = QComboBox()
        self.conflict_combo.addItems(["skip", "add to", "overwrite"])
        self.conflict_combo.currentIndexChanged.connect(self.update_conflict_mode)
        self.conflict_label = QLabel("On import conflict:")
        self.conflict_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.conflict_label.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        bottom_layout.addWidget(self.conflict_label)
        bottom_layout.addWidget(self.conflict_combo)
        self.conflict_combo.setToolTip('Determines how to handle importing items that already exist in the database. "Skip" will not import the item. "Add to" will add any new data to the existing item. "Overwrite" will replace the existing item with the imported data.')

        self.validate_button = QPushButton("Validate Import Data")
        self.validate_button.clicked.connect(self.validate_data)
        bottom_layout.addWidget(self.validate_button)
        self.validate_button.setToolTip('Validates that all required fields are filled, mainly that Sample, Aliquot, and Spot are filled in. Grains are optional')

        self.btn_import = QPushButton("Import to Database")
        self.btn_import.clicked.connect(self.check_and_import)
        bottom_layout.addWidget(self.btn_import)
        self.btn_import.setDisabled(True)
        self.btn_import.setToolTip(
            'Imports the data into the database. Will not be enabled until validation is successful.')
        self.import_clicked = False

        main_layout.addLayout(bottom_layout)

        self.setLayout(main_layout)

        # Dictionary of dataframes for each sheet
        self.dfs = {}
        # Mappings for each sheet
        self.sheet_mappings = {}
        # Map unknown values to static table values
        self.static_mappings = {}
        # Hidden ID columns from combo box values added to the U-Pb sheet
        self.hidden_mappings = {}
        # Original columns, maps the original index to the current index
        # Includes a key for each original column index, and a value of -1 if deleted, as well as a list of 'added' columns in their final index
        self.original_columns = {}

        # Load a dictionary for each category of fields
        self.field_dictionaries = [SQLUtils.sample_possible_user_input_fields,
                                   SQLUtils.gps_possible_user_input_fields,
                                   SQLUtils.column_possible_user_input_fields,
                                   SQLUtils.aliquot_grain_spot_possible_user_input_fields,
                                   SQLUtils.reference_possible_user_input_fields,
                                   SQLUtils.upb_possible_user_input_fields]

        # Dictionary of item types and their parent item types to use when importing
        self.item_parent_dict = {'Samples': None, 'Aliquots': 'Samples', 'Spots': 'Aliquots', 'Grains': 'Spots',
                            'UPbAnalyses': 'Spots', 'SampleAges': 'Samples', 'Columns': 'Samples',
                            'SampleGPSLocations': 'Samples', 'ColumnGPSLocations': 'Columns',
                                 'References': 'UPbAnalyses'}

        # Dictionary of items imported during the current import session
        self.upb_imports = {'SampleID': [], 'AliquotID': [], 'SpotID': [], 'UPbAnalysisID': []}
        self.skipped_conflict_ids = {}

        # Mapping of tag IDs to columns and rows in import sheets. This is built when the tags are imported.
        self.tag_ids = {}
        # Mapping of item IDs to columns in import sheets. This is built when the items are imported.
        self.item_ids = {}

        # False if a mapping has not been loaded since opening the current file
        self.mapping_loaded = False
        # Currently loaded mapping name
        self.current_mapping = None

        # sheets from the openpyxl workbook
        self.sheets = {}
        self.current_sheet_name = None
        # Name of sheet with U-Pb data
        self.upb_sheet_name = None

        # Icons for accepted/rejected
        self.rejected_icon = qtawesome.icon('fa5s.minus-circle', color='red', scale_factor=1.0)
        self.accepted_icon = qtawesome.icon('fa5s.check', color='green', scale_factor=1.0)

        # Sample IDs added or updated during import
        self.sample_ids = []

        # Flash fill connections
        self.left_table.cellChanged.connect(self.handle_left_cell_change)


        self.left_table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.left_table.horizontalHeader().customContextMenuRequested.connect(self.show_left_header_context_menu)

        self.conflict_mode = "skip"

        self.deactivate_widgets()

    def closeEvent(self, a0):
        self.combo_reference_comboBox.disconnect()
        self.combo_instrument_comboBox.disconnect()
        self.combo_lab_facility_comboBox.disconnect()
        self.combo_upb_analysis_method_comboBox.disconnect()
        super().closeEvent(a0)

    def populate_comboBoxes(self, combo=None):
        """
        Populates the combo boxes with values from the database.
        """
        if combo == self.combo_reference_comboBox or combo is None:
            self.combo_reference_comboBox.setFixedWidth(150)
            self.combo_reference_comboBox.set_single_click(True)
            ref_show_cols = settings.value('reference_view_columns')
            query_args = {'show_columns': ref_show_cols,
                          'group_col': f'{ref_show_cols[0]}', 'order_col': f'{ref_show_cols[get_name_column('ReferenceView')]}'}
            view_query = ViewQuery('References', False, **query_args)
            table_query = view_query.table_query
            self.combo_reference.setQuery(table_query)
            while self.combo_reference.canFetchMore():
                self.combo_reference.fetchMore()
            self.combo_reference_comboBox.setModel(self.combo_reference)
            self.combo_reference_comboBox.setModelColumn(get_name_column('ReferenceView'))
            self.combo_reference_comboBox.closing.connect(
                lambda: self.set_all_rows("Reference Display", self.combo_reference))
            self.combo_reference_comboBox.set_line_edit_text('')

        if combo == self.combo_instrument_comboBox or combo is None:
            self.combo_instrument_comboBox.setFixedWidth(150)
            self.combo_instrument_comboBox.set_single_click(True)
            self.combo_instrument = set_table(self.combo_instrument, "Instruments")
            self.combo_instrument_comboBox.setModel(self.combo_instrument)
            self.combo_instrument_comboBox.closing.connect(
                lambda: self.set_all_rows("Instrument Name", self.combo_instrument))
            self.combo_instrument_comboBox.set_line_edit_text('')

        if combo == self.combo_lab_facility_comboBox or combo is None:
            self.combo_lab_facility_comboBox.setFixedWidth(150)
            self.combo_lab_facility_comboBox.set_single_click(True)
            self.combo_lab_facility = set_table(self.combo_lab_facility, "LabFacilities")
            self.combo_lab_facility_comboBox.setModel(self.combo_lab_facility)
            self.combo_lab_facility_comboBox.closing.connect(
                lambda: self.set_all_rows("Lab Facility Name", self.combo_lab_facility))
            self.combo_lab_facility_comboBox.set_line_edit_text('')

        if combo == self.combo_upb_analysis_method_comboBox or combo is None:
            self.combo_upb_analysis_method_comboBox.setFixedWidth(150)
            self.combo_upb_analysis_method_comboBox.set_single_click(True)
            self.upb_analysis_method = set_table(self.upb_analysis_method, "UPbAnalysisMethods")
            self.combo_upb_analysis_method.setSourceModel(self.upb_analysis_method)
            self.combo_upb_analysis_method_comboBox.setModel(self.combo_upb_analysis_method)
            self.combo_upb_analysis_method_comboBox.set_single_click(True)
            self.combo_upb_analysis_method_comboBox.closing.connect(
                lambda: self.set_all_rows("UPb Analysis Method Name", self.combo_upb_analysis_method))
            self.combo_upb_analysis_method_comboBox.set_line_edit_text('')

    def deactivate_widgets(self):
        """
        Disables all widgets within the import wizard.
        """
        self.btn_save_mapping.setEnabled(False)
        self.btn_load_mapping.setEnabled(False)
        self.conflict_combo.setEnabled(False)
        self.btn_import.setEnabled(False)
        self.validate_button.setEnabled(False)
        self.btn_add_column.setEnabled(False)
        # self.delimiter_edit.setEnabled(False)
        # self.delimiter_checkbox.setEnabled(False)
        self.combo_reference_comboBox.setEnabled(False)
        self.combo_instrument_comboBox.setEnabled(False)
        self.combo_lab_facility_comboBox.setEnabled(False)
        self.combo_upb_analysis_method_comboBox.setEnabled(False)
        self.ratio_error_combobox.setEnabled(False)
        self.upb_age_error_combobox.setEnabled(False)
        self.age_unit_combobox.setEnabled(False)
        self.spot_size_unit_combobox.setEnabled(False)
        self.conc_format_combobox.setEnabled(False)

    def activate_widgets(self):
        """
        Enables all widgets within the import wizard.
        """
        self.btn_save_mapping.setEnabled(True)
        self.btn_load_mapping.setEnabled(True)
        self.conflict_combo.setEnabled(True)
        # self.btn_import.setEnabled(True)
        self.validate_button.setEnabled(True)
        self.btn_add_column.setEnabled(True)
        # self.delimiter_edit.setEnabled(True)
        # self.delimiter_checkbox.setEnabled(True)
        self.combo_reference_comboBox.setEnabled(True)
        self.combo_instrument_comboBox.setEnabled(True)
        self.combo_lab_facility_comboBox.setEnabled(True)
        self.combo_upb_analysis_method_comboBox.setEnabled(True)
        self.ratio_error_combobox.setEnabled(True)
        self.upb_age_error_combobox.setEnabled(True)
        self.age_unit_combobox.setEnabled(True)
        self.spot_size_unit_combobox.setEnabled(True)
        self.conc_format_combobox.setEnabled(True)

    def on_cell_clicked(self, index: QModelIndex):
        row = index.row()
        column = index.column()
        header_name = self.right_table.model().headerData(column, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)

        table_name_map = {
            "Reference Display": '"References"',
            "Instrument Name": "Instruments",
            "Lab Facility Name": "LabFacilities",
            "UPb Analysis Method Name": "UPbAnalysisMethods"
        }

        if header_name in table_name_map:
            self.show_listwidget_popup(row, column, header_name, table_name_map[header_name])

    def show_listwidget_popup(self, row, column, header_name, table_name):
        popup = QDialog(self)
        popup.setWindowTitle("Select a valid {} value".format(header_name))

        list_widget = QListWidget(popup)

        # Query database for ID and name
        query = QSqlQuery(f'SELECT * FROM {table_name}')
        data_map = {}  # Dictionary to store ID-Name mapping

        while query.next():
            item_id = query.value(0)
            item_name = query.value(get_name_column(get_view_from_table(table_name)))
            data_map[item_name] = item_id  # Store in map
            list_widget.addItem(item_name)  # Display only name in list

        # Handle item selection
        def on_item_clicked(item):
            selected_name = item.text()
            selected_id = data_map[selected_name]  # Get the associated ID
            self.create_table_item(row, column, selected_name, selected_id, header_name, table_name)
            popup.accept()

        list_widget.itemClicked.connect(on_item_clicked)

        # Set layout
        layout = QVBoxLayout()
        layout.addWidget(list_widget)
        popup.setLayout(layout)

        popup.exec()

    def update_conflict_mode(self):
        self.conflict_mode = self.conflict_combo.currentText()

    def create_table_item(self, row, column, name, id, header_name, table_name):
        """Creates a QTableWidgetItem with stored data."""
        field_to_column = {
            "ReferenceID": self.get_column_index("ReferenceID"),
            "Reference Display": self.get_column_index("Reference Display"),
            "InstrumentID": self.get_column_index("InstrumentID"),
            "Instrument Name": self.get_column_index("Instrument Name"),
            "LabFacilityID": self.get_column_index("LabFacilityID"),
            "Lab Facility Name": self.get_column_index("Lab Facility Name"),
            "UPbAnalysisMethodID": self.get_column_index("UPbAnalysisMethodID"),
            "UPb Analysis Method Name": self.get_column_index("UPb Analysis Method Name")
        }
        # item = QTableWidgetItem(name)
        index = self.right_table.model().index(row, field_to_column[header_name])
        self.right_table.model().setData(index, QTableWidgetItem(name), Qt.ItemDataRole.DisplayRole)
        index = self.right_table.model().index(row, field_to_column[header_name] + 1)
        self.right_table.model().setData(index, QTableWidgetItem(str(id)), Qt.ItemDataRole.DisplayRole)

    def get_valid_unit_formats(self):
        age_unit_query = QSqlQuery()
        age_unit_query.prepare(
            'SELECT AgeUnitAbbreviation, AgeUnitID From AgeUnits')
        self.age_formats = []

        if age_unit_query.exec():
            while age_unit_query.next():
                self.age_formats.append((age_unit_query.value(0), age_unit_query.value(1)))
        else:
            logger_setup.get_logger().error(f"Failed to find age units")
            logger_setup.get_logger().debug(f"Error: {age_unit_query.lastError().text()}")
            logger_setup.get_logger().debug(f"SQL query: {age_unit_query.lastQuery()}")

        distance_units_query = QSqlQuery()
        distance_units_query.prepare(
            'SELECT DistanceUnitAbbreviation, DistanceUnitID From DistanceUnits')
        self.distance_units = []

        if distance_units_query.exec():
            while distance_units_query.next():
                self.distance_units.append((distance_units_query.value(0), distance_units_query.value(1)))
        else:
            logger_setup.get_logger().error(f"Failed to find distance units")
            logger_setup.get_logger().debug(f"Error: {distance_units_query.lastError().text()}")
            logger_setup.get_logger().debug(f"SQL query: {distance_units_query.lastQuery()}")

        concordance_units_query = QSqlQuery()
        concordance_units_query.prepare(
            'SELECT ConcordanceFormatAbbreviation, ConcordanceFormatID From ConcordanceFormats')
        self.concordance_formats = []

        if concordance_units_query.exec():
            while concordance_units_query.next():
                self.concordance_formats.append((concordance_units_query.value(0), concordance_units_query.value(1)))
        else:
            logger_setup.get_logger().error(f"Failed to find concordance formats")
            logger_setup.get_logger().debug(f"Error: {concordance_units_query.lastError().text()}")
            logger_setup.get_logger().debug(f"SQL query: {concordance_units_query.lastQuery()}")

        error_type_format_query = QSqlQuery()
        error_type_format_query.prepare(
            'SELECT ErrorFormatAbbreviation, ErrorFormatID From ErrorFormats')
        self.error_formats = []

        if error_type_format_query.exec():
            while error_type_format_query.next():
                self.error_formats.append((error_type_format_query.value(0), error_type_format_query.value(1)))
        else:
            logger_setup.get_logger().error(f"Failed to find error formats")
            logger_setup.get_logger().debug(f"Error: {error_type_format_query.lastError().text()}")
            logger_setup.get_logger().debug(f"SQL query: {error_type_format_query.lastQuery()}")

    def validate_data(self):
        """
        Validate the data mapped before import.
        :return:
        """
        sender = self.sender()

        logger_setup.get_logger().info("Validating data for import")

        row_count = self.left_table.rowCount()
        if row_count == 0:
            dlg = QMessageBox.question(self, "No U-Pb Data",
                                       "There is no U-Pb data selected." "Continue without U-Pb data?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            dlg.exec_()
            if dlg.result() != QMessageBox.StandardButton.Yes:
                return False
            else:
                upb_data = False
        else:
            upb_data = True

        # Set all cells to transparent background
        for r in range(self.left_table.rowCount()):
            for c in range(self.left_table.columnCount()):
                item = self.left_table.item(r, c)
                if item:
                    item.setBackground(Qt.GlobalColor.transparent)
        for sheet in self.sheet_mappings.keys():
            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[sheet]))
            if self.right_table != self.right_tables[sheet]:
                logger_setup.get_logger().critical(f"Sheet {sheet} does not match the current right table")
            self.right_table.model().clear_all_background_colors()

        if not self.check_static_table_fields():
            self.btn_import.setDisabled(True)
            return False

        if upb_data:
            if not self.validate_ids():
                self.btn_import.setDisabled(True)
                return False
            if not self.check_duplicates_in_left_table():
                self.btn_import.setDisabled(True)
                return False
            if not self.check_for_conflicts():
                self.btn_import.setDisabled(True)
                return False

        if not self.check_unmapped_aliquots_spots_analyses():
            self.btn_import.setDisabled(True)
            return False
        if not self.check_unmapped_references():
            self.btn_import.setDisabled(True)
            return False
        if not self.check_unmapped_samples_columns():
            self.btn_import.setDisabled(True)
            return False

        if sender == self.validate_button:
            QMessageBox.information(self, "Initial Validation Complete", f"Initial data validation is complete.\n\n"
                                                                         "Make sure units and formats are defined for all necessary fields and that GPS data are separated into columns")
        logger_setup.get_logger().info("Data validation complete")
        self.btn_import.setDisabled(False)
        return True


    def validate_ids(self):
        """
        Validate UPb Analysis Name, Sample Name, Aliquot Name, (Grain Name) and Spot Name in the left_table against the database.
        Flag rows that have matching entries in the database.
        """

        logger_setup.get_logger().info("Validating IDs in the left table")

        # Step 1: Check for empty cells in the left table
        empty_cells = self.check_empty_cells_in_left_table()

        if empty_cells:
            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))

            # Step 2: Show dialog to ask if the user wants to use default values
            use_defaults = self.ask_to_use_default_values(empty_cells)

            if use_defaults:
                # Step 3: Fill empty cells with default values
                if self.fill_empty_cells_with_defaults(empty_cells):

                    # Optional: Give user a chance to review and adjust the values
                    QMessageBox.information(self, "Review",
                                            "Purple cells in the left UPb table have been filled with default values. Please review before clicking import again.")
                    logger_setup.get_logger().info("Empty cells have been filled with default values")
                return False
            else:
                logger_setup.get_logger().info("Empty cells, but opted not to use default values")
                return False
        logger_setup.get_logger().info("Validation of IDs complete")
        return True


    def edit_combo_box(self, pos):
        """
        Selected 'Edit' from the context menu for the combo box.
        Open the table edit dialog.
        Args:
            pos (QPoint): Position of the context menu request.
        """
        combo = self.sender()
        if combo == self.combo_reference_comboBox:
            table = "References"
        elif combo == self.combo_instrument_comboBox:
            table = "Instruments"
        elif combo == self.combo_lab_facility_comboBox:
            table = "LabFacilities"
        elif combo == self.combo_upb_analysis_method_comboBox:
            table = "UPbAnalysisMethods"
        else:
            return
        if table in SQLUtils.user_viewable_trees:
            dlg = EditTree(self, table)
        elif table != get_view_from_table(table):
            dlg = EditView(self, table)
        else:
            dlg = EditTable(self, table)
        dlg.exec()
        if dlg.updated:
            if not update_database():
                logger_setup.get_logger().critical('Error updating and displaying database')
                self.close()
            self.populate_comboBoxes(combo)

    def add_combo_box(self, pos, action: QAction | None = None):
        """
        Selected an add action from the context menu for the combo box.
        :param pos: QPoint of the context menu request.
        :param action: QAction that was triggered.
        :return:
        """
        combo = self.sender()
        if combo == self.combo_reference_comboBox:
            table = "References"
        elif combo == self.combo_instrument_comboBox:
            table = "Instruments"
        elif combo == self.combo_lab_facility_comboBox:
            table = "LabFacilities"
        elif combo == self.combo_upb_analysis_method_comboBox:
            table = "UPbAnalysisMethods"
        else:
            return

        dlg = None
        dlg_args = None
        if table in SQLUtils.user_viewable_trees:
            save_expanded_state(table, combo.treeView)
            if action:
                dlg_args = add_tree_popup(combo.treeView, action)
                self.loading_manager.show_loading_dialog('Loading', f'Opening add window for {table}...')
                dlg = AddTreeTags(self, table, **dlg_args)
        elif table == "References":
            self.loading_manager.show_loading_dialog('Loading', 'Opening add window for References...')
            dlg = NewReference(self)
        else:
            self.loading_manager.show_loading_dialog('Loading', f'Opening add window for {table}...')
            dlg = AddTags(self, table)
        if not dlg:
            return
        dlg.exec()
        if dlg.updated:
            if not update_database():
                logger_setup.get_logger().critical('Error updating and displaying database')
                self.close()
            self.populate_comboBoxes(combo)

    def show_left_header_context_menu(self, pos):
        """
        Show a context menu on the horizontal header to set column values or insert columns.
        Args:
            pos (QPoint): Position of the context menu request.
        """
        # Determine the column index under the cursor
        column = self.left_table.horizontalHeader().logicalIndexAt(pos)
        if column < 0:
            return

        menu = QMenu(self)

        # Add options to the menu
        set_value_action = menu.addAction("Set Entire Column to Value")
        set_blank_action = menu.addAction("Set Entire Column to Blank")

        # Execute the menu and get the selected action
        action = menu.exec(self.left_table.horizontalHeader().mapToGlobal(pos))
        if action:
            if action == set_value_action:
                self.set_column_to_value(column, self.left_table)
            elif action == set_blank_action:
                self.set_column_to_blank(column, self.left_table)

    def show_right_header_context_menu(self, pos):
        """
        Show a context menu on the horizontal header to set column values or insert columns.
        Args:
            pos (QPoint): Position of the context menu request.
        """
        # Determine the column index under the cursor
        column_index = self.right_table.horizontalHeader().logicalIndexAt(pos)
        if column_index < 0:
            return

        menu = QMenu(self)

        # Add options to the menu
        set_value_action = menu.addAction("Set Entire Column to Value")
        set_blank_action = menu.addAction("Set Entire Column to Blank")
        insert_before_action = menu.addAction("Insert Column Before")
        insert_after_action = menu.addAction("Insert Column After")

        # Execute the menu and get the selected action
        action = menu.exec(self.right_table.horizontalHeader().mapToGlobal(pos))
        if action:
            if action == set_value_action:
                self.set_column_to_value(column_index, self.right_table)
            elif action == set_blank_action:
                self.set_column_to_blank(column_index, self.right_table)
            elif action == insert_before_action:
                self.add_column(column_index, before=True)
            elif action == insert_after_action:
                self.add_column(column_index, before=False)

    def set_column_to_value(self, column, table: QTableView | QTableWidget):
        """
        Set all cells in the specified column to a user-provided value.
        Args:
            column (int): The column index to update.
        """
        # Prompt the user for a value
        value, ok = QInputDialog.getText(self, "Set Column Value", f"Enter value for column {column + 1}:")
        if not ok or value is None:
            return  # User canceled

        # Update all rows in the column
        table.blockSignals(True)
        if isinstance(table, QTableWidget):
            for row in range(table.rowCount()):
                item = table.item(row, column)
                if item is None:
                    item = QTableWidgetItem()
                    table.setItem(row, column, item)
                item.setText(value)
        elif isinstance(table, QTableView):
            for row in range(table.model().rowCount()):
                index = table.model().index(row, column)
                data = table.model().data(index, Qt.ItemDataRole.DisplayRole)
                if data in ['NULL', None, '']:
                    table.model().setData(index, value, Qt.ItemDataRole.DisplayRole)
        table.blockSignals(False)

    def set_column_to_blank(self, column, table: QTableWidget | QTableView):
        """
        Set all cells in the specified column to blank.
        Args:
            column (int): The column index to update.
        """
        # Update all rows in the column
        if isinstance(table, QTableWidget):
            for row in range(table.model().rowCount()):
                item = table.item(row, column)
                if item is None:
                    item = QTableWidgetItem()
                    table.setItem(row, column, item)
                item.setText("")
        elif isinstance(table, QTableView):
            for row in range(table.model().rowCount()):
                index = table.model().index(row, column)
                table.model().setData(index, "", Qt.ItemDataRole.DisplayRole)

    def handle_vertical_header_double_click(self, logical_index):
        """
        Handle double-clicks on vertical headers to mark rows as rejected.
        Args:
            logical_index (int): The row index corresponding to the double-clicked header.
        """
        status = self.right_table.model().return_row_status(logical_index)
        if logical_index and status == 'rejected':
            self.mark_selected_rows_rejected([logical_index], False)
        else:
            self.mark_selected_rows_rejected([logical_index], True)
        self.update_left_table_row_status()

    # ---------------------------
    #    Context Menu Methods
    # ---------------------------

    # def handle_cell_click(self, row, column):
    #     """
    #     Handle cell clicks for specific columns and show a popup with a QSqlTableModel.
    #     Args:
    #         row (int): Row index of the clicked cell.
    #         column (int): Column index of the clicked cell.
    #     """
    #     # Determine the clicked column's header
    #     column_name = self.right_table.model().headerData(column, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
    #     if not column_name:
    #         return
    #
    #     # Map column names to database tables
    #     column_to_table = {
    #         "Reference Display": "References",
    #         "Instrument Name": "Instruments",
    #         "Lab Facility Name": "LabFacilities",
    #         "UPb Analysis Method Name": "UPbAnalysisMethods"
    #     }
    #
    #     if column_name in column_to_table:
    #         table_name = column_to_table[column_name]
    #         # self.show_table_popup(table_name, row, column)

    # def show_table_popup(self, table_name, row, column_index):
    #     """
    #     Show a popup with a QSqlTableModel for the specified table.
    #     Args:
    #         :param table_name: Name of the database table to display.
    #         :param column_index:
    #         :param row:
    #     """
    #     if table_name == "Reference Display":
    #         table_name = '"References"'
    #     elif table_name == "Instrument Name":
    #         table_name = "Instruments"
    #     elif table_name == "Lab Facility Name":
    #         table_name = "LabFacilities"
    #     elif table_name == "UPb Analysis Method Name":
    #         table_name = "UPbAnalysisMethods"
    #
    #     # Create a QSqlTableModel and set the table
    #     model = CheckableSqlTableModel()
    #     model = set_table(model, table_name)
    #
    #     # Create a QTableView to display the model
    #     combobox = QComboBox()
    #     # combobox.set_single_click(True)
    #     # combobox.set_line_edit_text(None)
    #     # combobox.setModel(model)
    #     combobox.addItems(["Edit", "Add"])
    #
    #     # combobox.closing.connect(lambda: self.set_cell_combobox(model, row, column_index))
    #
    #     self.right_table.setCellWidget(row, column_index, combobox)
    #     self.right_table.setColumnWidth(column_index, 200)

    # def set_cell_combobox(self, model, row, column_index):
    #     name_col = WC.name_column(model.tableName())
    #
    #     field = str(self.right_table.model().headerData(column_index, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)).strip()
    #
    #     for temp_row in range(model.rowCount()):
    #         name_index = model.index(temp_row, name_col)
    #         id_index = model.index(temp_row, 0)
    #
    #         if model.data(name_index, QtCore.Qt.ItemDataRole.CheckStateRole) == Qt.CheckState.Checked:
    #             checked_item_name = model.data(name_index, Qt.ItemDataRole.DisplayRole)
    #             checked_item_id = model.data(id_index, Qt.ItemDataRole.DisplayRole)
    #
    #     field_to_column = {
    #         "ReferenceID": self.get_column_index("ReferenceID"),
    #         "Reference Display": self.get_column_index("Reference Display"),
    #         "InstrumentID": self.get_column_index("InstrumentID"),
    #         "Instrument Name": self.get_column_index("Instrument Name"),
    #         "LabFacilityID": self.get_column_index("LabFacilityID"),
    #         "Lab Facility Name": self.get_column_index("Lab Facility Name"),
    #         "UPbAnalysisMethodID": self.get_column_index("UPbAnalysisMethodID"),
    #         "UPb Analysis Method Name": self.get_column_index("UPb Analysis Method Name"),
    #     }
    #
    #     # Update all rows in the column
    #     self.right_table.blockSignals(True)
    #
    #     if field == "Reference Display":
    #         id_column = field_to_column.get("ReferenceID")
    #     elif field == "Instrument Name":
    #         id_column = field_to_column.get("InstrumentID")
    #     elif field == "Lab Facility Name":
    #         id_column = field_to_column.get("LabFacilityID")
    #     elif field == "UPb Analysis Method Name":
    #         id_column = field_to_column.get("UPbAnalysisMethodID")
    #     else:
    #         id_column = None
    #
    #     name_column = field_to_column.get(field)
    #     index = self.right_table.model().index(row, id_column)
    #     item = self.right_table.model().data(index)
    #     if item is None:
    #         item = QTableWidgetItem()
    #         index = self.right_table.model().index(row, id_column)
    #         self.right_table.model().setData(index, item)
    #     item.setText(str(checked_item_id))
    #
    #     index = self.right_table.model().index(row, name_column)
    #     item = self.right_table.model().data(index)
    #     if item is None:
    #         item = QTableWidgetItem()
    #         index = self.right_table.model().index(row, name_column)
    #         self.right_table.model().setData(index, item)
    #     item.setText(str(checked_item_name))
    #     self.right_table.blockSignals(False)

    def add_column(self, column_index=None, before=False, field=None):
        """
        Adds a column to the right QTableWidget
        :param column_index: Int index of the column to add.
        :param before: True if inserting before the specified column index.
        :param field: Optional field name to add without prompting the user.
        """
        # Open the Column Map Dialog to let the user select a column name and data type

        if self.sender() in [self.combo_reference_comboBox, self.combo_instrument_comboBox,
                             self.combo_lab_facility_comboBox, self.combo_upb_analysis_method_comboBox]:
            # Set the current sheet to the U-Pb data sheet
            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))

        if field is None:
            dialog = ColumnMapDialog("New Column", "None", self)
            if dialog.exec():
                selected_field = dialog.get_selected_value()
            else:
                return
        else:
            selected_field = field

        # Ensure the user selected a valid field
        if selected_field == "None":
            QMessageBox.warning(self, "Invalid Selection", "Please select a valid column field.")
            return

        for test_idx in range(self.right_table.model().columnCount()):
            header = self.right_table.model().headerData(test_idx, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
            if header and header.startswith(selected_field):
                QMessageBox.warning(self, "Duplicate Column", f"Column '{selected_field}' already exists.")
                return

        self.loading_manager.show_loading_dialog('Adding Column', f'Adding column {selected_field}...')
        if column_index is None:
            # Insert the new column at the end of the table
            column_index = self.right_table.model().columnCount()
        else:
            if not before:
                column_index += 1

        self.right_table.model().insertColumn(column_index)

        # Set the column header
        header_text = f"{selected_field}"
        self.right_table.model().setHeaderData(column_index, Qt.Orientation.Horizontal, header_text, Qt.ItemDataRole.EditRole)
        self.right_table.model().setHeaderData(column_index, Qt.Orientation.Vertical, QColor("#C0FFB8"), Qt.ItemDataRole.BackgroundRole)  # Green background for new column

        # Update the mappings for the current sheet
        # Shift existing mappings to the right, starting with the largest index
        ordered_columns = sorted(self.sheet_mappings[self.current_sheet_name].keys(), reverse=True)
        for column in ordered_columns:
            field = self.sheet_mappings[self.current_sheet_name][column]
            if column >= column_index:
                # Shift the mapping right by one
                self.sheet_mappings[self.current_sheet_name][column + 1] = field
        self.sheet_mappings[self.current_sheet_name][column_index] = selected_field
        # Update the original columns mapping
        for index, column in self.original_columns[self.current_sheet_name].items():
            if index == 'added':
                continue
            if column >= column_index:
                self.original_columns[self.current_sheet_name][index] = column + 1
        if 'added' not in self.original_columns[self.current_sheet_name].keys():
            self.original_columns[self.current_sheet_name]['added'] = []
        # For any existing added columns, shift them right as well
        new_added = []
        for index in self.original_columns[self.current_sheet_name]['added']:
            if index >= column_index:
                index += 1
            new_added.append(index)
        self.original_columns[self.current_sheet_name]['added'] = new_added
        # Add the new column index to the list of added columns
        self.original_columns[self.current_sheet_name]['added'].append(column_index)

        if selected_field == "Reference Display":
            field = "ReferenceID"
        elif selected_field == "Instrument Name":
            field = "InstrumentID"
        elif selected_field == "Lab Facility Name":
            field = "LabFacilityID"
        elif selected_field == "UPb Analysis Method Name":
            field = "UPbAnalysisMethodID"

        # add additional ID column if column is References, Instruments, Analysis Methods, or Lab Facilities
        if selected_field in ["Reference Display", "Instrument Name", "Lab Facility Name", "UPb Analysis Method Name"]:
            # ADD ID Column to the tablewidget
            # Insert the new column at the end of the table
            column_index = self.right_table.model().columnCount()
            self.right_table.model().insertColumn(column_index)

            # Set the column header
            header_text = f"{field}"
            self.right_table.model().setHeaderData(column_index, Qt.Orientation.Horizontal, header_text, Qt.ItemDataRole.EditRole)

            # Update the mappings for the current sheet
            self.sheet_mappings[self.current_sheet_name][column_index] = field
            self.original_columns[self.current_sheet_name]['added'].append(column_index)

            self.right_table.hideColumn(column_index)
            # self.right_table.resizeColumnsToContents()

        self.loading_manager.close_loading_dialog('Adding Column', f'Adding column {selected_field}...')

        # Notify the user only when method is initated from the user
        if field is None:
            QMessageBox.information(self, "Column Added", f"Column '{header_text}' added successfully.")

    def show_left_table_context_menu(self, pos: QPoint):
        """
        Context menu for the left table (Sample ID, Aliquot ID, Spot ID).
        Allows setting all selected cells to a user-defined value.
        """
        menu = QMenu(self)
        set_value_action = menu.addAction("Set Selected Cells to Value...")

        action = menu.exec(self.left_table.mapToGlobal(pos))
        if action:
            if action == set_value_action:
                new_value, ok = QInputDialog.getText(self, "Set Value", "Enter new value:")
                if ok:
                    for item in self.left_table.selectedItems():
                        item.setText(new_value)

    def show_right_table_context_menu(self, pos: QPoint):
        """
        Context menu for the right table.
        Includes remove rows, mark rows rejected, unmark,
        and set selected cells to a user-defined value.
        """
        logger_setup.get_logger().info("Showing right table context menu")
        menu = QMenu(self)
        remove_action = menu.addAction("Remove Selected Rows")
        disable_action = menu.addAction("Disable Selected Rows")
        reject_action = menu.addAction("Mark Selected Rows as Rejected")
        accept_action = menu.addAction("Mark Selected Rows as Accepted")
        set_value_action = menu.addAction("Set Selected Cells to Value...")
        remove_column = menu.addAction('Remove Selected Column')

        action = menu.exec(self.right_table.mapToGlobal(pos))
        if action:
            if not self.right_table.model().selectedIndexes():
                return
            selected_rows = []
            selected_columns = []
            for index in self.right_table.model().selectedIndexes():
                if index.row() not in selected_rows:
                    selected_rows.append(index.row())
                if index.column() not in selected_columns:
                    selected_columns.append(index.column())
            if action == remove_action:
                self.remove_selected_rows(selected_rows)
            elif action == disable_action:
                self.disable_selected_rows(selected_rows)
            elif action == reject_action:
                self.mark_selected_rows_rejected(selected_rows, True)
            elif action == accept_action:
                self.mark_selected_rows_rejected(selected_rows, False)
            elif action == set_value_action:
                new_value, ok = QInputDialog.getText(self, "Set Value", "Enter new value:")
                if ok:
                    for index in self.right_table.model().selectedIndexes():
                        self.right_table.model().setData(index, new_value, Qt.ItemDataRole.DisplayRole)
            elif action == remove_column:
                self.remove_selected_columns(selected_columns)

        self.repaint()

    def show_right_table_vertical_header_context_menu(self, pos: QPoint):
        """
        Context menu for the right table.
        Includes remove rows, mark rows rejected, unmark,
        and set selected cells to a user-defined value.
        """
        row = self.right_table.verticalHeader().logicalIndexAt(pos)
        if row == -1:  # Ensure a valid header was clicked
            return

        menu = QMenu(self)
        remove_action = menu.addAction("Remove Selected Rows")
        disable_action = menu.addAction("Disable Selected Rows")
        reject_action = menu.addAction("Mark Selected Rows as Rejected")
        accept_action = menu.addAction("Mark Selected Rows as Accepted")

        action = menu.exec(self.right_table.mapToGlobal(pos))
        if action:
            selected_rows = []
            for index in self.right_table.selectedIndexes():
                if index.row() not in selected_rows:
                    selected_rows.append(index.row())
            if action == remove_action:
                self.remove_selected_rows(selected_rows)
            elif action == disable_action:
                self.disable_selected_rows(selected_rows)
            elif action == reject_action:
                self.mark_selected_rows_rejected(selected_rows, True)
            elif action == accept_action:
                self.mark_selected_rows_rejected(selected_rows, False)

        self.repaint()

    # ---------------------------
    #     File & Sheet Loading
    # ---------------------------

    def handle_left_cell_change(self, row, column):
        """
        Handle cell value changes in the left table. Ask the user if they want to flash fill downward.
        """

        # Get the current value of the cell
        current_value = self.left_table.item(row, column).text().strip()
        # if self.left_table.item(row + 1, column) is None:
        #     return

        header_name = self.left_table.horizontalHeaderItem(column).text().strip()
        # headers for U-Pb data in right table
        right_column = None
        for r_column in range(self.right_tables[self.upb_sheet_name].model().columnCount()):
            r_header_name = self.right_tables[self.upb_sheet_name].model().headerData(r_column, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
            if r_header_name == header_name:
                right_column = r_column
                break
        if right_column is not None:
            # Update the value in the right table as well
            index = self.right_tables[self.upb_sheet_name].model().index(row, right_column)
            if self.right_tables[self.upb_sheet_name].model().data(index, Qt.ItemDataRole.DisplayRole) != current_value:
                self.right_tables[self.upb_sheet_name].model().setData(index, current_value, Qt.ItemDataRole.DisplayRole)

            # next_value = self.left_table.item(row + 1, column).text().strip()
            # # If the value is empty or invalid, ignore
            # if (not current_value or
            #         next_value == current_value or
            #         len(next_value) > 0 or
            #         len(self.left_table.selectionModel().selectedIndexes()) > 1):
            #     return

            ## Check if the user wants to flash fill - disabled for now
            # reply = QMessageBox.question(
            #     self, "Flash Fill Downward",
            #     "Do you want to auto-fill downward with this value for blank cells?",
            #     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            #     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            # )
            #
            # if reply == QMessageBox.StandardButton.Yes:
            #     self.flash_fill_downward(target_table, row, column, current_value)

    def handle_right_cell_change(self, index):
        if self.sender() == self.right_tables[self.upb_sheet_name].model():
            right_table = self.right_tables[self.upb_sheet_name]
            # Get the current value of the cell
            current_value = str(right_table.model().data(index, Qt.ItemDataRole.DisplayRole)).strip()

            header_name = self.right_table.model().headerData(index.column(), Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)

            if self.right_table == self.right_tables[self.upb_sheet_name]:
                for l_column in range(self.left_table.columnCount()):
                    l_header_name = self.left_table.horizontalHeaderItem(l_column).text().strip()
                    if l_header_name == header_name:
                        # Update the value in the left table as well
                        item = self.left_table.item(index.row(), l_column)
                        if item is None:
                            item = QTableWidgetItem()
                            self.left_table.setItem(index.row(), l_column, item)
                        if item.text().strip() != current_value:
                            item.setText(current_value)
                        break

            # index_below = right_table.model().index(index.row() + 1, index.column())
            # next_value = right_table.model().index(index.row() + 1, index.column).data(Qt.ItemDataRole.DisplayRole)
            # # If the value is empty or invalid, ignore
            # if (not current_value or
            #         next_value == current_value or
            #         len(next_value) > 0 or
            #         len(right_table.selectionModel().selectedIndexes()) > 1):
            #     return

            ## Check if the user wants to flash fill - disabled for now
            # reply = QMessageBox.question(
            #     self, "Flash Fill Downward",
            #     "Do you want to auto-fill downward with this value for blank cells?",
            #     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            #     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            # )
            #
            # if reply == QMessageBox.StandardButton.Yes:
            #     self.flash_fill_downward(right_table, row, column, current_value)
        else:
            return


    def flash_fill_downward(self, target_table, start_row, column, value):
        """
        Flash fill a column downward starting from a given row.
        """

        target_table.blockSignals(True)

        # Start from the next row and go downward
        for row in range(start_row + 1, target_table.rowCount()):
            item = target_table.item(row, column)

            # If the cell is blank, fill it with the given value
            if item is None or not item.text().strip():
                # If the cell doesn't exist, create it
                if item is None:
                    item = QTableWidgetItem()
                    target_table.setItem(row, column, item)

                # Set the value
                item.setText(value)
            else:
                # Stop when a non-blank value is encountered
                break
        target_table.blockSignals(False)

    def select_file(self):
        dlg = QFileDialog(self)
        path, _ = dlg.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
        if path:
            logger_setup.get_logger().debug("Selected Excel File: " + path)
            self.selected_file_path = path
            self.label_file.setText(f"Selected File: {os.path.basename(path)}")
            file_size_bytes = os.path.getsize(path)
            file_size_mb = file_size_bytes / (1024 * 1024)
            if file_size_mb > 5:
                reply = QMessageBox.question(self, "Large File Warning",
                                             f"The selected file is {file_size_mb:.2f} MB.\n\n"
                                             "Loading large files may take a while and could cause the application to become unresponsive.\n\n"
                                             "Do you want to continue?",
                                             QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                             QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.No:
                    return
            try:
                logger_setup.get_logger().debug(f"Loading sheets from Excel file {os.path.basename(path)}")
                self.loading_manager.show_loading_dialog("Loading", f"Loading {os.path.basename(path)}...")
                load_workbook_start = time.time()
                wb = load_workbook(path, data_only=True, rich_text=True)
                # wb = load_workbook(path, data_only=True, keep_vba=False, read_only=True, rich_text=True, keep_links=False)
                logger_setup.get_logger().info(f"Excel file {os.path.basename(path)} loaded in {(time.time() - load_workbook_start):.2f} seconds")
                self.combo_sheets.clear()
                self.sheet_mappings.clear()
                self.static_mappings.clear()
                self.item_ids.clear()
                self.right_tables.clear()
                self.combo_sheets.addItems(wb.sheetnames)
                combo_sheets = QComboBox()
                combo_sheets.clear()
                combo_sheets.addItems(wb.sheetnames)
                self.workbook_tabs.clear()
                for sheet in wb.worksheets:
                    sheet_name = sheet.title.strip()
                    self.sheets[sheet_name] = sheet

                # Ask the user which sheet contains U-Pb data if multiple sheets exist
                if len(self.sheets.values()) > 1:
                    sheet_dialog = QDialog(self)
                    sheet_dialog.setWindowTitle("Select U-Pb sheet")
                    sheet_layout = QVBoxLayout()
                    sheet_label = QLabel("Multiple sheets found. Please select the sheet containing U-Pb data:")
                    sheet_layout.addWidget(sheet_label)
                    sheet_layout.addWidget(combo_sheets)
                    button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
                    button_box.accepted.connect(sheet_dialog.accept)
                    button_box.rejected.connect(sheet_dialog.reject)
                    sheet_layout.addWidget(button_box)
                    sheet_dialog.setLayout(sheet_layout)
                    if sheet_dialog.exec() == QDialog.DialogCode.Rejected:
                        self.loading_manager.close_loading_dialog("Loading", f"Loading {os.path.basename(path)}...")
                        return
                    selected_sheet = combo_sheets.currentText()
                    if selected_sheet:
                        self.combo_sheets.setCurrentText(selected_sheet)
                for sheet in self.sheets.values():
                    row_count = sheet.max_row
                    sheet.title = sheet.title.strip()
                    if row_count > 10000:
                        msg_box = QMessageBox()
                        msg_box.setIcon(QMessageBox.Icon.Warning)
                        msg_box.setWindowTitle("Large Sheet Warning")
                        msg_box.setText(f"Sheet {sheet.title} has {row_count} rows.\n\n"
                                        f"Loading the data will take a while and may cause the application to become unresponsive.\n\n"
                                        f"Do you want to continue?")
                        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                        reply = msg_box.exec()
                        if reply != QMessageBox.StandardButton.Yes:
                            continue
                    logger_setup.get_logger().debug(f"Loading sheet: {sheet.title}")
                    self.loading_manager.show_loading_dialog(f"Loading {sheet.title}", f"Loading sheet {sheet.title} with {row_count} rows...")
                    # Add a new tab for each sheet
                    self.sheet_mappings[sheet.title] = {}
                    self.static_mappings[sheet.title] = {}
                    self.right_table = QTableView()
                    self.right_tables[sheet.title] = self.right_table
                    self.workbook_tabs.addTab(self.right_table, sheet.title)
                    self.current_sheet_name = sheet.title

                    # Display data on the right table
                    self.display_right_table_with_styles(sheet.title)

                    # Record the original mapping of column indexes
                    for col in range(self.right_table.model().columnCount()):
                        self.original_columns[self.current_sheet_name] = {}
                        self.original_columns[self.current_sheet_name][col] = col

                    self.right_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
                    self.right_table.customContextMenuRequested.connect(self.show_right_table_context_menu)
                    self.right_table.model().userDataChanged.connect(self.handle_right_cell_change)

                    self.right_table.doubleClicked.connect(self.on_cell_clicked)

                    self.right_table.verticalHeader().sectionDoubleClicked.connect(
                        self.handle_vertical_header_double_click)

                    self.right_table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
                    self.right_table.horizontalHeader().customContextMenuRequested.connect(
                        self.show_right_header_context_menu)

                    self.right_table.verticalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
                    self.right_table.verticalHeader().customContextMenuRequested.connect(
                        self.show_right_table_vertical_header_context_menu)

                    # Connect the header double-click signal to the handler
                    header = self.right_table.horizontalHeader()
                    header.sectionDoubleClicked.connect(self.handle_header_double_clicked)

                    if sheet.title == self.combo_sheets.currentText():
                        # This is the U-Pb sheet, so sync it with the left table
                        # Make sure the left table has the same number of rows
                        self.left_table.setRowCount(self.right_table.model().rowCount())
                        # Scroll synchronization (vertical)
                        self.left_table.verticalScrollBar().valueChanged.connect(
                            self.right_table.verticalScrollBar().setValue
                        )
                        self.right_table.verticalScrollBar().valueChanged.connect(
                            self.left_table.verticalScrollBar().setValue
                        )

                    self.loading_manager.close_loading_dialog(f"Loading {sheet.title}",
                                                             f"Loading sheet {sheet.title} with {row_count} rows...")

                self.workbook_tabs.setCurrentIndex(self.combo_sheets.currentIndex())
                wb.close()
            except Exception as e:
                logger_setup.get_logger().critical("Error", f"Failed to read Excel file:\n{e}")
                return
        self.activate_widgets()
        self.resize_tables()
        self.mapping_loaded = False
        self.loading_manager.close_loading_dialog("Loading", f"Loading {os.path.basename(path)}...")

    def resize_tables(self):
        """
        Resize the tables to fit their contents and match the top positions.
        """
        tab_bar_height = self.workbook_tabs.tabBar().size().height()
        self.left_top_spacer.changeSize(0, tab_bar_height, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

    def update_upb_sheet(self):
        """
        Update the name of the sheet with U-Pb data.
        """
        self.upb_sheet_name = self.combo_sheets.currentText()
        if self.upb_sheet_name in self.right_tables.keys():
            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))
            self.update_current_right_table()
            self.sync_left_table_rows()
            if self.current_mapping:
                self.load_mapping(self.current_mapping)

    def update_current_right_table(self):
        """
        Update the current right table widget.
        """
        self.current_sheet_name = self.workbook_tabs.tabText(self.workbook_tabs.currentIndex())
        self.right_table = self.right_tables[self.current_sheet_name]

    def on_tab_changed(self, index):
        """
        Handle tab changes in the workbook tabs.
        """
        if not self.workbook_tabs.tabText(index):
            # No valid tab selected
            return
        self.update_current_right_table()
        if self.current_sheet_name == self.upb_sheet_name:
            self.left_widget.show()
            self.sync_left_table_rows()
        else:
            self.left_widget.hide()

    def display_right_table_with_styles(self, sheet_name):
        """
        Display the right table with openpyxl-based formatting
        + add 4 extra columns for Lab Facilities, Source, Analysis Method, Instrument (editable).
        """
        # self.loading_manager.show_loading_dialog('Loading', f'Displaying {sheet_name}...')
        logger_setup.get_logger().debug(f"Displaying sheet: {sheet_name}")
        display_start_time = time.time()
        check_style = self.identify_rejected.isChecked()

        dataframe_model = ImportSheetModel(self.sheets[sheet_name], check_style)
        print(f'Rows: {dataframe_model.rowCount()}, Columns: {dataframe_model.columnCount()}')
        self.right_table.blockSignals(True)
        self.right_table.setModel(dataframe_model)
        self.right_table.blockSignals(False)

        self.right_table.resizeColumnsToContents()
        self.loading_manager.close_loading_dialog('Loading', f'Displaying {sheet_name}...')
        logger_setup.get_logger().info(f"Displayed sheet '{sheet_name}' in {(time.time() - display_start_time):.2f} seconds")

    def sync_left_table_rows(self):
        """
        Make the left table have the same row count as the right table
        and add editable cells for Sample ID, Aliquot ID, Spot ID.
        """
        self.left_table.blockSignals(True)
        right_table = self.right_tables[self.upb_sheet_name]
        row_count = right_table.model().rowCount()
        # row_count = right_table.model().rowCount()
        self.left_table.setRowCount(row_count)
        for r in range(row_count):
            for c in range(3):
                if not self.left_table.item(r, c):
                    self.left_table.setItem(r, c, QTableWidgetItem(""))
        self.left_table.blockSignals(False)

        self.left_table.resizeColumnsToContents()

    def set_all_rows(self, field, model):
        """
        Set all rows in the specified column to the given value.
        Args:
            field (str): The field name (e.g., 'Reference', 'Instrument').
            model (checkable model): The model to retrieve checks from.
        """
        if field == "Reference Display":
            id_name = "ReferenceID"
        elif field == "Instrument Name":
            id_name = "InstrumentID"
        elif field == "Lab Facility Name":
            id_name = "LabFacilityID"
        elif field == "UPb Analysis Method Name":
            id_name = "UPbAnalysisMethodID"
        else:
            id_name = None

        try:
            table = model.tableName()
            name_column = get_name_column(get_view_from_table(table))
        except AttributeError:
            table = model.table  # for trees
            name_column = 0  # for trees
        # Determine the column index for the field
        checked_item_name = None
        checked_item_id = None
        if table not in SQLUtils.user_viewable_trees:
            for row in range(model.rowCount()):
                name_index = model.index(row, name_column)
                id_index = model.index(row, 0)

                if model.data(name_index, QtCore.Qt.ItemDataRole.CheckStateRole) == Qt.CheckState.Checked:
                    checked_item_name = model.data(name_index, Qt.ItemDataRole.DisplayRole)
                    checked_item_id = model.data(id_index, Qt.ItemDataRole.DisplayRole)
                    break
            logger_setup.get_logger().info(f"Checked item: {checked_item_name}, ID: {checked_item_id}")
        else:
            # Get the checked item from the tree
            checked_items, partially_checked_items, checked_indices, partially_checked_indices = model.traverse_checkable_tree(
                QtCore.QModelIndex())
            if checked_items:
                checked_item_name = checked_indices[0].data(QtCore.Qt.ItemDataRole.DisplayRole)
                checked_item_id = checked_items[0]

        id_column = self.get_column_index(id_name)  # Column with ID header in the right table
        name_col = self.get_column_index(field)  # Column with name header in the right table

        if (checked_item_id is None or checked_item_name is None) and (id_column or name_col):
            # Remove the columns that were added before
            self.loading_manager.show_loading_dialog('Removing Column', f'Removing column {field}...')
            self.remove_selected_columns([id_column, name_col])
            self.hidden_mappings.pop(id_column, None)
            self.loading_manager.close_loading_dialog('Removing Column', f'Removing column {field}...')
            return
        elif self.get_column_index(field) is None:
            self.add_column(field=field)
            id_column = self.get_column_index(id_name)  # Column with ID header in the right table
            name_col = self.get_column_index(field)  # Column with name header in the right table

        # Update all rows in the column
        self.right_table.blockSignals(True)

        for row in range(self.right_table.model().rowCount()):
            index = self.right_table.model().index(row, id_column)
            self.right_table.model().setData(index, str(checked_item_id), Qt.ItemDataRole.DisplayRole)
            index = self.right_table.model().index(row, name_col)
            self.right_table.model().setData(index, str(checked_item_name), Qt.ItemDataRole.DisplayRole)

        self.right_table.blockSignals(False)
        self.right_table.hideColumn(id_column)
        self.hidden_mappings[id_column] = {}
        self.hidden_mappings[id_column][id_name] = checked_item_id
        # self.right_table.resizeColumnsToContents()
        QMessageBox.information(self, "Success", f"All rows updated with '{str(checked_item_name)}' for {field}.")

    def get_column_index(self, header_name):
        """
        Get the column index of a header by its name.
        Args:
            header_name (str): The name of the header.
        Returns:
            int: The column index, or None if not found.
        """
        for col in range(self.right_table.model().columnCount()):
            header = self.right_table.model().headerData(col, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
            if header and header.startswith(header_name):
                return col
        return None

    def get_column_name(self, column_index):
        return self.right_table.model().headerData(column_index, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)

    def update_left_table_row_status(self):
        """
        Update the left table row text appearance based on the right table row status.
        """
        if self.current_sheet_name == self.upb_sheet_name:
            # get the default style text color
            if not self.style().standardPalette():
                text_color = QColor("#000000")  # Default to black
            else:
                text_color = self.style().standardPalette().color(QPalette.ColorRole.Text)

            for row in range(self.right_table.model().rowCount()):
                status = self.right_table.model().return_row_status(row)
                if status == 'disabled':
                    color = QColor("#A0A0A0")  # Gray
                else:
                    color = text_color  # Default text color
                for c in range(self.left_table.columnCount()):
                    item = self.left_table.item(row, c)
                    if item:
                        item.setForeground(QBrush(color))  # Gray text

    def remove_selected_rows(self, rows=None):
        if rows is None:
            selected_rows = [index.row() for index in self.right_table.selectedIndexes()]
            if not selected_rows:
                return
        else:
            selected_rows = rows

        sr = sorted(selected_rows, reverse=True)
        for r in sr:
            self.right_table.model().removeRow(r)
            if self.current_sheet_name == self.upb_sheet_name:
                self.left_table.removeRow(r)

        # self.right_table.resizeColumnsToContents()
        # if self.current_sheet_name == self.upb_sheet_name:
        #     self.left_table.resizeColumnsToContents()

    def remove_selected_columns(self, columns=None):
        """
        Remove selected columns from the right table, the column mappings,
        and associated data structures.
        """
        if columns is None:
            selected_columns = [index.column() for index in self.right_table.selectedIndexes()]
            if not selected_columns:
                return
        else:
            selected_columns = columns

        # Sort selected columns in descending order to remove from the rightmost column
        sorted_columns = sorted(selected_columns, reverse=True)

        adjusted_mappings = {}
        # Preserve mappings for columns to the left of the minimum selected column. Shift others left.
        for index in self.sheet_mappings[self.current_sheet_name]:
            if index < min(sorted_columns):
                adjusted_mappings[index] = self.sheet_mappings[self.current_sheet_name][index]
            if index not in sorted_columns and index > min(sorted_columns):
                field = self.sheet_mappings[self.current_sheet_name][index]
                shift = sum(1 for deleted_index in sorted_columns if deleted_index < index)
                adjusted_mappings[index - shift] = field

        for column_index in sorted_columns:
            # Remove the column from the right table
            self.right_table.model().removeColumn(column_index)

            # Check if this is one of the added columns
            if column_index in self.original_columns[self.current_sheet_name]['added']:
                # If so, remove it from the added list
                self.original_columns[self.current_sheet_name]['added'].remove(column_index)
            else:
                # It's an original column, so find the key with this current index and change it to -1
                for key, value in self.original_columns[self.current_sheet_name].items():
                    if value == column_index:
                        self.original_columns[self.current_sheet_name][key] = -1
                        break
                # Adjust the indices of original columns after removal
                for index, column in self.original_columns.items():
                    column = list(column)[0]
                    shift = sum(1 for deleted_index in sorted_columns if deleted_index < column)
                    self.original_columns[index] = column - shift
                # Adjust the indices in the 'added' list
                new_added = []
                for index in self.original_columns[self.current_sheet_name]['added']:
                    shift = sum(1 for deleted_index in sorted_columns if deleted_index < index)
                    new_added.append(index - shift)
                self.original_columns[self.current_sheet_name]['added'] = new_added


        self.sheet_mappings[self.current_sheet_name] = adjusted_mappings
        # self.right_table.resizeRowsToContents()
        # Notify the user
        QMessageBox.information(self, "Columns Removed", "Selected columns have been successfully removed.")

    def update_vertical_headers(self):
        """
        Update the vertical headers to ensure they match the current row indices.
        """
        row_count = self.right_table.model().rowCount()
        for row_idx in range(row_count):
            self.right_table.model().setHeaderData(row_idx, Qt.Orientation.Vertical, row_idx + 1)
        self.repaint()

    def disable_selected_rows(self, rows=None):
        if rows is None:
            selected_rows = [index.row() for index in self.right_table.selectedIndexes()]
            if not selected_rows:
                return
        else:
            selected_rows = rows

        for r in selected_rows:
            self.right_table.model().setHeaderData(r, Qt.Orientation.Vertical, "disabled", Qt.ItemDataRole.DecorationRole)
        if self.current_sheet_name == self.upb_sheet_name:
            self.update_left_table_row_status()

        self.repaint()

    def mark_selected_rows_rejected(self, rows, rejected: bool):
        if not rows:
            return
        if rejected:
            for r in rows:
                self.right_table.model().setHeaderData(r, Qt.Orientation.Vertical, "rejected", Qt.ItemDataRole.DecorationRole)
        else:
            for r in rows:
                self.right_table.model().setHeaderData(r, Qt.Orientation.Vertical, "accepted", Qt.ItemDataRole.DecorationRole)
        if self.current_sheet_name == self.upb_sheet_name:
            self.update_left_table_row_status()

        self.repaint()


    # ---------------------------
    #     Header Double Click
    # ---------------------------

    def handle_header_double_clicked(self, logical_index):
        """
        Double-click on a right table header => open mapping dialog.
        """
        value = self.right_table.model().headerData(logical_index, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
        if not value:
            return
        original_header_text = str(logical_index)
        curr_map = self.sheet_mappings[self.current_sheet_name].get(logical_index, "None")
        dialog = ColumnMapDialog(original_header_text, curr_map, self)
        if dialog.exec():
            new_field = dialog.get_selected_value()
            if new_field == "None" or not new_field:
                if logical_index in self.sheet_mappings[self.current_sheet_name]:
                    if self.sheet_mappings[self.current_sheet_name][logical_index] == "Grain Name" and self.current_sheet_name == self.upb_sheet_name:
                        # Remove the Grain Name column from the left table if it exists
                        left_headers = [self.left_table.horizontalHeaderItem(i).text() for i in
                                        range(self.left_table.columnCount())]
                        if "Grain Name" in left_headers:
                            grain_col_index = left_headers.index("Grain Name")
                            self.left_table.removeColumn(grain_col_index)
                    del self.sheet_mappings[self.current_sheet_name][logical_index]
                # Reset the text and background color
                self.right_table.model().setHeaderData(logical_index, Qt.Orientation.Horizontal, str(logical_index), Qt.ItemDataRole.EditRole)
                self.right_table.model().setHeaderData(logical_index, Qt.Orientation.Horizontal, Qt.GlobalColor.transparent, Qt.ItemDataRole.BackgroundRole)
            else:
                self.sheet_mappings[self.current_sheet_name][logical_index] = (new_field)
                self.right_table.model().setHeaderData(logical_index, Qt.Orientation.Horizontal, f"{new_field}", Qt.ItemDataRole.EditRole)
                self.right_table.model().setHeaderData(logical_index, Qt.Orientation.Horizontal, QColor("#C0FFB8"), Qt.ItemDataRole.BackgroundRole)

                # If it’s Sample Name / Aliquot Name / Grain Name / Spot Name / UPb Analysis Name, auto-populate left table
                if (new_field in ["Sample Name", "Aliquot Name", "Grain Name", "Spot Name", "UPb Analysis Name"] and
                        self.current_sheet_name == self.upb_sheet_name):
                    self.update_left_table_on_header_change(new_field, logical_index)

    def update_left_table_on_header_change(self, field, logical_index):
        sample_col = None
        aliquot_col = None
        spot_col = None
        grain_col = None
        upb_analysis_col = None
        for column in range(self.left_table.columnCount()):
            if self.left_table.horizontalHeaderItem(column).text() == "Sample Name":
                sample_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Aliquot Name":
                aliquot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Spot Name":
                spot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Grain Name":
                grain_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "UPb Analysis Name":
                upb_analysis_col = column

        if field == "Sample Name":
            for r in range(self.right_tables[self.upb_sheet_name].model().rowCount()):
                index = self.right_tables[self.upb_sheet_name].model().index(r, logical_index)
                cell_data = self.right_tables[self.upb_sheet_name].model().data(index)
                if not cell_data:
                    continue
                sample_id_value = str(cell_data).strip()

                # Update the left table
                self.left_table.blockSignals(True)
                self.left_table.setItem(r, sample_col, QTableWidgetItem(sample_id_value))  # Sample ID
                self.left_table.blockSignals(False)
        elif field == "Aliquot Name":
            for r in range(self.right_tables[self.upb_sheet_name].model().rowCount()):
                index = self.right_tables[self.upb_sheet_name].model().index(r, logical_index)
                cell_data = self.right_tables[self.upb_sheet_name].model().data(index, Qt.ItemDataRole.DisplayRole)
                if not cell_data:
                    continue
                aliquot_id_value = str(cell_data).strip()

                # Update the left table
                self.left_table.blockSignals(True)
                self.left_table.setItem(r, aliquot_col, QTableWidgetItem(aliquot_id_value))  # Aliquot ID
                self.left_table.blockSignals(False)
        elif field == "Grain Name":
            left_headers = [self.left_table.horizontalHeaderItem(i).text() for i in
                            range(self.left_table.columnCount())]
            # Only show a grain name column if a Grain Name is mapped
            if field == "Grain Name" and "Grain Name" not in left_headers:
                # Insert a new column for Grain Name at index 2, give it the header "Grain Name"
                self.left_table.insertColumn(2)
                self.left_table.setHorizontalHeaderItem(2, QTableWidgetItem("Grain Name"))
                spot_col = 3  # Spot ID is now at index 3
                upb_analysis_col = 4  # UPb Analysis Name is now at index 4
            for r in range(self.right_tables[self.upb_sheet_name].model().rowCount()):
                index = self.right_tables[self.upb_sheet_name].model().index(r, logical_index)
                cell_data = self.right_tables[self.upb_sheet_name].model().data(index, Qt.ItemDataRole.DisplayRole)
                if not cell_data:
                    continue
                grain_id_value = str(cell_data).strip()

                # Update the left table
                self.left_table.blockSignals(True)
                self.left_table.setItem(r, 2, QTableWidgetItem(grain_id_value))  # Grain ID
                self.left_table.blockSignals(False)
        elif field == "Spot Name":
            for r in range(self.right_tables[self.upb_sheet_name].model().rowCount()):
                index = self.right_tables[self.upb_sheet_name].model().index(r, logical_index)
                cell_data = self.right_tables[self.upb_sheet_name].model().data(index, Qt.ItemDataRole.DisplayRole)
                if not cell_data:
                    continue
                spot_id_value = str(cell_data).strip()

                # Update the left table
                self.left_table.blockSignals(True)
                self.left_table.setItem(r, spot_col, QTableWidgetItem(spot_id_value))  # Spot ID
                self.left_table.blockSignals(False)
        elif field == "UPb Analysis Name":
            for r in range(self.right_tables[self.upb_sheet_name].model().rowCount()):
                index = self.right_tables[self.upb_sheet_name].model().index(r, logical_index)
                cell_data = self.right_tables[self.upb_sheet_name].model().data(index, Qt.ItemDataRole.DisplayRole)
                if not cell_data:
                    continue
                upb_analysis_value = str(cell_data).strip()

                # Update the left table
                self.left_table.blockSignals(True)
                self.left_table.setItem(r, upb_analysis_col, QTableWidgetItem(upb_analysis_value))  # UPb Analysis Name
                self.left_table.blockSignals(False)
        self.left_table.resizeColumnsToContents()

    def update_left_table_background(self, item, color_hex):
        """
        Update the background color of a specific cell in the left table.
        """
        if item and color_hex:
            item.setBackground(QBrush(QColor(color_hex)))

    # def update_left_table_on_delimiter_change(self):
    #     """
    #     Update the left table's Sample ID and Spot ID columns whenever the delimiter value changes.
    #     """
    #     # Find the right table column mapped to "Spot ID"
    #
    #     if self.delimiter_checkbox.isChecked():
    #         spot_id_column = None
    #         for col_idx, (field_name) in self.sheet_mappings[self.current_sheet_name].items():
    #             if field_name == "Spot Name":
    #                 spot_id_column = col_idx
    #                 break
    #
    #         if spot_id_column is not None:
    #             self.auto_split_sample_spot(spot_id_column)
    #     else:
    #         spot_id_column = None
    #         for col_idx, (field_name) in self.sheet_mappings[self.current_sheet_name].items():
    #             if field_name == "Spot Name":
    #                 spot_id_column = col_idx
    #                 break
    #
    #         if spot_id_column is not None:
    #             row_count = self.right_table.model().rowCount()
    #             for r in range(row_count):
    #                 index = self.right_table.model().index(r, spot_id_column)
    #                 cell_data = self.right_table.model().data(index)
    #                 if not cell_data:
    #                     continue
    #
    #                 spot_id_value = str(cell_data).strip()
    #
    #                 # Update the left table
    #                 self.left_table.blockSignals(True)
    #                 self.left_table.setItem(r, 0, QTableWidgetItem(""))
    #                 self.left_table.setItem(r, 2, QTableWidgetItem(spot_id_value))  # Spot ID
    #                 self.left_table.blockSignals(False)
    #     # self.left_table.resizeColumnsToContents()
    #
    # def auto_split_sample_spot(self, col_idx):
    #     """
    #     Split the right table's Spot ID column values into Sample ID and Spot ID
    #     using the delimiter, and populate the left table accordingly.
    #     """
    #     delimiter = self.delimiter_edit.text().strip()
    #     row_count = self.right_table.model().rowCount()
    #
    #     for r in range(row_count):
    #         index = self.right_table.model().index(r, col_idx)
    #         cell_data = self.right_table.model().data(index)
    #         if not cell_data:
    #             continue
    #
    #         spot_id_value = str(cell_data).strip()
    #
    #         if delimiter in spot_id_value and delimiter:
    #             # Split based on the delimiter
    #             sample_id, spot_id = spot_id_value.split(delimiter, 1)
    #         else:
    #             # No delimiter found, treat the entire value as Spot ID
    #             sample_id = ""
    #             spot_id = spot_id_value
    #
    #         # Update the left table
    #         self.left_table.blockSignals(True)
    #         if not self.left_table.item(r, 0) or self.left_table.item(r, 0).text() == "":
    #             self.left_table.setItem(r, 0, QTableWidgetItem(sample_id))  # Sample ID
    #         else:
    #             # data already exists in this column, skip it and prefer user entered data.
    #             pass
    #         left_headers = [self.left_table.horizontalHeaderItem(i).text() for i in
    #                         range(self.left_table.columnCount())]
    #         if "Grain Name" in left_headers:
    #             spot_idx = 3
    #         else:
    #             spot_idx = 2
    #         self.left_table.setItem(r, spot_idx, QTableWidgetItem(spot_id))  # Spot ID
    #         self.left_table.blockSignals(False)
    #     # self.left_table.resizeColumnsToContents()

    def save_mapping(self):
        n_mappings = 0
        for sheet in self.sheet_mappings.keys():
            column_mappings = self.sheet_mappings[sheet]
            n_mappings += len(column_mappings)
        if n_mappings == 0:
            QMessageBox.warning(self, "No Mapping", "No columns have been mapped yet.")
            return

        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    try:
                        configs = json.load(f)
                    except json.JSONDecodeError:
                        # If the file is empty, it cannot load
                        configs = {}
            else:
                configs = {}
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save mapping:\n{e}")

        if configs == {}:
            name, ok = QInputDialog.getText(self, "Save Mapping", "Enter a name for this mapping:")
        else:
            items = list(configs.keys())
            recent_mappings = settings.value("recent_mappings", [])
            items_sorted = []
            if recent_mappings:
                for name in recent_mappings:
                    if name in items:
                        items_sorted.append(name)
            for name in items:
                if name not in items_sorted:
                    items_sorted.append(name)
            if not items:
                name, ok = QInputDialog.getText(self, "Save Mapping", "Enter a name for this mapping:")
            else:
                dlg = CompleterInputDialog(self, "Save Mapping", "Enter or select a name for this mapping:", items_sorted,
                                           True)
                if dlg.exec() == QDialog.DialogCode.Accepted:
                    name = dlg.get_input()
                    if not name:
                        return
                    else:
                        if name in items:
                            ok = True
                            reply = QMessageBox.question(self, "Overwrite Mapping",
                                                         f"Mapping '{name}' already exists. Overwrite?",
                                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                                         QMessageBox.StandardButton.No)
                            if reply != QMessageBox.StandardButton.Yes:
                                return
                        else:
                            ok = True
                else:
                    return

        if ok and name:
            sheets = {}
            for sheet, column_mappings in self.sheet_mappings.items():
                if column_mappings:
                    if sheet not in sheets:
                        sheets[sheet] = {}
                    for k, v in column_mappings.items():
                        if v not in ["Reference Display", "ReferenceID", "Instrument Name", "InstrumentID",
                                 "Lab Facility Name", "LabFacilityID", "UPb Analysis Method Name", "UPbAnalysisMethodID"]:
                            # Do not save the mappings for combo box added columns as these values rely on the database and can change
                            sheets[sheet][str(k)] = {"field": v}
            combos = {}
            for key, combo in self.static_combos.items():
                combos[key] = combo.currentText()
            configs[name] = {"Sheets": sheets, "Units/Formats": combos, "OriginalMappings": self.original_columns}
            with open(CONFIG_FILE, 'w') as f:
                json.dump(configs, f, indent=4)
            QMessageBox.information(self, "Saved", f"Mapping '{name}' saved successfully.")
            self.load_mapping(name)

    def load_mapping(self, name: str=''):
        logger_setup.get_logger().info("Loading Mapping")
        # Check if columns have been added or removed between loading the file and loading the first mapping
        if not self.mapping_loaded:
            # This is the first time loading a mapping
            new_original_columns = []
            for sheet in self.sheets.keys():
                if (not any(index != column for index, column in self.original_columns[sheet].items()) and
                        'added' not in self.original_columns[sheet].keys()):
                    # This should only be true if no changes have been made to the original columns
                    new_original_columns.append(False)

            if len(new_original_columns) != len(self.sheets.keys()):
                # Changes have been made to the original columns, so we cannot apply the loaded mapping of the original columns
                dlg = QMessageBox()
                dlg.setIcon(QMessageBox.Icon.Warning)
                dlg.setWindowTitle("Columns Modified")
                dlg.setText(f"Column additions/removals have been detected since loading the file.\n\n"
                            f"The saved mapping may not map correctly.\n"
                            "Do you want to continue loading the mapping?"
                            )
                continue_button = dlg.addButton("Continue", QMessageBox.ButtonRole.YesRole)
                cancel_button = dlg.addButton("Cancel", QMessageBox.ButtonRole.RejectRole)
                dlg.setDefaultButton(cancel_button)
                dlg.exec()
                if dlg.clickedButton() != continue_button:
                    return

        if not os.path.exists(CONFIG_FILE):
            QMessageBox.warning(self, "No Config", "No configuration file found.")
            return

        try:
            with open(CONFIG_FILE, 'r') as f:
                configs = json.load(f)
            if not configs:
                QMessageBox.warning(self, "No Mappings", "No mappings found in configuration.")
                return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load mapping:\n{e}")
            return

        # items = list(configs.keys())
        # recent_mappings = settings.value("recent_mappings", [])
        # items_sorted = []
        # if recent_mappings:
        #     for name in recent_mappings:
        #         if name in items:
        #             items_sorted.append(name)
        # for name in items:
        #     if name not in items_sorted:
        #         items_sorted.append(name)
        #
        # name, ok = QInputDialog.getItem(self, "Load Mapping", "Select a mapping to load:", items_sorted, 0, False)
        if not name:
            dlg = LoadMappingDialog(self)
            if dlg.exec() == QDialog.DialogCode.Accepted:
                name = dlg.selected_name
            else:
                return
        if not name:
            logger_setup.get_logger().info("No mapping selected.")
            return
        logger_setup.get_logger().info(f"Loading Mapping {name}")
        self.loading_manager.show_loading_dialog('Loading', f'Loading mapping: {name}...')
        loaded = configs[name]
        # Preserve mappings for columns added by combo box selections
        combo_mappings = {}
        for sheet, column_mappings in self.sheet_mappings.items():
            for col_idx, field_name in column_mappings.items():
                if col_idx not in loaded.get("Sheets", {}).get(sheet, {}):
                    if field_name in ["ReferenceID", "InstrumentID", "LabFacilityID", "UPbAnalysisMethodID"]:
                        # Preserve this mapping
                        if sheet not in combo_mappings:
                            combo_mappings[sheet] = {}
                        combo_mappings[sheet][col_idx] = field_name
                        # Save the name column mapping as well
                        combo_mappings[sheet][col_idx-1] = self.sheet_mappings[sheet][col_idx-1]
        self.sheet_mappings.clear()
        for sheet in self.sheets.keys():
            if sheet in combo_mappings:
                self.sheet_mappings[sheet] = combo_mappings[sheet].copy()
            else:
                self.sheet_mappings[sheet] = {}


        ## For the future, account for added and deleted columns when saving mappings
        # if "OriginalMappings" in loaded.keys():
        #     loaded_original = loaded["OriginalMappings"]
        #
        #
        #     if len(new_original_columns) == len(self.sheets.values()):
        #         # No changes have been made to the original columns, so we can safely apply the loaded mapping of the original columns
        #         logger_setup.get_logger().info("Original Mappings loaded.")
        #         # Ask they user if they want to automatically insert and remove columns or use the existing columns
        #         dlg = QMessageBox()
        #         dlg.setIcon(QMessageBox.Icon.Question)
        #         dlg.setWindowTitle("Add/Remove Columns")
        #         dlg.setText(f"The loaded mapping used a table with added/removed columns.\n\nWould you like to automatically add/remove columns to match the mapping or use existing columns?")
        #         add_remove_button = dlg.addButton("Add/Remove", QMessageBox.ButtonRole.YesRole)
        #         keep_existing_button = dlg.addButton("Keep Existing", QMessageBox.ButtonRole.NoRole)
        #         cancel_button = dlg.addButton("Cancel", QMessageBox.ButtonRole.RejectRole)
        #         dlg.setDefaultButton('Keep Existing')
        #         dlg.exec()
        #         if dlg.clickedButton() == add_remove_button:
        #             # Add/remove columns to match the original mapping
        #             for sheet in self.sheets.keys():
        #                 original = self.original_columns[sheet]
        #                 # First remove any deleted indexes
        #                 cols_to_remove = [idx for key, idx in original.items() if key != 'added' and idx == -1]
        #                 for col in sorted(cols_to_remove, reverse=True):
        #                     self.right_tables[sheet].model().removeColumn(col)
        #                 # Now insert any added indexes
        #                 cols_to_add = sorted(original['added'])
        #                 for col in cols_to_add:
        #                     self.right_tables[sheet].model().insertColumn(col)
        #             logger_setup.get_logger().info("Added/removed columns successfully.")
        #         elif dlg.clickedButton() == keep_existing_button:
        #             pass
        #         else:
        #             return


        if "Sheets" and "Units/Formats" in loaded.keys():
            loaded_sheets = loaded["Sheets"]
            loaded_combos = loaded["Units/Formats"]
            for sheet, mappings in loaded_sheets.items():
                if sheet in self.sheet_mappings.keys():
                    for k, v in mappings.items():
                        self.sheet_mappings[sheet][int(k)] = v["field"]
            for key, combo in self.static_combos.items():
                if key in loaded_combos:
                    val = loaded_combos[key]
                    idx = combo.findText(val)
                    if idx != -1:
                        combo.setCurrentIndex(idx)
                        # Trigger the change event for any combos with this signal
                        try: combo.closing.emit()
                        except Exception as e:
                            continue
        else:
            # GeoCORK v1.0.0 format, apply to the upb sheet only
            for k_str, v in loaded.items():
                self.sheet_mappings[self.upb_sheet_name][int(k_str)] = (v["field"])

        for sheet in self.sheet_mappings.keys():
            right_table = self.right_tables[sheet]
            total_cols = right_table.model().columnCount()
            column_mappings = self.sheet_mappings[sheet]
            for col_idx in range(total_cols):
                hdr_data = right_table.model().headerData(col_idx, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                if not hdr_data:
                    continue
                if col_idx in column_mappings and (sheet not in combo_mappings.keys() or col_idx not in combo_mappings[sheet]):
                    # Overwrite any existing mapping that is not from a combo box added column
                    f_name = column_mappings[col_idx]
                    right_table.model().setHeaderData(col_idx, Qt.Orientation.Horizontal, f"{f_name}", Qt.ItemDataRole.EditRole)
                    right_table.model().setHeaderData(col_idx, Qt.Orientation.Horizontal, QColor("#B8CFFF"), Qt.ItemDataRole.BackgroundRole)
                    # If it’s Sample Name / Aliquot Name / Spot Name / Grain Name, auto-populate left table
                    if f_name in ["Sample Name", "Aliquot Name", "Spot Name", "Grain Name", "UPb Analysis Name"] and sheet == self.upb_sheet_name:
                        self.update_left_table_on_header_change(f_name, col_idx)
                else:
                    if (right_table.model().headerData(col_idx, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                            not in ["Reference Display", "ReferenceID", "Instrument Name", "InstrumentID",
                                    "Lab Facility Name", "LabFacilityID", "UPb Analysis Method Name", "UPbAnalysisMethodID"]):
                        # Set text to column number and background to transparent
                        right_table.model().setHeaderData(col_idx, Qt.Orientation.Horizontal, str(col_idx), Qt.ItemDataRole.EditRole)
                        right_table.model().setHeaderData(col_idx, Qt.Orientation.Horizontal, Qt.GlobalColor.transparent, Qt.ItemDataRole.BackgroundRole)
                    else:
                        # This is a combo box added column, preserve its mapping but set background to transparent so the user knows it is not saved
                        right_table.model().setHeaderData(col_idx, Qt.Orientation.Horizontal, Qt.GlobalColor.transparent, Qt.ItemDataRole.BackgroundRole)
        self.update_mapping_list(name, configs)
        # self.right_table.resizeColumnsToContents()
        self.loading_manager.close_loading_dialog('Loading', f'Loading mapping: {name}...')
        logger_setup.get_logger().info(f"Mapping {name} loaded successfully.")
        QMessageBox.information(self, "Loaded", f"Mapping '{name}' loaded successfully.")
        self.mapping_loaded = True
        self.current_mapping = name

    def edit_mapping(self, name):
        ok, new_name = QInputDialog.getText(self, "Edit Mapping", "Enter new name for this mapping:")

    def update_mapping_list(self, mapping_name, configs):
        # Update the order of mappings in the list and config file
        recent_mappings = settings.value("recent_mappings", [])
        for name in configs.keys():
            if name not in recent_mappings:
                recent_mappings.append(name)
        if mapping_name not in recent_mappings:
            recent_mappings.insert(0, mapping_name)
        else:
            recent_mappings.remove(mapping_name)
            recent_mappings.insert(0, mapping_name)
        settings.setValue("recent_mappings", recent_mappings)

    def check_empty_cells_in_left_table(self):
        """
        Check for empty cells in the left table, excluding the "Grain Name" column which is optional.
        :return:
        List of tuples (row, col) for empty cells.
        """
        logger_setup.get_logger().info("Checking empty cells in left table")
        empty_cells = []
        disabled_rows = self.right_tables[self.upb_sheet_name].model().rows_for_status('disabled')
        for row in range(self.left_table.rowCount()):
            if row in disabled_rows:
                continue
            for col in range(self.left_table.columnCount()):
                cell = self.left_table.item(row, col)
                if (cell is None or cell.text().strip() in ["", 'NULL']) and self.left_table.horizontalHeaderItem(col).text() != "Grain Name":
                    empty_cells.append((row, col))
        return empty_cells

    def ask_to_use_default_values(self, empty_cells):
        """
        Prompt the user to confirm whether to use autogenerated default values for empty cells.
        """
        missing_count = len(empty_cells)
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Question)
        msg_box.setWindowTitle("Missing Values")
        msg_box.setText(f"{missing_count} required cells in the left UPb table are empty. Would you like to use autogenerated default values?")
        msg_box.setDetailedText(f"""
        Sample names cannot be missing. 
        Aliquot names will be set to their associated sample name if missing. 
        If only one of Spot names or UPb analysis names is missing, the other will be set to the existing name.
        If grain name is mapped but empty, it will be set to the spot name.
        Spot names and UPb Analysis names will be set based on Grain names or Aliquot names with a counter if both are missing.""")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if msg_box.exec() == QMessageBox.StandardButton.Yes:
            return True
        else:
            return False

    def fill_empty_cells_with_defaults(self, empty_cells):
        """
        Method to fill the left table with default values.
        Sample names cannot be missing.
        Aliquot names will be set to their associated sample name if missing.
        If only one of Spot names or UPb analysis names is missing, the other will be set to the existing name.
        If grain name is mapped but empty, it will be set to the spot name.
        Spot names and UPb Analysis names will be set based on Grain names or Aliquot names with a counter if both are missing.
        :param empty_cells: list of empty items
        """
        self.left_table.blockSignals(True)

        sample_col = None
        aliquot_col = None
        grain_col = None
        spot_col = None
        upb_analysis_col = None
        for column in range(self.left_table.columnCount()):
            if self.left_table.horizontalHeaderItem(column).text() == "Sample Name":
                sample_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Aliquot Name":
                aliquot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Grain Name":
                grain_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Spot Name":
                spot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "UPb Analysis Name":
                upb_analysis_col = column

        # Initialize variables for tracking SampleID and counter
        current_aliquot_name = None
        current_grain_name = None
        spot_counter = 0
        empty_rows = set(cell[0] for cell in empty_cells)
        for row in empty_rows:
            sample_item = self.left_table.item(row, sample_col)
            sample_name = sample_item.text().strip() if sample_item else ""
            aliquot_item = self.left_table.item(row, aliquot_col)
            aliquot_name = aliquot_item.text().strip() if aliquot_item else ""
            if grain_col is not None:
                grain_item = self.left_table.item(row, grain_col)
            else:
                grain_item = None
            grain_name = grain_item.text().strip() if grain_item else ""
            spot_item = self.left_table.item(row, spot_col)
            spot_name = spot_item.text().strip() if spot_item else ""
            upb_analysis_item = self.left_table.item(row, upb_analysis_col)
            upb_analysis_name = upb_analysis_item.text().strip() if upb_analysis_item else ""

            if sample_name in ["", 'NULL', None]:
                sample_item.setBackground(QColor("#FFB8B8"))
                logger_setup.get_logger().info(f"Sample Name empty in row {row}.") # Light red
                QMessageBox.warning(self, "No Sample Name", f"Please enter a Sample Name in row {row+1} (red) before proceeding.")
                self.left_table.scrollToItem(sample_item)
                return False
            if aliquot_name in ["", 'NULL', None]:
                # If Aliquot Name is missing, set equal to Sample Name value
                self.left_table.setItem(row, aliquot_col, QTableWidgetItem(sample_name))
                self.update_left_table_background(self.left_table.item(row, aliquot_col), "#D4B8FF")  # Light purple
            if spot_name in ["", 'NULL', None] and upb_analysis_name not in ["", 'NULL', None]:
                # If the Spot Name is missing but the UPb Analysis Name exists, set equal to UPb Analysis Name
                self.left_table.setItem(row, spot_col, QTableWidgetItem(upb_analysis_name))
                self.update_left_table_background(self.left_table.item(row, spot_col), "#D4B8FF")
            elif spot_name not in ["", 'NULL', None] and upb_analysis_name in ["", 'NULL', None]:
                # If the UPb Analysis Name is missing but the Spot Name exists, set equal to Spot Name
                self.left_table.setItem(row, upb_analysis_col, QTableWidgetItem(spot_name))
                self.update_left_table_background(self.left_table.item(row, upb_analysis_col), "#D4B8FF")
            elif grain_name not in ["", 'NULL', None] and spot_name in ["", 'NULL', None] and upb_analysis_name in ["", 'NULL', None]:
                # If the grain name exists but both spot name and upb analysis name are missing, set both to GrainName-counter
                if grain_name != current_grain_name:
                    current_grain_name = grain_name
                    spot_counter = 0  # Reset counter for new Grain Name
                spot_counter += 1
                self.left_table.setItem(row, spot_col, QTableWidgetItem(f"{grain_name}-{spot_counter}"))
                self.left_table.setItem(row, upb_analysis_col, QTableWidgetItem(f"{grain_name}-{spot_counter}"))
                self.update_left_table_background(self.left_table.item(row, spot_col), "#D4B8FF")
                self.update_left_table_background(self.left_table.item(row, upb_analysis_col), "#D4B8FF")
            elif grain_item and grain_name in ["", 'NULL', None] and spot_name in ["", 'NULL', None] and upb_analysis_name in ["", 'NULL', None]:
                # If the grain column is defined but grain name, spot name, and upb analysis name are missing, set all three to AliquotName-counter
                aliquot_name_item = self.left_table.item(row, aliquot_col)
                if aliquot_name_item and aliquot_name_item.text().strip() not in ["", 'NULL', None]:
                    aliquot_name = aliquot_name_item.text().strip()
                    if aliquot_name != current_aliquot_name:
                        current_aliquot_name = aliquot_name
                        spot_counter = 0  # Reset counter for new Aliquot Name
                    spot_counter += 1
                    self.left_table.setItem(row, grain_col, QTableWidgetItem(f"{aliquot_name}-{spot_counter}"))
                    self.left_table.setItem(row, spot_col, QTableWidgetItem(f"{aliquot_name}-{spot_counter}"))
                    self.left_table.setItem(row, upb_analysis_col, QTableWidgetItem(f"{aliquot_name}-{spot_counter}"))
                    self.update_left_table_background(self.left_table.item(row, grain_col), "#D4B8FF")
                    self.update_left_table_background(self.left_table.item(row, spot_col), "#D4B8FF")
                    self.update_left_table_background(self.left_table.item(row, upb_analysis_col), "#D4B8FF")
            spot_name = self.left_table.item(row, spot_col).text()
            if grain_item and self.left_table.item(row, grain_col).text() in ["", 'NULL', None] and spot_name not in ["", 'NULL', None]:
                # If grain column is defined but Grain Name is still missing, set it equal to the spot name
                self.left_table.setItem(row, grain_col, QTableWidgetItem(spot_name))
                self.update_left_table_background(self.left_table.item(row, grain_col), "#D4B8FF")

        self.left_table.blockSignals(False)
        # self.left_table.resizeColumnsToContents()
        return True

    def check_duplicates_in_left_table(self):
        """
        Check for duplicates in the left table for Sample Name, Aliquot Name, Spot Name, Grain Name, and UPb Analysis Name.
        First, check within the data to be imported, then check against the values in the database.
        :return:
        """

        # Look for duplicates within the left table
        logger_setup.get_logger().info("Checking for duplicates in left table")
        upb_analyses = []
        duplicates = {}
        grain_duplicates = {}
        sample_col = None
        aliquot_col = None
        spot_col = None
        grain_col = None
        upb_analysis_col = None
        disabled_rows = self.right_tables[self.upb_sheet_name].model().rows_for_status('disabled')

        for column in range(self.left_table.columnCount()):
            if self.left_table.horizontalHeaderItem(column).text() == "Sample Name":
                sample_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Aliquot Name":
                aliquot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Spot Name":
                spot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Grain Name":
                grain_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "UPb Analysis Name":
                upb_analysis_col = column
        for row in range(self.left_table.rowCount()):
            if row in disabled_rows:
                continue
            sample = self.left_table.item(row, sample_col).text().strip()
            aliquot = self.left_table.item(row, aliquot_col).text().strip()
            if grain_col:
                grain = self.left_table.item(row, grain_col).text().strip()
                if grain in ['NULL', '']:
                    grain = None
            else:
                grain = None
            spot = self.left_table.item(row, spot_col).text().strip()
            upb_analysis = self.left_table.item(row, upb_analysis_col).text().strip()
            if sample not in duplicates.keys():
                duplicates[sample] = {}
            if aliquot not in duplicates[sample].keys():
                duplicates[sample][aliquot] = {}
            if spot not in duplicates[sample][aliquot].keys():
                duplicates[sample][aliquot][spot] = []
            if spot not in grain_duplicates.keys():
                grain_duplicates[spot] = []
            duplicates[sample][aliquot][spot].append(upb_analysis)
            if grain:
                grain_duplicates[spot].append(grain)
            upb_analyses.append(upb_analysis)
        # check for duplicates in each list
        # Check for duplicate UPb Analysis Names
        upb_analysis_duplicates = list(set(
            [upb_analysis for upb_analysis in upb_analyses if upb_analyses.count(upb_analysis) > 1]))
        # Check for duplicate Spot Names with different Aliquot Names
        spot_duplicates = []
        grain_duplicates = []
        aliquot_duplicates = []
        sample_duplicates = []
        distinct_spots = set()
        distinct_aliquots = set()
        for sample in duplicates.keys():
            for aliquot in duplicates[sample].keys():
                for spot in duplicates[sample][aliquot].keys():
                    # Check for duplicate Spot Names with different Aliquot Names
                    if spot in distinct_spots:
                        spot_duplicates.append(spot)
                    else:
                        distinct_spots.add(spot)
                    # Check for different grains with the same spot
                    if grain_duplicates:
                        if len(grain_duplicates[spot]) > 1:
                            grain_duplicates.append(spot)
                # Check for duplicate Aliquot Names with different Sample Names
                if aliquot in distinct_aliquots:
                    aliquot_duplicates.append(aliquot)
                else:
                    distinct_aliquots.add(aliquot)
            # Check for different samples with the same aliquot
            aliquot_duplicates = []
            distinct_aliquots = set()
            for sample in duplicates.keys():
                for aliquot in duplicates[sample].keys():
                    if aliquot in distinct_aliquots:
                        aliquot_duplicates.append(aliquot)
                    else:
                        distinct_aliquots.add(aliquot)

        if upb_analysis_duplicates or spot_duplicates or grain_duplicates or aliquot_duplicates:
            logger_setup.get_logger().info("Duplicates found in left table")
            # Highlight the duplicate cells
            self.left_table.blockSignals(True)
            for row in range(self.left_table.rowCount()):
                if row in disabled_rows:
                    continue
                sample_item = self.left_table.item(row, sample_col)
                sample_name = sample_item.text().strip()
                aliquot_item = self.left_table.item(row, aliquot_col)
                aliquot_name = aliquot_item.text().strip()
                if aliquot_name in aliquot_duplicates:
                    self.update_left_table_background(aliquot_item, '#FFB8B8')  # Light red
                else:
                    aliquot_item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default
                spot_item = self.left_table.item(row, spot_col)
                spot_name = spot_item.text().strip()
                if spot_name in spot_duplicates:
                    self.update_left_table_background(spot_item, '#FFB8B8')  # Light red
                else:
                    spot_item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default
                if grain_col:
                    grain_item = self.left_table.item(row, grain_col)
                    grain_name = grain_item.text().strip()
                    if grain_name in grain_duplicates:
                        self.update_left_table_background(grain_item, '#FFB8B8')  # Light red
                    else:
                        grain_item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default
                upb_analysis_item = self.left_table.item(row, upb_analysis_col)
                upb_analysis_name = upb_analysis_item.text().strip()
                if upb_analysis_name in upb_analysis_duplicates:
                    self.update_left_table_background(upb_analysis_item, '#FFB8B8')  # Light red
                else:
                    upb_analysis_item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))
            QMessageBox(QMessageBox.Icon.Warning, f'Conflicts Detected',
                        f'Red cells in the left table are duplicates with different parent items\n\n'
                        'Ensure unique names before importing').exec()
            # Set the current tab to the UPb tab
            return False

        # Now that there are no duplicates in the import data, look for duplicates in the database
        logger_setup.get_logger().info("Checking for duplicates against database")
        # Prepare SQL queries for validation with existing database values
        # Find Samples where SampleName in the database matches import SampleName
        sample_query = QSqlQuery()
        sample_query.prepare("SELECT SampleID FROM Samples WHERE SampleName = :sample_name COLLATE NOCASE")

        # Find Aliquots and Samples where AliquotName matches in the database.
        aliquot_query = QSqlQuery()
        aliquot_query.prepare(
            "SELECT AliquotID, SampleID FROM Aliquots WHERE AliquotName = :aliquot_name COLLATE NOCASE")

        # Find Spots and Aliquots where SpotName matches in the database.
        spot_query = QSqlQuery()
        spot_query.prepare(
            "SELECT SpotID, AliquotID FROM Spots WHERE SpotName = :spot_name COLLATE NOCASE")

        # Find Grains and Spots where GrainName matches in the database.
        grain_query = QSqlQuery()
        grain_query.prepare(
            f'SELECT GrainID, SpotID FROM Spots JOIN Grains ON Spots.GrainID = Grains.GrainID\n'
            f'WHERE GrainName = :grain_name COLLATE NOCASE')

        # Find UPb Analyses and Spots where UPbAnalysisName matches in the database.
        upb_analysis_query = QSqlQuery()
        upb_analysis_query.prepare(
            "SELECT UPbAnalysisID, SpotID FROM UPbAnalyses WHERE UPbAnalysisName = :upb_analysis_name COLLATE NOCASE")

        distinct_samples = []
        aliquots_different_sample = []
        spots_different_aliquots = []
        grains_different_spots = []
        upb_analyses_different_spots = []
        sample_match = False
        aliquot_match = False
        spot_match = False
        grain_match = False
        upb_analysis_match = False
        conflicts = False

        for row in range(self.left_table.rowCount()):
            if row in disabled_rows:
                continue
            sample_name = self.left_table.item(row, sample_col).text()
            aliquot_name = self.left_table.item(row, aliquot_col).text()
            spot_name = self.left_table.item(row, spot_col).text()
            grain_name = self.left_table.item(row, grain_col).text() if grain_col else None
            upb_analysis_name = self.left_table.item(row, upb_analysis_col).text()

            self.left_table.blockSignals(True)
            # Highlight the row if any match is found
            # Highlight matching cells

            # Check Sample Name
            if sample_name in distinct_samples:
                sample_match = True
            else:
                sample_match = False
                sample_query.bindValue(':sample_name', sample_name)
                sample_match = sample_query.exec() and sample_query.next()
                sample_id = sample_query.value(0) if sample_match else None
                if sample_match:
                    distinct_samples.append(sample_name)

            if sample_match:
                item = self.left_table.item(row, sample_col)
                if item:
                    item.setBackground(QColor('#FFFAB8'))  # Light yellow
            else:
                item = self.left_table.item(row, sample_col)
                if item:
                    item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            # Check if any has already been identified as a conflict
            if aliquot_name in aliquots_different_sample:
                aliquot_match = True
            else:
                # Check Aliquot Name
                aliquot_match = False
                aliquot_query.bindValue(':aliquot_name', aliquot_name)
                aliquot_match = aliquot_query.exec() and aliquot_query.next()
                aliquot_id = aliquot_query.value(0) if aliquot_match else None
                aliquot_sample_id = aliquot_query.value(1) if aliquot_match else None
                # If the SampleID from the Aliquot does not match the SampleID from the Sample, the aliquot exists but for a different sample
                if aliquot_id:
                    if aliquot_sample_id and aliquot_sample_id != sample_id:
                        aliquots_different_sample.append(aliquot_name)

            if aliquot_match:
                item = self.left_table.item(row, aliquot_col)
                if item:
                    if aliquot_name in aliquots_different_sample:
                        item.setBackground(QColor('#FFB8B8'))  # Light red
                        conflicts = True
                    else:
                        item.setBackground(QColor('#FFFAB8'))  # Light yellow
            else:
                item = self.left_table.item(row, aliquot_col)
                if item:
                    item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            # Check Spot Name
            if spot_name in spots_different_aliquots:
                spot_match = True
            else:
                spot_match = False
                spot_query.bindValue(':spot_name', spot_name)
                spot_query.bindValue(':aliquot_id', aliquot_id)
                spot_match = spot_query.exec() and spot_query.next()
                spot_id = spot_query.value(0) if spot_match else None
                spot_aliquot_id = spot_query.value(1) if spot_match else None
                # If the AliquotID from the Spot does not match the AliquotID from the Aliquot, the spot exists but for a different aliquot
                if spot_id:
                    if spot_aliquot_id and spot_aliquot_id != aliquot_id:
                        spots_different_aliquots.append(spot_name)

            if spot_match:
                item = self.left_table.item(row, spot_col)
                if item:
                    if spot_name in spots_different_aliquots:
                        item.setBackground(QColor('#FFB8B8'))  # Light red
                        conflicts = True
                    else:
                        item.setBackground(QColor('#FFFAB8'))  # Light yellow
            else:
                item = self.left_table.item(row, spot_col)
                if item:
                    item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            # Check Grain Name
            if grain_name:
                if grain_name in grains_different_spots:
                    grain_match = True
                else:
                    grain_match = False
                    grain_query.bindValue(':grain_name', grain_name)
                    grain_query.bindValue(':spot_id', spot_id)
                    grain_match = grain_query.exec() and grain_query.next()
                    grain_id = grain_query.value(0) if grain_match else None
                    grain_spot_id = grain_query.value(1) if grain_match else None
                    # If the SpotID from the Grain does not match the SpotID from the Spot, the grain exists but for a different spot
                    if grain_id:
                        if grain_spot_id and grain_spot_id != spot_id:
                            grains_different_spots.append(grain_spot_id)

                if grain_match:
                    item = self.left_table.item(row, grain_col)
                    if item:
                        if grain_name in grains_different_spots:
                            item.setBackground(QColor('#FFB8B8'))  # Light red
                            conflicts = True
                        else:
                            item.setBackground(QColor('#FFFAB8'))  # Light yellow
                else:
                    item = self.left_table.item(row, grain_col)
                    if item:
                        item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            # Check UPb Analysis Name
            if upb_analysis_name in upb_analyses_different_spots:
                upb_analysis_match = True
            else:
                upb_analysis_match = False
                upb_analysis_query.bindValue(':upb_analysis_name', upb_analysis_name)
                upb_analysis_query.bindValue(':spot_id', spot_id)
                upb_analysis_match = upb_analysis_query.exec() and upb_analysis_query.next()
                upb_analysis_id = upb_analysis_query.value(0) if upb_analysis_match else None
                upb_analysis_spot_id = upb_analysis_query.value(1) if upb_analysis_match else None
                # If the SpotID from the UPb Analysis does not match the SpotID from the Spot, the UPb Analysis exists but for a different spot
                if upb_analysis_id:
                    if upb_analysis_spot_id and upb_analysis_spot_id != spot_id:
                        upb_analyses_different_spots.append(upb_analysis_name)

            if upb_analysis_match:
                item = self.left_table.item(row, upb_analysis_col)
                if item:
                    if upb_analysis_name in upb_analyses_different_spots:
                        item.setBackground(QColor('#FFB8B8'))  # Light red
                        conflicts = True
                    else:
                        item.setBackground(QColor('#FFFAB8'))  # Light yellow
            else:
                item = self.left_table.item(row, upb_analysis_col)
                if item:
                    item.setBackground(QBrush(Qt.GlobalColor.transparent))  # Reset to default

            self.left_table.blockSignals(False)

        if conflicts:
            logger_setup.get_logger().info("Conflict with left table values and database detected")
            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))
            message = QMessageBox()
            message.setIcon(QMessageBox.Icon.Warning)
            message.setWindowTitle('Conflicts Detected')
            message.setText('Yellow cells in the left table match existing entries in the database.\n'
                            'If these matches are not intended, rename with a unique name before importing.\n\n'
                            'Red cells in the left table already exist with a different parent item\n'
                            'Resolve red conflicts before importing')
            message.exec()
            return False
        elif any([sample_match, aliquot_match, spot_match, grain_match, upb_analysis_match]) and not self.import_clicked:
            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))
            logger_setup.get_logger().info("Matches between the left table and the database detected")
            message = QMessageBox()
            message.setIcon(QMessageBox.Icon.Information)
            message.setWindowTitle('Potential Conflicts Detected')
            message.setText('Yellow cells in the left table match existing entries in the database.\n\n'
                            'If these matches are not intended, rename with a unique name before importing.')
            message.exec()
        elif self.import_clicked:
            logger_setup.get_logger().info("Import clicked, no conflicts between the left table and the database detected")
        else:
            logger_setup.get_logger().info("No conflicts between the left table and the database detected")
            QMessageBox(QMessageBox.Icon.Information, f'No conflicts detected', f'No conflicts detected.')
        return True

    def check_duplicates_gps(self):
        """
        Checks for same Sample Name or Column Name with different GPS coordinates.
        :return:
        """

        def check_item_gps(item_gps_dictionary: str , item_import_header: str, item_db_header: str, item: str, items: str):
            logger_setup.get_logger().info(f"Checking for duplicate {items} with different GPS coordinates")

            for sheet, column_mappings in self.sheet_mappings.items():
                disabled_rows = self.right_tables[self.upb_sheet_name].model().rows_for_status('disabled')
                gps_columns = [list(column_mappings.keys())[list(column_mappings.values()).index(field)]
                               for field in SQLUtils.gps_possible_user_input_fields[item_gps_dictionary].keys()
                               if field in column_mappings.values()]
                db_gps_column_headers = get_headers('GPSLocations')
                gps_columns_db_equivalent = []
                for column in gps_columns:
                    field_name = column_mappings[column]
                    if field_name in SQLUtils.gps_possible_user_input_fields[item_gps_dictionary].keys():
                        gps_columns_db_equivalent.append(SQLUtils.gps_possible_user_input_fields[item_gps_dictionary][field_name][1])

                if item_import_header in column_mappings.values() and any(SQLUtils.gps_possible_user_input_fields[item_gps_dictionary].keys()):
                    item_col = list(column_mappings.keys())[list(column_mappings.values()).index(item_import_header)]
                    # Look for duplicate names in the column
                    duplicate_items = {}
                    item_names = {}
                    for row in range(self.right_tables[sheet].model().rowCount()):
                        if row in disabled_rows:
                            continue
                        index = self.right_tables[sheet].model().index(row, item_col)
                        item_name = self.right_tables[sheet].model().data(index)
                        if item_name not in item_names.keys():
                            item_names[item_name] = []
                            item_names[item_name].append(row)
                        else:
                            duplicate_items[item_name] = [row]

                    different_item_gps = False
                    if duplicate_items:
                        # Now check the GPS coordinates for each duplicate item
                        for item_name, gps_values1 in duplicate_items.items():
                            for row in item_names[item_name]:
                                gps_values2 = [self.right_tables[sheet].model().index(row, col).data() for col in gps_columns]
                                if set(gps_values1) != set(gps_values2):
                                    different_item_gps = True
                                    for col in gps_columns:
                                        index = self.right_tables[sheet].model().index(row, col)
                                        self.right_tables[sheet].model().setData(index, QColor('#FFB8B8'), Qt.ItemDataRole.ForegroundRole)  # Light red
                    if different_item_gps:
                        QMessageBox(QMessageBox.Icon.Warning, f'Conflicts Detected',
                                    f'Red cells in the right table indicate GPS coordinate conflicts for the same {item_import_header}.\n\n'
                                    'Resolve these red conflicts before importing')
                        logger_setup.get_logger().info(f"{items} with conflicting GPS coordinates detected in the right table")
                        # Set the current tab to the sheet with the conflict
                        self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[sheet]))
                        return False
                    logger_setup.get_logger().info(f"No {items} with conflicting GPS coordinates detected in the right table")

                    # Now check if these items are in the database with different GPS coordinates
                    logger_setup.get_logger().info(f"Checking for duplicate {items} with different GPS coordinates in the database")
                    query = QSqlQuery()
                    item_query = f'SELECT {item}GPSLocationID FROM {items} WHERE {item_db_header} = :item_name COLLATE NOCASE'
                    gps_query = f'SELECT * FROM GPSLocations WHERE GPSLocationID = :gps_id'
                    for item_name, row in item_names:
                        query.prepare(item_query)
                        query.bindValue(':item_name', item_name)
                        if not query.exec(item_query):
                            logger_setup.get_logger().error("Error checking for duplicates in the database")
                            logger_setup.get_logger().debug(f"Error: {query.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {item_query}")
                            return False
                        if query.next():
                            gps_id =query.value(0)
                            query.prepare(gps_query)
                            query.bindValue(':gps_id', gps_id)
                            if not query.exec(gps_query):
                                logger_setup.get_logger().error("Error checking for duplicates in the database")
                                logger_setup.get_logger().debug(f"Error: {query.lastError().text()}")
                                logger_setup.get_logger().debug(f"SQL query: {gps_query}")
                                return False
                            if query.next():
                                db_gps_values = [query.value(i) for i in range(1, query.record().count()) if query.record().fieldName(i) in get_headers('GPSLocations')]
                                gps_values_import = [self.right_tables[sheet].model().index(row, col).data() for col in gps_columns]
                                for header in db_gps_column_headers:
                                    if header in gps_columns_db_equivalent:
                                        import_idx = gps_columns_db_equivalent.index(header)
                                        db_value = db_gps_values[db_gps_column_headers.index(header)-1]
                                        import_value = gps_values_import[import_idx]
                                        if str(db_value) != str(import_value):
                                            different_item_gps = True
                                            for col in gps_columns:
                                                index = self.right_tables[sheet].model().index(row, col)
                                                self.right_tables[sheet].setData(index, QColor('#FFB8B8'), Qt.ItemDataRole.ForegroundRole)  # Light red
                    if different_item_gps:
                        QMessageBox(QMessageBox.Icon.Warning, f'Conflicts Detected',
                                    f'Red cells in the right table indicate GPS coordinate conflicts for the same {item_import_header} with existing database entries.\n\n'
                                    'Resolve these red conflicts before importing')
                        logger_setup.get_logger().info(f"{items} with conflicting GPS coordinates detected in the right table with existing database entries")
                        # Set the current tab to the sheet with the conflict
                        self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[sheet]))
                        return False

                logger_setup.get_logger().info("Checking for duplicate Columns with different GPS coordinates")
                if 'Column Name' in column_mappings.values() and any(SQLUtils.gps_possible_user_input_fields['Column GPS'].keys()):
                    column_col = list(column_mappings.keys())[list(column_mappings.values()).index('Column Name')]
                    # Look for duplicate Column Names in the column
                    duplicate_columns = {}
                    column_names = {}
                    for row in range(self.right_tables[sheet].model().rowCount()):
                        if row in disabled_rows:
                            continue
                        index = self.right_tables[sheet].model().index(row, column_col)
                        column_name = self.right_tables[sheet].data(index)
                        if column_name not in column_names.keys():
                            column_names[column_name] = []
                            column_names[column_name].append(row)
                        else:
                            duplicate_columns[column_name] = [row]

                    different_column_gps = False
                    if duplicate_columns:
                        # Now check the GPS coordinates for each duplicate column
                        gps_columns = [list(column_mappings.keys())[list(column_mappings.values()).index(field)]
                                       for field in SQLUtils.gps_possible_user_input_fields['Column GPS'].keys()
                                       if field in column_mappings.values()]
                        for column_name, gps_values1 in duplicate_columns.items():
                            for row in column_names[column_name]:
                                gps_values2 = [self.right_tables[sheet].model().index(row, col).data() for col in gps_columns]
                                if set(gps_values1) != set(gps_values2):
                                    different_column_gps = True
                                    for col in gps_columns:
                                        index = self.right_tables[sheet].model().index(row, col)
                                        self.right_tables[sheet].setData(index, QColor('#FFB8B8'), Qt.ItemDataRole.ForegroundRole)  # Light red
                    if different_column_gps:
                        QMessageBox(QMessageBox.Icon.Warning, f'Conflicts Detected',
                                    'Red cells in the right table indicate GPS coordinate conflicts for the same Column Name.\n\n'
                                    'Resolve these red conflicts before importing')
                        logger_setup.get_logger().info("Columns with conflicting GPS coordinates detected in the right table")
                        # Set the current tab to the sheet with the conflict
                        self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[sheet]))
                        return False
                logger_setup.get_logger().info("No Columns with conflicting GPS coordinates detected in the right table")
            return True

        if not check_item_gps(item_gps_dictionary='Sample GPS', item_import_header='Sample Name', item_db_header='SampleName', item='Sample', items='Samples'):
            return False
        if not check_item_gps(item_gps_dictionary='Column GPS', item_import_header='Column Name', item_db_header='ColumnName', item='Column', items='Columns'):
            return False
        return True


    def check_and_import(self) -> None:
            """
            Main method used when the import button is clicked. Used to first check the left table for empty values, prompt
            the user for default values, checks then for conflicts in the database. If all checks are passed then the list
            of rows to be imported is inserted into the database.
            """
            # First run the validation checks again
            self.import_clicked = True
            if not self.validate_data():
                return

            # Proceed with import
            self.import_to_db()


    def check_for_conflicts(self):
        """Checks values of SampleName, AliquotName, and SpotName in the left table for import against the database.
         Ensures no values are attempted to be inserted that could raise Unique Constraint Errors by the database.
         If values are found, the list of rows to be imported gets amended with primary IDs from inside the database."""
        # Find existing aliquot names belonging to other samples and existing spot names belonging to other aliquots

        sample_col = None
        aliquot_col = None
        spot_col = None
        grain_col = None
        upb_analysis_col = None
        disabled_rows = self.right_tables[self.upb_sheet_name].model().rows_for_status('disabled')
        for column in range(self.left_table.columnCount()):
            if self.left_table.horizontalHeaderItem(column).text() == "Sample Name":
                sample_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Aliquot Name":
                aliquot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Spot Name":
                spot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Grain Name":
                grain_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "UPb Analysis Name":
                upb_analysis_col = column

        existing_aliquots = set()
        existing_spots = set()
        existing_upb_analyses = set()
        existing_grains = set()
        for row_idx in range(self.left_table.rowCount()):
            if row_idx in disabled_rows:
                continue
            sample_name = self.left_table.item(row_idx, sample_col).text().strip()
            aliquot_name = self.left_table.item(row_idx, aliquot_col).text().strip()
            spot_name = self.left_table.item(row_idx, spot_col).text().strip()
            if grain_col:
                grain_name = self.left_table.item(row_idx, grain_col).text().strip()
            else:
                grain_name = None
            if not grain_name or grain_name in ['NULL', '']:
                grain_name = None
            upb_analysis_name = self.left_table.item(row_idx, upb_analysis_col).text().strip()
            # If the aliquot name exists in the database, does it have the same sample name?
            aliquot_id = self.find_matching_id('Aliquots', 'AliquotName', aliquot_name)
            if aliquot_id:
                query = QSqlQuery()
                query.prepare('SELECT SampleName FROM Aliquots JOIN Samples ON Aliquots.SampleID = Samples.SampleID WHERE AliquotID=:aliquot_id')
                query.bindValue(':aliquot_id', aliquot_id)
                if query.exec():
                    if query.next():
                        existing_sample_name = query.value(0)
                        if existing_sample_name != sample_name:
                            existing_aliquots.add(aliquot_name)
                            self.left_table.item(row_idx, aliquot_col).setBackground(QColor('#FFB8B8'))  # Light red
            spot_id = self.find_matching_id('Spots', 'SpotName', spot_name)
            if spot_id:
                query = QSqlQuery()
                query.prepare('SELECT AliquotName FROM Spots JOIN Aliquots ON Spots.AliquotID = Aliquots.AliquotID WHERE SpotID=:spot_id')
                query.bindValue(':spot_id', spot_id)
                if query.exec():
                    if query.next():
                        existing_aliquot_name = query.value(0)
                        if existing_aliquot_name != aliquot_name:
                            existing_spots.add(spot_name)
                            self.left_table.item(row_idx, spot_col).setBackground(QColor('#FFB8B8'))  # Light red
            if grain_name:
                grain_id = self.find_matching_id('Grains', 'GrainName', grain_name)
                if grain_id:
                    query = QSqlQuery()
                    query.prepare('SELECT SpotName FROM Grains JOIN Spots ON Grains.SpotID = Spots.SpotID WHERE GrainID=:grain_id')
                    query.bindValue(':grain_id', grain_id)
                    if query.exec():
                        if query.next():
                            existing_spot_name = query.value(0)
                            if existing_spot_name != spot_name:
                                existing_grains.add(grain_name)
                                self.left_table.item(row_idx, grain_col).setBackground(QColor('#FFB8B8'))  # Light red
            upb_analysis_id = self.find_matching_id('UPbAnalyses', 'UPbAnalysisName', upb_analysis_name)
            if upb_analysis_id:
                query = QSqlQuery()
                query.prepare('SELECT SpotName FROM UPbAnalyses JOIN Spots ON UPbAnalyses.SpotID = Spots.SpotID WHERE UPbAnalysisID=:upb_analysis_id')
                query.bindValue(':upb_analysis_id', upb_analysis_id)
                if query.exec():
                    if query.next():
                        existing_spot_name = query.value(0)
                        if existing_spot_name != spot_name:
                            existing_upb_analyses.add(upb_analysis_name)
                            self.left_table.item(row_idx, upb_analysis_col).setBackground(QColor('#FFB8B8'))  # Light red
        if existing_aliquots or existing_spots or existing_grains or existing_upb_analyses:
            msg = f"Items highlighted in red exist in the database with other parent data:\n"
            msg += "Please resolve these conflicts before importing."
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.setWindowTitle("Conflicts Found")
            msg_box.setText(msg)
            msg_box.addButton(QMessageBox.StandardButton.Ok)
            msg_box.exec()
            return False

        if existing_upb_analyses:
            msg = f"{len(existing_upb_analyses)}/{self.left_table.rowCount()} existing UPb analyses found for spots being imported."
            msg_box = QMessageBox()
            msg_box.setWindowTitle("Existing Values")
            msg_box.setText(msg)
            msg_box.addButton(QMessageBox.StandardButton.Cancel)
            skip_button = QPushButton("Skip duplicates")
            msg_box.addButton(skip_button, QMessageBox.ButtonRole.NoRole)
            overwrite_button = QPushButton("Overwrite duplicates")
            msg_box.addButton(overwrite_button, QMessageBox.ButtonRole.YesRole)
            msg_box.setDefaultButton(QMessageBox.StandardButton.Cancel)
            msg_box.exec()
            if msg_box.clickedButton() == QMessageBox.StandardButton.Cancel:
                return False
            elif msg_box.clickedButton() == skip_button:
                self.conflict_mode = 'skip'
                return True
            elif msg_box.clickedButton() == overwrite_button:
                self.conflict_mode = 'overwrite'
                return True
        return True

    def check_unmapped_references(self):
        """
        Checks if any columns are mapped to 'Reference Display' and if so, ensures that at least one other Reference field
        is also mapped in the same sheet(s).
        :return:
        """
        # Look for any columns mapped to 'Reference Display' and check if there are other columns mapped to other Reference fields
        reference_display_mapped = False
        for sheet, column_mapping in self.sheet_mappings.items():
            if "Reference Display" in column_mapping.values():
                reference_display_mapped = True
                break
        if reference_display_mapped:
            reference_field_mapped = False
            reference_fields = list(SQLUtils.reference_possible_user_input_fields['Reference'].keys())
            # Include the ID field for any columns populated from the combo box
            reference_fields.append('ReferenceID')
            # Remove Reference Display from the list
            reference_fields.remove('Reference Display')
            # Check if any other Reference fields are mapped in the same sheet as Reference Display
            for sheet, column_mapping in self.sheet_mappings.items():
                if any(field in column_mapping.values() for field in reference_fields):
                    if "Reference Display" in column_mapping.values():
                        reference_field_mapped = True
                        break
            if not reference_field_mapped:
                QMessageBox.warning(self, "Reference Mapping Error",
                                    f"Reference Display is mapped but no columns are mapped to other Reference fields in the same sheet(s).\n\n"
                                    "At least one other Reference field must be mapped (e.g., Authors, Year, Title, etc.) in the same sheet as Reference Display.")
                return False
        return True

    def check_unmapped_samples_columns(self):
        # Check if any Sample Info, Sample Age, or Sample GPS fields are mapped without 'Sample Name' in the same sheet
        sample_mapped = True
        sample_info = list(SQLUtils.sample_possible_user_input_fields['Sample Info'].keys())
        sample_info.remove('Sample Name')
        sample_age_info = list(SQLUtils.sample_possible_user_input_fields['Default Sample Age'].keys())
        sample_gps_info = list(SQLUtils.gps_possible_user_input_fields['Sample GPS'].keys())
        all_sample_info = []
        all_sample_info.extend(sample_info)
        all_sample_info.extend(sample_age_info)
        all_sample_info.extend(sample_gps_info)
        for sheet, column_mapping in self.sheet_mappings.items():
            if any(field in column_mapping.values() for field in all_sample_info):
                if "Sample Name" not in column_mapping.values():
                    sample_mapped = False
                    break
        if not sample_mapped:
            QMessageBox.warning(self, "Sample Mapping Error",
                                f"One or more columns mapped to Sample Info, Sample Age, or Sample GPS fields, but no columns are mapped to 'Sample Name' in the same sheet(s).\n\n"
                                "At least one column must be mapped to 'Sample Name' in the same sheet(s) as any Sample Info, Sample Age, or Sample GPS fields.")
            return False

        # Check if any Column GPS or Column Info fields are mapped without 'Column Name' in the same sheet
        column_mapped = True
        column_info = list(SQLUtils.column_possible_user_input_fields['Column Info'].keys())
        column_info.remove('Column Name')
        column_gps_info = list(SQLUtils.gps_possible_user_input_fields['Column GPS'].keys())
        all_column_info = []
        all_column_info.extend(column_info)
        all_column_info.extend(column_gps_info)
        for sheet, column_mapping in self.sheet_mappings.items():
            if any(field in column_mapping.values() for field in all_column_info):
                if "Column Name" not in column_mapping.values():
                    column_mapped = False
                    break
        if not column_mapped:
            QMessageBox.warning(self, "Column Mapping Error",
                                f"One or more columns mapped to Column Info or Column GPS fields, but no columns are mapped to 'Column Name' in the same sheet(s).\n\n"
                                "At least one column must be mapped to 'Column Name' in the same sheet(s) as any Column Info or Column GPS fields.")
            return False
        return True

    def check_unmapped_aliquots_spots_analyses(self):
        # Check if any Aliquot Info fields are mapped without 'Aliquot Name' in the same sheet
        aliquot_mapped = True
        aliquot_info = list(SQLUtils.aliquot_grain_spot_possible_user_input_fields['Aliquot Info'].keys())
        aliquot_info.remove('Aliquot Name')
        for sheet, column_mapping in self.sheet_mappings.items():
            if any(field in column_mapping.values() for field in aliquot_info):
                if "Aliquot Name" not in column_mapping.values():
                    aliquot_mapped = False
                    break
        if not aliquot_mapped:
            QMessageBox.warning(self, "Aliquot Mapping Error",
                                f"One or more columns mapped to Aliquot Info fields, but no columns are mapped to 'Aliquot Name' in the same sheet(s).\n\n"
                                "At least one column must be mapped to 'Aliquot Name' in the same sheet(s) as any Aliquot Info fields.")
            return False

        # Check if any Spot Info fields are mapped without 'Spot Name' in the same sheet
        spot_mapped = True
        spot_info = list(SQLUtils.aliquot_grain_spot_possible_user_input_fields['Spot Info'].keys())
        spot_info.remove('Spot Name')
        for sheet, column_mapping in self.sheet_mappings.items():
            if any(field in column_mapping.values() for field in spot_info):
                if "Spot Name" not in column_mapping.values():
                    spot_mapped = False
                    break
        if not spot_mapped:
            QMessageBox.warning(self, "Spot Mapping Error",
                                f"One or more columns mapped to Spot Info fields, but no columns are mapped to 'Spot Name' in the same sheet(s).\n\n"
                                "At least one column must be mapped to 'Spot Name' in the same sheet(s) as any Spot Info fields.")
            return False

        # Check if any Grain Info fields are mapped without 'Grain Name' in the same sheet
        grain_mapped = True
        grain_info = list(SQLUtils.aliquot_grain_spot_possible_user_input_fields['Grain Info'].keys())
        grain_info.remove('Grain Name')
        for sheet, column_mapping in self.sheet_mappings.items():
            if any(field in column_mapping.values() for field in grain_info):
                if "Grain Name" not in column_mapping.values():
                    grain_mapped = False
                    break
        if not grain_mapped:
            QMessageBox.warning(self, "Grain Mapping Error",
                                f"One or more columns mapped to Grain Info fields, but no columns are mapped to 'Grain Name' in the same sheet(s).\n\n"
                                "At least one column must be mapped to 'Grain Name' in the same sheet(s) as any Grain Info fields.")
            return False

        # Check if any UPb Analysis Info fields are mapped without 'UPb Analysis Name' in the same sheet
        upb_analysis_mapped = True
        upb_base_info = list(SQLUtils.upb_possible_user_input_fields['U-Pb Base Info'].keys())
        upb_base_info.remove('UPb Analysis Name')
        upb_ratio_info = list(SQLUtils.upb_possible_user_input_fields['Ratios'].keys())
        upb_age_info = list(SQLUtils.upb_possible_user_input_fields['Ages'].keys())
        upb_count_info = list(SQLUtils.upb_possible_user_input_fields['Isotope Counts'].keys())
        all_upb_info = []
        all_upb_info.extend(upb_base_info)
        all_upb_info.extend(upb_ratio_info)
        all_upb_info.extend(upb_age_info)
        all_upb_info.extend(upb_count_info)
        for sheet, column_mapping in self.sheet_mappings.items():
            if any(field in column_mapping.values() for field in all_upb_info):
                if "UPb Analysis Name" not in column_mapping.values():
                    upb_analysis_mapped = False
                    break
        if not upb_analysis_mapped:
            QMessageBox.warning(self, "U-Pb Analysis Mapping Error",
                                f"One or more columns mapped to U-Pb Analysis Info fields, but no columns are mapped to 'UPb Analysis Name' in the same sheet(s).\n\n"
                                "At least one column must be mapped to 'UPb Analysis Name' in the same sheet(s) as any U-Pb Analysis Info fields.")
            return False
        return True

    def check_static_table_fields(self):
        """
        Checks if any static fields in the left table are mapped and if so, tries to match to one of the existing values.
        If not match can be found, the user is asked to select the correct value from a dropdown.
        :return:
        """

        static_fields = {'Direct Age Error Format': 'ErrorFormats', 'Direct Age Unit': 'AgeUnits',
                         'Oldest Relative Age': 'Ages', 'Youngest Relative Age': 'Ages',
                         'Sample Latitude direction': 'DirectionUnits', 'Sample Longitude direction': 'DirectionUnits',
                         'Sample Elevation Unit': 'DistanceUnits', 'Column Latitude direction': 'DirectionUnits',
                         'Column Longitude direction': 'DirectionUnits', 'Column Elevation Unit': 'DistanceUnits',
                         'Column Total Height/Depth Unit': 'DistanceUnits', 'Sample Height/Depth Unit': 'DistanceUnits',
                         'Spot Size Unit': 'DistanceUnits', 'Ratio Error Format': 'ErrorFormats', 'Rejected': 'UPbAnalyses',
                         'Concordance Format': 'ConcordanceFormats', 'Age Error Format': 'ErrorFormats', 'Age Unit': 'AgeUnits'}

        query = QSqlQuery()
        ambiguous_static_values = {}

        for sheet, column_mappings in self.sheet_mappings.items():
            ambiguous_static_values[sheet] = {}
            disabled_rows = self.right_tables[sheet].model().rows_for_status('disabled')
            for column, field in column_mappings.items():
                if field in static_fields.keys():
                    static_table = static_fields[field]
                    if field == 'Rejected':
                        static_values = {}
                        static_values[0] = 'False'
                        static_values[1] = 'True'
                    else:
                        # Get the unique values in this column
                        static_name_header = get_headers(static_table)[get_name_column(static_table)]
                        id_header = get_headers(static_table)[0]
                        query.prepare(f'SELECT "{id_header}", "{static_name_header}" FROM "{static_table}"')
                        if not query.exec():
                            logger_setup.get_logger().critical(f'Could not load values from database')
                            logger_setup.get_logger().debug(f'Failed to query the {static_table} values')
                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                        static_values = {}
                        while query.next():
                            static_values[query.value(1)] = query.value(0)
                    unique_values = set()
                    for row in range(self.right_tables[sheet].model().rowCount()):
                        if row not in disabled_rows:
                            value = self.right_tables[sheet].model().index(row, column).data().strip()
                            if value:
                                unique_values.add(value)
                    if unique_values:
                        # Try to match each unique value to an existing value in the database
                        if column not in self.static_mappings[sheet].keys():
                            self.static_mappings[sheet][column] = {}
                        if column not in ambiguous_static_values[sheet].keys():
                            ambiguous_static_values[sheet][column] = {}
                        for value in unique_values:
                            if value not in self.static_mappings[sheet][column].keys():
                                if value in static_values.keys():
                                    value_id = static_values[value]
                                    self.static_mappings[sheet][column][value] = value_id
                                else:
                                    # Put together a list of ambiguous fields to ask the user about at the end
                                    ambiguous_static_values[sheet][column][value] = static_table

        if ambiguous_static_values:
            from ui.EditView import SetSelectedValues
            from Functions.Widget_classes import (CheckableTreeCombobox, CheckableComboBox, populate_combo_box)
            for sheet, columns in ambiguous_static_values.items():
                for column, values in columns.items():
                    if not values:
                        continue
                    else:
                        for value, static_table in values.items():
                            field = self.sheet_mappings[sheet][column]
                            if field == 'Rejected':
                                # Special case for Rejected field
                                combo = QComboBox()
                                combo.addItem('Accepted')
                                combo.addItem('Rejected')
                                dlg_text = f'Select the best match for {field} "{value}":'
                            else:
                                if static_table in SQLUtils.user_viewable_trees:
                                    combo = CheckableTreeCombobox()
                                else:
                                    combo = CheckableComboBox()
                                combo.model_modifiable = False
                                combo.enable_context_menu(False)
                                combo.set_single_click(True)
                                populate_combo_box(combo, **{'table': static_table})
                                dlg_text = f'{static_table} are fixed in the database.\n\nSelect the best match for {field} "{value}":'
                            dlg = SetSelectedValues(self, combo)
                            dlg.setWindowTitle(f'Select value for "{value}"')
                            dlg.main_layout.insertWidget(0, QLabel(dlg_text))
                            if dlg.exec() == QDialog.DialogCode.Accepted:
                                combo = dlg.widget
                                combo: CheckableTreeCombobox | CheckableComboBox
                                selected_value = combo.currentText()
                                if field == 'Rejected':
                                    # Special case for Rejected field (boolean)
                                    if selected_value == 'Rejected':
                                        selected_id = 1
                                    else:
                                        selected_id = 0
                                else:
                                    selected_id = get_id_from_name(static_table, selected_value)
                                if selected_value:
                                    self.static_mappings[sheet][column][value] = selected_id
                                else:
                                    self.static_mappings[sheet][column][value] = None
                            else:
                                cancel_dlg = QMessageBox()
                                cancel_dlg.setIcon(QMessageBox.Icon.Question)
                                cancel_dlg.setWindowTitle('Cancel or Skip')
                                cancel_dlg.setText('Do you want to cancel the validation or skip this value?')
                                cancel_dlg.addButton('Cancel Validation', QMessageBox.ButtonRole.RejectRole)
                                cancel_dlg.addButton('Skip Value', QMessageBox.ButtonRole.AcceptRole)
                                cancel_dlg.setDefaultButton(QMessageBox.StandardButton.Cancel)
                                if cancel_dlg.exec() == QDialog.DialogCode.Accepted:
                                    self.static_mappings[sheet][column][value] = None
                                else:
                                    return False

        return True

    def import_to_db(self):
        """
        Main method to import values in the QTableWidgets into the SQLite Database. Assumes using QSqlDatabase() default
        connection.
        """
        row_count = self.left_table.rowCount()
        if row_count == 0:
            dlg = QMessageBox.question(self, "No U-Pb Data",
                                       "There is no U-Pb data selected." "Continue without U-Pb data?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            dlg.exec_()
            if dlg.result() != QMessageBox.StandardButton.Yes:
                return
            else:
                upb_data = False
        else:
            upb_data = True

        # For each sheet, check the number of rows minus the disabled rows. Report the largest number of rows to be imported
        inserted_count = 0
        for sheet, table in self.sheet_mappings.items():
            sheet_row_count = self.right_tables[sheet].model().rowCount() - len(self.right_tables[sheet].model().rows_for_status('disabled'))
            if sheet_row_count > inserted_count:
                inserted_count = sheet_row_count

        item_tables = []
        many_tables = []
        for table in SQLUtils.database_ordered_tables:
            # if table in item_tables or table in analysis_tables or table in SQLUtils.static_tables:
            if table in SQLUtils.static_tables:
                continue
            elif "_" in table:
                # Many-to-many tables are in the format table1_table2
                many_tables.append(table)
            elif table == 'GPSLocations':
                item_tables.append('SampleGPSLocations')
                item_tables.append('ColumnGPSLocations')
            else:
                # tag_tables.append(table)
                item_tables.append(table)

        # We don't need to import the static table values, and their IDs are already mapped in self.static_mappings.

        create_savepoint('before_import')

        if upb_data:
            # Ensure the chain from Sample -> Aliquot -> Spot -> UPbAnalysis is imported correctly
            self.upb_imports = {'SampleID': [], 'AliquotID': [], 'SpotID': [], 'GrainID': [], 'UPbAnalysisID': []}
            if not self.import_upb_to_db():
                rollback_savepoint('before_import')
                logger_setup.get_logger().error('Import canceled')
                self.import_clicked = False
                return

        if not self.import_items_to_db(item_tables):
            rollback_savepoint('before_import')
            logger_setup.get_logger().error('Import canceled')
            self.import_clicked = False
            return

        if not self.import_many_to_many_to_db(many_tables):
            rollback_savepoint('before_import')
            logger_setup.get_logger().error('Import canceled')
            self.import_clicked = False
            return

        QMessageBox.information(self, "Success", f"Imported {inserted_count} rows into the database.")

        if not update_database():
            logger_setup.get_logger().critical('Error updating and displaying database')
            self.close()
        self.data_imported.emit(self.sample_ids)
        self.close()

    def import_upb_to_db(self):
        """
        Method to import the Samples, Aliquots, Spots, (Grains,) and UPbAnalyses names for the UPb table. Keeps a record
        of which IDs were imported so that other fields and tags will be filled when importing items.
        Assumes using QSqlDatabase() default connection.
        Also assumes that the left table has been populated with the Sample, Aliquot, Spot, (Grain,) and UPbAnalysis names
        :return:
        """

        from Functions.Alter_database import get_columns

        sample_col = None
        aliquot_col = None
        spot_col = None
        grain_col = None
        upb_analysis_col = None
        disabled_rows = self.right_tables[self.upb_sheet_name].model().rows_for_status('disabled')
        for column in range(self.left_table.columnCount()):
            if self.left_table.horizontalHeaderItem(column).text() == "Sample Name":
                sample_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Aliquot Name":
                aliquot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Spot Name":
                spot_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "Grain Name":
                grain_col = column
            elif self.left_table.horizontalHeaderItem(column).text() == "UPb Analysis Name":
                upb_analysis_col = column

        row_count = self.right_tables[self.upb_sheet_name].model().rowCount()
        import_count = row_count - len(disabled_rows)

        # Create a modal progress dialog
        progress_dialog = QProgressDialog(
            "Importing items...", "Cancel", 0, import_count, self
        )
        create_savepoint('before_upb_import')
        inserted_count = 0
        try:
            for row_idx in range(row_count):
                # Skip disabled rows
                if row_idx in disabled_rows:
                    continue
                progress_dialog.setValue(row_idx + 1)
                # Let the event loop process the dialog's updates
                QApplication.processEvents()
                # If the user clicked "Cancel", we can break out
                if progress_dialog.wasCanceled():
                    rollback_savepoint('before_upb_import')
                    return False

                # Build a record dict with every key initialized to None
                record = {field: None for field in SQLUtils.upb_possible_user_input_fields}

                # Populate the left-table items (sample_id, aliquot_id, spot_id)
                sample_id_item = self.left_table.item(row_idx, sample_col)
                aliquot_id_item = self.left_table.item(row_idx, aliquot_col)
                spot_id_item = self.left_table.item(row_idx, spot_col)
                grain_id_item = self.left_table.item(row_idx, grain_col) if grain_col else None
                upb_analysis_item = self.left_table.item(row_idx, upb_analysis_col)

                record["Sample Name"] = sample_id_item.text().strip() if sample_id_item else None
                record["Aliquot Name"] = aliquot_id_item.text().strip() if aliquot_id_item else None
                record["Spot Name"] = spot_id_item.text().strip() if spot_id_item else None
                record["Grain Name"] = grain_id_item.text().strip() if grain_id_item else None
                record["UPb Analysis Name"] = upb_analysis_item.text().strip() if upb_analysis_item else None

                # Find matching SampleID or create new
                if record["Sample Name"]:
                    logger_setup.get_logger().info(f"Sample Name: {record['Sample Name']}")
                    sample_query = QSqlQuery()
                    if not sample_query.prepare(
                        'SELECT SampleID FROM Samples WHERE SampleName=:name COLLATE NOCASE'):
                        logger_setup.get_logger().critical(f"Error importing Sample {record['Sample Name']}")
                        logger_setup.get_logger().debug(f"Failed to prepare query to find sample")
                        logger_setup.get_logger().debug(f"Error: {sample_query.lastError().text()}")
                        logger_setup.get_logger().debug(f"SQL query: {sample_query.lastQuery()}")
                        rollback_savepoint('before_upb_import')
                        return False
                    sample_query.bindValue(":name", record["Sample Name"])
                    if not sample_query.exec():
                        logger_setup.get_logger().critical(f"Error importing Sample {record['Sample Name']}")
                        logger_setup.get_logger().debug(f'Failed search for sample {record["Sample Name"]}')
                        logger_setup.get_logger().debug(f'Error: {sample_query.lastError().text()}')
                        logger_setup.get_logger().debug(f'SQL query: {sample_query.executedQuery()}')
                        rollback_savepoint('before_upb_import')
                        return False
                    sample_id = None
                    if sample_query.next():
                        sample_id = sample_query.value(0)
                    if sample_id:
                        # found matching samplename in database, will use that sample ID
                        record["SampleID"] = sample_id
                        self.sample_ids.append(record["SampleID"])
                        logger_setup.get_logger().info(f"Existing Sample: {record["Sample Name"]}")
                    else:
                        # no matching samplename in database, will create new one.
                        create_sample = QSqlQuery()
                        if not create_sample.prepare('INSERT INTO Samples (SampleName) VALUES (:name)'):
                            logger_setup.get_logger().critical(f"Error importing Sample {record['Sample Name']}")
                            logger_setup.get_logger().debug(f"Failed to prepare query to create sample")
                            logger_setup.get_logger().debug(f"Error: {create_sample.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {create_sample.executedQuery()}")
                            rollback_savepoint('before_upb_import')
                            return False
                        create_sample.bindValue(":name", record["Sample Name"])
                        if not create_sample.exec():
                        # if not create_sample.exec() and 'UNIQUE constraint failed' not in create_sample.lastError().text():
                            logger_setup.get_logger().critical(f"Error importing Sample {record['Sample Name']}")
                            logger_setup.get_logger().debug(f"Failed to execute query to create sample")
                            logger_setup.get_logger().debug(f"Error: {create_sample.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {create_sample.executedQuery()}")
                            logger_setup.get_logger().debug(f"Bound values: {create_sample.boundValues()}")
                            rollback_savepoint('before_upb_import')
                            return False
                        else:
                            record["SampleID"] = create_sample.lastInsertId()
                            self.sample_ids.append(record["SampleID"])
                            self.upb_imports['SampleID'].append(record["SampleID"])
                            logger_setup.get_logger().info(f"Imported Sample: {record['Sample Name']}")

                # Find matching Aliquot Name or create new
                if record["Aliquot Name"] and record["SampleID"]:
                    logger_setup.get_logger().info(f"Aliquot Name: {record['Aliquot Name']}")
                    aliquot_query = QSqlQuery()
                    if not aliquot_query.prepare(
                        'SELECT AliquotID, ParentAliquotID, AliquotParentRow, SampleID FROM Aliquots WHERE AliquotName=:name COLLATE NOCASE'):
                        logger_setup.get_logger().critical(f"Error importing Aliquot {record['Aliquot Name']}")
                        logger_setup.get_logger().debug(f"Failed to prepare query to find aliquot")
                        logger_setup.get_logger().debug(f"Error: {aliquot_query.lastError().text()}")
                        logger_setup.get_logger().debug(f"SQL query: {aliquot_query.lastQuery()}")
                        rollback_savepoint('before_upb_import')
                        return False
                    aliquot_query.bindValue(":name", record["Aliquot Name"])
                    if not aliquot_query.exec():
                        logger_setup.get_logger().critical(f"Error importing Aliquot {record['Aliquot Name']}")
                        logger_setup.get_logger().debug(f'Failed search for aliquot {record["Aliquot Name"]}')
                        logger_setup.get_logger().debug(f'Error: {aliquot_query.lastError().text()}')
                        logger_setup.get_logger().debug(f'SQL query: {aliquot_query.executedQuery()}')
                        logger_setup.get_logger().debug(f'Bound values: {aliquot_query.boundValues()}')
                        rollback_savepoint('before_upb_import')
                        return False
                    if aliquot_query.next():
                        # Check that existing aliquot matches the sample ID
                        if aliquot_query.value(3) != record["SampleID"]:
                            logger_setup.get_logger().error(
                                f'Aliquot {record["Aliquot Name"]} exists but is already associated with a different Sample.\nAliquot names must be unique.')
                            # Highlight the cell in the left table
                            self.left_table.item(row_idx, aliquot_col).setBackground(QColor('#FFB8B8'))  # Light red
                            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))
                            # scroll the left table to the row
                            self.left_table.scrollToItem(self.left_table.item(row_idx, aliquot_col))
                            rollback_savepoint('before_upb_import')
                            return False
                        # found matching aliquot name in database, will use that aliquot ID
                        record["AliquotID"] = aliquot_query.value(0)
                        record["AliquotParentRow"] = aliquot_query.value(2)
                        logger_setup.get_logger().info(f"Existing Aliquot: {record['Aliquot Name']}")
                    else:
                        # no matching aliquot in database, will create new one.
                        # Check if the sample has other aliquots to determine the parent row
                        query = QSqlQuery()
                        if not query.exec(
                                f'SELECT AliquotID, AliquotParentRow FROM Aliquots WHERE SampleID = {record["SampleID"]}'):
                            logger_setup.get_logger().critical(f'Error importing Aliquot {record["Aliquot Name"]}')
                            logger_setup.get_logger().debug(
                                f'Failed to query existing aliquots for sample ID {record["SampleID"]}')
                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                            logger_setup.get_logger().debug(f'values: {query.boundValues()}')
                            rollback_savepoint('before_upb_import')
                            return False
                        existing_rows = []
                        while query.next():
                            existing_rows.append(query.value(1))
                        if existing_rows:
                            record["AliquotParentRow"] = max(existing_rows) + 1
                        else:
                            record["AliquotParentRow"] = 0
                        create_aliquot = QSqlQuery()
                        if not create_aliquot.prepare(
                            'INSERT INTO Aliquots (AliquotName, AliquotParentRow, SampleID) VALUES (:name, :parent_row, :sample_id)'):
                            logger_setup.get_logger().critical(f"Error importing Aliquot {record['Aliquot Name']}")
                            logger_setup.get_logger().debug(f"Failed to prepare query to create aliquot")
                            logger_setup.get_logger().debug(f"Error: {create_aliquot.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {create_aliquot.executedQuery()}")
                            rollback_savepoint('before_upb_import')
                            return False
                        create_aliquot.bindValue(':name', record["Aliquot Name"])
                        create_aliquot.bindValue(':parent_row', record["AliquotParentRow"])
                        create_aliquot.bindValue(':sample_id', record["SampleID"])
                        if not create_aliquot.exec():
                        # if not create_aliquot.exec() and 'UNIQUE constraint failed' not in create_aliquot.lastError().text():
                            logger_setup.get_logger().critical(f"Error importing Aliquot {record['Aliquot Name']}")
                            logger_setup.get_logger().debug(f"Failed to create aliquot {record['Aliquot Name']}")
                            logger_setup.get_logger().debug(f"Error: {create_aliquot.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {create_aliquot.executedQuery()}")
                            logger_setup.get_logger().debug(f"values: {create_aliquot.boundValues()}")
                            rollback_savepoint('before_upb_import')
                            return False
                        else:
                            record["AliquotID"] = create_aliquot.lastInsertId()
                            self.upb_imports['AliquotID'].append(record["AliquotID"])
                            logger_setup.get_logger().info(f"Imported Aliquot: {record['Aliquot Name']}")


                # Find matching SpotID or create new
                if record["Spot Name"] and record["AliquotID"]:
                    logger_setup.get_logger().info(f"Spot Name: {record['Spot Name']}")
                    spot_query = QSqlQuery()
                    if not spot_query.prepare(
                        'SELECT SpotID, AliquotID FROM Spots WHERE SpotName=:name COLLATE NOCASE'):
                        logger_setup.get_logger().critical(f"Error importing Spot {record['Spot Name']}")
                        logger_setup.get_logger().debug(f"Failed to prepare query to find spot")
                        logger_setup.get_logger().debug(f"Error: {spot_query.lastError().text()}")
                        logger_setup.get_logger().debug(f"SQL query: {spot_query.lastQuery()}")
                        rollback_savepoint('before_upb_import')
                        return False
                    spot_query.bindValue(':name', record["Spot Name"])
                    if not spot_query.exec():
                        logger_setup.get_logger().critical(f"Error importing Spot {record['Spot Name']}")
                        logger_setup.get_logger().debug(f'Failed search for spot {record["Spot Name"]}')
                        logger_setup.get_logger().debug(f'Error: {spot_query.lastError().text()}')
                        logger_setup.get_logger().debug(f'SQL query: {spot_query.executedQuery()}')
                        logger_setup.get_logger().debug(f'values: {spot_query.boundValues()}')
                        rollback_savepoint('before_upb_import')
                        return False
                    if spot_query.next():
                        # Check that existing spot matches the aliquot ID
                        if spot_query.value(1) != record["AliquotID"]:
                            logger_setup.get_logger().error(
                                f'Spot {record["Spot Name"]} exists but is already associated with a different Aliquot.\nSpot names must be unique.')
                            # Highlight the cell in the left table
                            self.left_table.item(row_idx, spot_col).setBackground(QColor('#FFB8B8'))  # Light red
                            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))
                            # scroll the left table to the row
                            self.left_table.scrollToItem(self.left_table.item(row_idx, spot_col))
                            rollback_savepoint('before_upb_import')
                            return False
                        # found matching spot name in database, will use that spot ID
                        record["SpotID"] = spot_query.value(0)
                        logger_setup.get_logger().info(f"Existing Spot: {record['Spot Name']}")
                    else:
                        # no matching spot name in database, will create new one.
                        create_spot = QSqlQuery()
                        if not create_spot.prepare(
                            'INSERT INTO Spots (SpotName, AliquotID) VALUES (:name, :aliquot_id)'):
                            logger_setup.get_logger().critical(f"Error importing Spot {record['Spot Name']}")
                            logger_setup.get_logger().debug(f"Failed to prepare query to create spot")
                            logger_setup.get_logger().debug(f"Error: {create_spot.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {create_spot.executedQuery()}")
                            rollback_savepoint('before_upb_import')
                            return False
                        create_spot.bindValue(':name', record["Spot Name"])
                        create_spot.bindValue(':aliquot_id', record["AliquotID"])
                        if not create_spot.exec():
                        # if not create_spot.exec() and 'UNIQUE constraint failed' not in create_spot.lastError().text():
                            logger_setup.get_logger().critical(f"Error importing {record['Spot Name']}")
                            logger_setup.get_logger().debug(f"Failed to execute query to create spot")
                            logger_setup.get_logger().debug(f"Error: {create_spot.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {create_spot.executedQuery()}")
                            logger_setup.get_logger().debug(f"values: {create_spot.boundValues()}")
                            rollback_savepoint('before_upb_import')
                            return False
                        else:
                            record["SpotID"] = create_spot.lastInsertId()
                            self.upb_imports['SpotID'].append(record["SpotID"])
                            logger_setup.get_logger().info(f"Imported Spot: {record['Spot Name']}")


                # Find matching UPbAnalysisID or create new
                if record["UPb Analysis Name"] and record["SpotID"]:
                    logger_setup.get_logger().info(f"UPb Analysis Name: {record['UPb Analysis Name']}")
                    upb_query = QSqlQuery()
                    if not upb_query.prepare(
                        'SELECT UPbAnalysisID, SpotID FROM UPbAnalyses WHERE UPbAnalysisName=:name COLLATE NOCASE'):
                        logger_setup.get_logger().critical(f"Error importing UPbAnalysis {record['UPb Analysis Name']}")
                        logger_setup.get_logger().debug(f"Failed to prepare query to find UPbAnalysis")
                        logger_setup.get_logger().debug(f"Error: {upb_query.lastError().text()}")
                        logger_setup.get_logger().debug(f"SQL query: {upb_query.lastQuery()}")
                        rollback_savepoint('before_upb_import')
                        return False
                    upb_query.bindValue(':name', record["UPb Analysis Name"])
                    if not upb_query.exec():
                        logger_setup.get_logger().critical(f"Error importing {record['UPb Analysis Name']}")
                        logger_setup.get_logger().debug(f"Error searching for existing UPbAnalysis")
                        logger_setup.get_logger().debug(f"Error: {upb_query.lastError().text()}")
                        logger_setup.get_logger().debug(f"SQL query: {upb_query.executedQuery()}")
                        logger_setup.get_logger().debug(f"Bound values: {upb_query.boundValues()}")
                        rollback_savepoint('before_upb_import')
                        return False
                    if upb_query.next():
                        # Check that existing UPbAnalysis matches the spot ID
                        if upb_query.value(1) != record["SpotID"]:
                            logger_setup.get_logger().error(
                                f'UPb Analysis {record["UPb Analysis Name"]} exists but is already associated with a different Spot.\nUPb Analysis names must be unique.')
                            # Highlight the cell in the left table
                            self.left_table.item(row_idx, upb_analysis_col).setBackground(QColor('#FFB8B8'))  # Light red
                            self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[self.upb_sheet_name]))
                            # scroll the left table to the row
                            self.left_table.scrollToItem(self.left_table.item(row_idx, upb_analysis_col))
                            rollback_savepoint('before_upb_import')
                            return False
                        # found matching UPb Analysis name in database, will use that UPb Analysis ID
                        record["UPbAnalysisID"] = upb_query.value(0)
                        logger_setup.get_logger().info(f"Existing UPb Analysis: {record['UPb Analysis Name']}")
                    else:
                        record["UPbAnalysisID"] = None
                    if not record["UPbAnalysisID"]:
                        insert_sql = f'INSERT INTO UPbAnalyses (UPbAnalysisName, SpotID) VALUES (:name, :spot_id)'
                        insert_query = QSqlQuery()
                        if not insert_query.prepare(insert_sql):
                            logger_setup.get_logger().critical(f"Error importing UPb Analysis {record['UPb Analysis Name']}")
                            logger_setup.get_logger().debug(f"Failed to prepare data for spot {record['Spot Name']}")
                            logger_setup.get_logger().debug(f"Error: {insert_query.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {insert_query.executedQuery()}")
                            rollback_savepoint('before_upb_import')
                            return False
                        insert_query.bindValue(':name', record["UPb Analysis Name"])
                        insert_query.bindValue(':spot_id', record["SpotID"])
                        if not insert_query.exec():
                        # if not insert_query.exec() and 'UNIQUE constraint failed' not in insert_query.lastError().text():
                            logger_setup.get_logger().critical(f"Error importing UPb Analysis {record['UPb Analysis Name']}")
                            logger_setup.get_logger().debug(f"Failed to insert data for spot {record['Spot Name']}")
                            logger_setup.get_logger().debug(f"Error: {insert_query.lastError().text()}")
                            logger_setup.get_logger().debug(f"SQL query: {insert_query.executedQuery()}")
                            logger_setup.get_logger().debug(f"Values: {insert_query.boundValues()}")
                            rollback_savepoint('before_upb_import')
                            return False

                        record['UPbAnalysisID'] = insert_query.lastInsertId()
                        self.upb_imports['UPbAnalysisID'].append(record["UPbAnalysisID"])
                        logger_setup.get_logger().info(f"Imported UPb Analysis: {record['UPb Analysis Name']}")

                        inserted_count += 1

        except Exception as e:
            logger_setup.get_logger().debug(f"Error: {e}")
            rollback_savepoint('before_upb_import')
            return False
        # QSqlDatabase().commit()
        logger_setup.get_logger().info(f"Imported {inserted_count} UPb Analysis records")
        release_savepoint('before_upb_import')
        return True


    def import_items_to_db(self, item_tables):
        """
        Main method to import items from the item_tables list of tables. Works whether or not there are analyses.
        Data for each table is collated from all sheets then imported in the order of item_tables, which should be in
        parent->child order.
        :param item_tables:
        :return:
        """

        from Functions.Widget_classes import get_columns

        combo_values = {
            'GPSElevUnitID': self.elevation_unit_combobox.itemData(
                self.elevation_unit_combobox.currentIndex())
            ,
            'HeightDepthUnitID': self.heightdepth_unit_combobox.itemData(
                self.heightdepth_unit_combobox.currentIndex())
            ,
            'TotalHeightDepthUnitID': self.heightdepth_unit_combobox.itemData(
                self.heightdepth_unit_combobox.currentIndex())
            ,
            'DirectAgeErrorFormatID': self.sample_age_error_combobox.itemData(
                self.sample_age_error_combobox.currentIndex())
            ,
            'DirectAgeUnitID': self.age_unit_combobox.itemData(self.age_unit_combobox.currentIndex())
            ,
            'AgeUnitID': self.age_unit_combobox.itemData(self.age_unit_combobox.currentIndex())
            ,
            'AgeErrorFormatID': self.upb_age_error_combobox.itemData(
                self.upb_age_error_combobox.currentIndex())
            ,
            'RatioErrorFormatID': self.ratio_error_combobox.itemData(
                self.ratio_error_combobox.currentIndex())
            ,
            'SpotSizeUnitID': self.spot_size_unit_combobox.itemData(
                self.spot_size_unit_combobox.currentIndex())
            ,
            'ConcordanceFormatID': self.conc_format_combobox.itemData(
                self.conc_format_combobox.currentIndex())
        }

        query = QSqlQuery()

        organize_progress_dialog = QProgressDialog(
            "Organizing data...", "Cancel", 0, len(item_tables), self
        )
        organize_count = 0

        self.item_ids = {}
        item_data = {}
        logger_setup.get_logger().info(f'Organizing data for import')
        for table in item_tables:
            organize_count += 1
            organize_progress_dialog.setValue(organize_count)
            # Let the event loop process the dialog's updates
            QApplication.processEvents()
            # If the user clicked "Cancel", we can break out
            if organize_progress_dialog.wasCanceled():
                logger_setup.get_logger().info('Canceled organizing data')
                return False
            # Check if any columns are mapped to this table in any sheet
            if table == 'SampleGPSLocations':
                db_table = 'GPSLocations'
                prefix = 'Sample'
            elif table == 'ColumnGPSLocations':
                db_table = 'GPSLocations'
                prefix = 'Column'
            else:
                db_table = table
                prefix = None
            logger_setup.get_logger().info(f'Importing tags for {table}')
            for sheet, column_mappings in self.sheet_mappings.items():
                logger_setup.get_logger().info(f'Importing tags for {table} from sheet {sheet}')
                if column_mappings:
                    for column, field in column_mappings.items():
                        # retrieve the table the field represents
                        for field_dict in self.field_dictionaries:
                            for category, fields in field_dict.items():
                                if field in fields.keys():
                                    if db_table in fields[field]:
                                        if prefix and prefix not in field:
                                            # This field is for a different prefix, so skip it
                                            continue
                                        item_header = fields[field][1]
                                        if table not in self.item_ids.keys():
                                            self.item_ids[table] = {}
                                        if sheet not in self.item_ids[table].keys():
                                            self.item_ids[table][sheet] = {}
                                        if column not in self.item_ids[table][sheet].keys():
                                            self.item_ids[table][sheet][column] = {}
                                        self.item_ids[table][sheet][column][item_header] = {}
            if self.hidden_mappings:
                for column, id_header in self.hidden_mappings.items():
                    id_header = list(id_header.keys())[0]
                    if id_header == get_headers(db_table)[0]:
                        # This is the ID column for this table, it was added by the combo box to the U-Pb sheet
                        if table not in self.item_ids.keys():
                            self.item_ids[table] = {}
                        if self.upb_sheet_name not in self.item_ids[table].keys():
                            self.item_ids[table][self.upb_sheet_name] = {}
                        if column not in self.item_ids[table][self.upb_sheet_name].keys():
                            self.item_ids[table][self.upb_sheet_name][column] = {}
                        self.item_ids[table][self.upb_sheet_name][column][id_header] = {}
                        break

        import_progress_dialog = QProgressDialog(
            "Importing data...", "Cancel", 0, len(self.item_ids.keys())+1, self
        )
        create_savepoint('before_import_items')
        import_table_count = 0

        for table in self.item_ids.keys():
            import_progress_dialog.setLabelText(f'Importing {table}')
            import_table_count += 1
            import_progress_dialog.setValue(import_table_count)
            # Let the event loop process the dialog's updates
            QApplication.processEvents()
            # If the user clicked "Cancel", we can break out
            if import_progress_dialog.wasCanceled():
                logger_setup.get_logger().info('Canceled importing data')
                rollback_savepoint('before_import_items')
                return False
            logger_setup.get_logger().info(f'Importing {table}')
            table_name_columns = {}
            if table == 'SampleGPSLocations' or table == 'ColumnGPSLocations':
                db_table = 'GPSLocations'
            else:
                db_table = table
            query, virtual, stored, columns = get_columns(db_table)
            # Get a list of all columns except the ID column. This list already excludes calculated and automatically set columns
            query_columns = [col.replace('"', '') for col in columns if f'"{get_headers(db_table)[0]}"' not in col and
                             'Created' not in col and 'Modified' not in col and 'Parent' not in col and 'Converted' not in col]
            item_data[table] = {}
            foreign_keys = {}
            if not query.exec(f'PRAGMA foreign_key_list("{db_table}")'):
                logger_setup.get_logger().critical(f'Could not retrieve related table info for {table}')
                logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                rollback_savepoint('before_import_items')
                return False
            while query.next():
                foreign_keys[query.value(3)] = {'foreign_table': query.value(2), 'foreign_column': query.value(4)}
            if table not in ['SampleAges', 'References', 'SampleGPSLocations', 'ColumnGPSLocations']:
                name_header = get_headers(table)[get_name_column(table)]
            elif table == 'References' and search_dictionary(self.item_ids[table], get_headers(db_table)[0]):
                # ID column is mapped, so make sure that ReferenceDisplay is too
                if search_dictionary(self.item_ids[table], 'ReferenceDisplay'):
                    name_header = 'ReferenceDisplay'
                else:
                    name_header = None
            else:
                name_header = None
            for sheet in self.item_ids[table].keys():
                name_columns = []
                disabled_rows = self.right_tables[sheet].model().rows_for_status('disabled')
                rejected_rows = self.right_tables[sheet].model().rows_for_status('rejected')
                if name_header:
                    for column, item_header in self.item_ids[table][sheet].items():
                        item_header = list(item_header.keys())[0]
                        if item_header == name_header:
                            name_columns.append(column)
                else:
                    parent_table = None
                    if self.item_parent_dict[table]:
                        # For tables with no name column, we need to find the parent name column
                        parent_table = self.item_parent_dict[table]
                    if not parent_table:
                        name_columns = []
                    parent_name_header = get_headers(parent_table)[get_name_column(parent_table)]
                    for column, item_header in self.item_ids[parent_table][sheet].items():
                        item_header = list(item_header.keys())[0]
                        if item_header == parent_name_header:
                            name_columns.append(column)
                    if not name_columns:
                        logger_setup.get_logger().critical(f"Unknown mapping for table {table} items")
                        logger_setup.get_logger().debug(f"Could not find name header {name_header} or  parent column {parent_name_header} in sheet {sheet}")
                        rollback_savepoint('before_import_items')
                        return False
                if name_header and not search_dictionary(self.item_ids[table][sheet], name_header):
                    # This table requires a name column, but none of the mapped columns correspond to the name column
                    logger_setup.get_logger().error(
                        f'Cannot import info into {table} without a value for the {name_header} field in sheet {sheet}')
                    rollback_savepoint('before_import_items')
                    self.workbook_tabs.setCurrentIndex(self.workbook_tabs.indexOf(self.right_tables[sheet]))
                    return False
                # Get any data directly related to the columns in this table
                table_name_columns[sheet] = name_columns
                if name_header not in item_data[table].keys():
                    item_data[table][name_header] = {}
                for row in range(self.right_tables[sheet].model().rowCount()):
                    if row in disabled_rows:
                        continue
                    for name_column in name_columns:
                        item_name = self.right_tables[sheet].model().index(row, name_column).data()
                        logger_setup.get_logger().info(f'Gathering data for {table} item "{item_name}" from sheet {sheet}, row {row}')
                        item_id = None
                        if item_name in ['NULL', '', None]:
                            continue
                        if item_name not in item_data[table][name_header].keys():
                            item_data[table][name_header][item_name] = {}
                        if sheet == self.upb_sheet_name and self.hidden_mappings and search_dictionary(
                                self.item_ids[table][sheet], get_headers(db_table)[0]):
                            # This table has an ID column added to the U-Pb sheet via the hidden mappings, so get the ID directly
                            for column, header in self.item_ids[table][sheet].items():
                                header = list(header.keys())[0]
                                if header == get_headers(db_table)[0]:
                                    item_id = self.right_tables[sheet].model().index(row, column).data()
                                    break
                            if item_id:
                                # Now set the other columns to this ID value
                                for column, header in self.item_ids[table][sheet].items():
                                    header = list(header.keys())[0]
                                    item_data[table][name_header][item_name][header] = item_id
                                    self.item_ids[table][sheet][column][header][item_name] = item_id
                        else:
                            for column, item_header in self.item_ids[table][sheet].items():
                                item_header = list(item_header.keys())[0]
                                if (item_header not in item_data[table][name_header][item_name].keys() or
                                        item_name not in self.item_ids[table][sheet][column][item_header].keys()):
                                    self.item_ids[table][sheet][column][item_header][
                                        item_name] = None  # Placeholder, will set after insert
                                    if item_header == name_header:
                                        item = item_name
                                    else:
                                        item_input = self.right_tables[sheet].model().index(row, column).data()
                                        if item_header == 'Rejected':
                                            item = self.static_mappings[sheet][column][item_input]
                                        else:
                                            item = item_input
                                    if item_header in foreign_keys.keys():
                                        # This column is a foreign key
                                        foreign_table = foreign_keys[item_header]['foreign_table']
                                        if foreign_table in SQLUtils.static_tables:
                                            item = self.static_mappings[sheet][column][item_input]
                                        else:
                                            if foreign_table == 'GPSLocations':
                                                if table == 'Samples':
                                                    item = self.item_ids['SampleGPSLocations'][sheet][column][item_input]
                                                elif table == 'Columns':
                                                    item = self.item_ids['ColumnGPSLocations'][sheet][column][item_input]
                                            else:
                                                item = self.item_ids[foreign_table][sheet][column][item_input]
                                    item_data[table][name_header][item_name][item_header] = item
                            for item_header in foreign_keys.keys():
                                if item_header not in item_data[table][name_header][item_name].keys() or item_data[table][name_header][item_name][item_header] in ['NULL', '', None]:
                                    foreign_table = foreign_keys[item_header]['foreign_table']
                                    foreign_query, foreign_virtual, foreign_stored, foreign_columns = get_columns(foreign_table)
                                    # Get a list of all columns except the ID column. This list already excludes calculated and automatically set columns
                                    foreign_query_columns = [col.replace('"', '') for col in foreign_columns if
                                                     f'"{get_headers(foreign_table)[0]}"' not in col and
                                                     'Created' not in col and 'Modified' not in col and 'Parent' not in col and 'Converted' not in col]
                                    if get_headers(foreign_table)[get_name_column(foreign_table)] in foreign_query_columns:
                                        foreign_name_header = get_headers(foreign_table)[get_name_column(foreign_table)]
                                        foreign_name_item = None
                                    elif (foreign_table == 'References' and 'References' in self.item_ids.keys() and
                                          search_dictionary(self.item_ids[foreign_table][sheet], 'ReferenceDisplay')):
                                        foreign_name_header = 'ReferenceDisplay'
                                        foreign_name_item = None
                                    else:
                                        # This should be the parent table of the foreign table, so we can use its name column
                                        foreign_name_header = name_header
                                        foreign_name_item = item_name
                                    if foreign_table == 'GPSLocations':
                                        if table == 'Samples':
                                            search_table = 'SampleGPSLocations'
                                        elif table == 'Columns':
                                            search_table = 'ColumnGPSLocations'
                                        else:
                                            search_table = foreign_table
                                    else:
                                        search_table = foreign_table
                                    if search_table in self.item_ids.keys():
                                        if sheet in self.item_ids[search_table].keys():
                                            for column, header in self.item_ids[search_table][sheet].items():
                                                header = list(header.keys())[0]
                                                if foreign_name_item:
                                                    column_data = self.item_ids[search_table][sheet][column][header][foreign_name_item]
                                                    break
                                                elif header == foreign_name_header:
                                                    foreign_name_item = self.right_tables[sheet].model().index(row, column).data()
                                                    if foreign_name_item in self.item_ids[foreign_table][sheet][column][header].keys():
                                                        column_data = self.item_ids[foreign_table][sheet][column][header][foreign_name_item]
                                                        break
                                                    else:
                                                        column_data = 'NULL'
                                        else:
                                            column_data = 'NULL'
                                    elif item_header in combo_values.keys():
                                        # This is a static value for all rows
                                        column_data = combo_values[item_header]
                                    else:
                                        column_data = 'NULL'
                                    item_data[table][name_header][item_name][item_header] = column_data
                            if table == 'UPbAnalyses':
                                if 'Rejected' not in item_data[table][name_header][item_name].keys():
                                    if row in rejected_rows:
                                        item_data[table][name_header][item_name]['Rejected'] = 1
                                    else:
                                        item_data[table][name_header][item_name]['Rejected'] = 0
                                else:
                                    if (item_data[table][name_header][item_name]['Rejected'] == 'Rejected'
                                        or item_data[table][name_header][item_name]['Rejected'] == 1):
                                        item_data[table][name_header][item_name]['Rejected'] = 1
                                    elif (item_data[table][name_header][item_name]['Rejected'] == 'Accepted'
                                          or item_data[table][name_header][item_name]['Rejected'] == 0):
                                        item_data[table][name_header][item_name]['Rejected'] = 0

            import_table_count = 0
            item_progress_dialog = QProgressDialog(
                f"Importing {len(item_data[table][name_header].keys())} {table}...", "Cancel", 0, len(item_data[table][name_header].keys()), self
            )
            item_count = 0

            # Now insert the items into the database
            for item_name in item_data[table][name_header].keys():
                item_count += 1
                item_progress_dialog.setValue(item_count)
                # Let the event loop process the dialog's updates
                QApplication.processEvents()
                # If the user clicked "Cancel", we can break out
                if item_progress_dialog.wasCanceled():
                    logger_setup.get_logger().info('Canceled importing data')
                    rollback_savepoint('before_import_items')
                    return False
                input_values = {}
                item_id = None
                all_null = True
                for column_header in query_columns:
                    if column_header in item_data[table][name_header][item_name].keys():
                        input_values[column_header] = item_data[table][name_header][item_name][column_header]
                    else:
                        input_values[column_header] = 'NULL'
                    if 'ID' not in column_header:
                        if all_null and input_values[column_header] not in ['NULL', '', None]:
                            all_null = False
                if item_name in ['NULL', '', None]:
                    if all_null:
                        # This is a blank entry with no data, so skip it
                        continue
                if 'GPSLocations' in table:
                    gps_format_id = self.determine_gps_format(input_values)
                    if not gps_format_id:
                        logger_setup.get_logger().error(f'Could not determine GPS format for {table} entry "{item_name}"')
                        rollback_savepoint('before_import_items')
                        return False
                    input_values['GPSFormatID'] = gps_format_id
                if get_headers(db_table)[get_name_column(db_table)] in query_columns:
                    name = item_name
                else:
                    name = None
                # Search for existing item in database
                if get_headers(db_table)[0] in item_data[table][name_header][item_name].keys():
                    # The ID column is mapped, so use this ID
                    item_id = item_data[table][name_header][item_name][get_headers(db_table)[0]]
                    for sheet in self.item_ids[table].keys():
                        for column, header in self.item_ids[table][sheet].items():
                            header_name = list(header.keys())[0]
                            if item_name in self.item_ids[table][sheet][column][header_name].keys():
                                self.item_ids[table][sheet][column][header_name][item_name] = item_id
                    continue
                else:
                    if name:
                        search_query = f'SELECT {get_headers(db_table)[0]} FROM "{db_table}" WHERE {name_header} = :name COLLATE NOCASE'
                        if not query.prepare(search_query):
                            logger_setup.get_logger().critical(f'Error importing {table} "{name}"')
                            logger_setup.get_logger().debug(f'Failed to prepare query to find existing item')
                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                            rollback_savepoint('before_import_items')
                            return False
                        query.bindValue(':name', name)
                    else:
                        # This table does not have a name field. Search for all values instead
                        non_null = False
                        search_query = f'SELECT {get_headers(db_table)[0]} FROM "{db_table}" WHERE '
                        for column_name in query_columns:
                            if column_name in item_data[table][name_header][item_name].keys():
                                value = item_data[table][name_header][item_name][column_name]
                                if value in ['NULL', '', None]:
                                    search_query += f'"{column_name}" IS NULL AND '
                                else:
                                    search_query += f'"{column_name}" = :{column_name} COLLATE NOCASE AND '
                                    non_null = True
                            else:
                                search_query += f'"{column_name}" IS NULL AND '
                        search_query = search_query[:-5]  # Remove the last " AND "
                        if not non_null:
                            # All values are null, so skip this entry
                            continue
                        if not query.prepare(search_query):
                            logger_setup.get_logger().critical(f'Error importing {table} entry')
                            logger_setup.get_logger().debug(f'Failed to prepare query to find existing item')
                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                            rollback_savepoint('before_import_items')
                            return False
                        for column_name in query_columns:
                            if column_name in item_data[table][name_header][item_name].keys():
                                value = item_data[table][name_header][item_name][column_name]
                                if value not in ['NULL', '', None]:
                                    query.bindValue(f':{column_name}', value)
                    if not query.exec():
                        logger_setup.get_logger().critical(f'Could not search for existing {table} in database')
                        logger_setup.get_logger().debug(f'Failed to query the {table} values')
                        logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                        logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                        logger_setup.get_logger().debug(f'Bound values: {query.boundValues()}')
                        rollback_savepoint('before_import_items')
                        return False
                    if query.next():
                        item_id = query.value(0)
                    else:
                        item_id = None
                if item_id:
                    # The item already exists in the database. Follow conflict resolution strategy
                    if search_dictionary(self.item_ids[table], get_headers(db_table)[0]):
                        # The ID column is mapped and assigned above, so assume this is the correct item
                        continue
                    if table in ['SampleAges', 'References', 'SampleGPSLocations', 'ColumnGPSLocations']:
                        # Have already checked it against all values
                        for sheet in self.item_ids[table].keys():
                            for column, header in self.item_ids[table][sheet].items():
                                header_name = list(header.keys())[0]
                                if item_name in self.item_ids[table][sheet][column][header_name].keys():
                                    self.item_ids[table][sheet][column][header_name][item_name] = item_id
                        continue
                    else:
                        # Now check all values to see if they are identical
                        existing_values = {}
                        query_column_list = ', '.join([f'"{col}"' for col in query_columns])
                        existing_query = f'SELECT {query_column_list} FROM "{db_table}" WHERE "{get_headers(db_table)[0]}" = {item_id}'
                        if not query.prepare(existing_query):
                            logger_setup.get_logger().critical(f'Error importing {table} "{name}"')
                            logger_setup.get_logger().debug(f'Failed to prepare query to find existing item values')
                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                            rollback_savepoint('before_import_items')
                            return False
                        if not query.exec():
                            logger_setup.get_logger().critical(f'Error importing {table} "{name}"')
                            logger_setup.get_logger().debug(f'Failed to query the {table} values')
                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                            logger_setup.get_logger().debug(f'Bound values: {query.boundValues()}')
                            rollback_savepoint('before_import_items')
                            return False
                        if query.next():
                            for col_idx in range(query.record().count()):
                                col_name = query.record().fieldName(col_idx)
                                if not query.value(col_idx):
                                    existing_values[col_name] = 'NULL'
                                else:
                                    existing_values[col_name] = query.value(col_idx)
                        if existing_values == input_values:
                            # All values are identical, so update the ID in all mapping places and move on
                            logger_setup.get_logger().info(f'Skipping identical existing {table} "{name}"')
                            for sheet in self.item_ids[table].keys():
                                for column, header in self.item_ids[table][sheet].items():
                                    header_name = list(header.keys())[0]
                                    # Only update if the item_name exists in this mapping (it may not if multiple name columns are used)
                                    if item_name in self.item_ids[table][sheet][column][header_name].keys():
                                        self.item_ids[table][sheet][column][header_name][item_name] = item_id
                            continue
                        elif ((get_headers(table)[0] in self.upb_imports.keys() and item_id in self.upb_imports[get_headers(table)[0]])
                              or self.conflict_mode == 'add to'):
                            # Add on to what was imported with the UPb chain in the previous method
                            update_values = {}
                            for column in query_columns:
                                if existing_values[column] == 'NULL' and input_values[column] != 'NULL':
                                    update_values[column] = input_values[column]
                            if db_table in SQLUtils.trigger_tables:
                                error, header = validate_update(db_table, list(update_values.keys()),
                                                                list(update_values.values()),
                                                                f'{get_headers(db_table)[0]} = {item_id}')
                                if error:
                                    logger_setup.get_logger().error(f'Error updating {db_table} for {name}: {error}')
                                    logger_setup.get_logger().debug(
                                        f'Error validating data to update for table {table}')
                                    rollback_savepoint('before_import_items')
                                    return False
                            if update_values:
                                update_query = f'UPDATE "{db_table}" SET '
                                for column in update_values.keys():
                                    column_placeholder = f':{column.replace('/', '').replace('*', '').replace(' ', '_')}'
                                    update_query += f'"{column}" = {column_placeholder}, '
                                update_query = update_query[:-2]  # Remove the last ", "
                                update_query += f' WHERE {get_headers(db_table)[0]} = {item_id}'
                                if not query.prepare(update_query):
                                    logger_setup.get_logger().critical(f'Error importing {table} "{name}"')
                                    logger_setup.get_logger().debug(f'Failed to prepare query to update existing item')
                                    logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                                    logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                                    rollback_savepoint('before_import_items')
                                    return False
                                for column in update_values.keys():
                                    column_placeholder = f':{column.replace('/', '').replace('*', '').replace(' ', '_')}'
                                    if input_values[column] not in ['NULL', '', None]:
                                        query.bindValue(f'{column_placeholder}', input_values[column])
                                    else:
                                        query.bindValue(f'{column_placeholder}', QVariant())
                                if not query.exec():
                                    logger_setup.get_logger().critical(f'Could not update existing {table} in the database')
                                    logger_setup.get_logger().debug(f'Failed to update values in {table} for ID {item_id}')
                                    logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                                    logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                                    logger_setup.get_logger().debug(f'Bound values: {query.boundValues()}')
                                    rollback_savepoint('before_import_items')
                                    return False
                            for sheet in self.item_ids[table].keys():
                                for column, header in self.item_ids[table][sheet].items():
                                    header_name = list(header.keys())[0]
                                    if item_name in self.item_ids[table][sheet][column][header_name].keys():
                                        self.item_ids[table][sheet][column][header_name][item_name] = item_id
                        elif self.conflict_mode == 'skip':
                            logger_setup.get_logger().info(f'Skipping duplicate {table} "{name}"')
                            for sheet in self.item_ids[table].keys():
                                for column, header in self.item_ids[table][sheet].items():
                                    header_name = list(header.keys())[0]
                                    if item_name in self.item_ids[table][sheet][column][header_name].keys():
                                        self.item_ids[table][sheet][column][header_name][item_name] = item_id
                            if table not in self.skipped_conflict_ids.keys():
                                self.skipped_conflict_ids[db_table] = []
                            self.skipped_conflict_ids[db_table].append(item_id)
                            continue
                        elif self.conflict_mode == 'overwrite':
                            logger_setup.get_logger().info(f'Overwriting existing {table} "{name}"')
                            # Delete the existing item and re-insert it
                            delete_query = f'DELETE FROM "{db_table}" WHERE {get_headers(db_table)[0]} = {item_id}'
                            if not query.prepare(delete_query):
                                logger_setup.get_logger().critical(f'Error importing {table} "{name}"')
                                logger_setup.get_logger().debug(f'Failed to prepare query to delete existing item')
                                logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                                logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                                rollback_savepoint('before_import_items')
                                return False
                            if not query.exec():
                                logger_setup.get_logger().critical(f'Could not overwrite existing {table} in the database')
                                logger_setup.get_logger().debug(f'Failed to delete values from {table} for ID {item_id}')
                                logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                                logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                                logger_setup.get_logger().debug(f'Bound values: {query.boundValues()}')
                                rollback_savepoint('before_import_items')
                                return False
                            item_id = None

                if not item_id:
                    logger_setup.get_logger().info(f'Inserting new entry into {table} for "{item_name}"')
                    if db_table in SQLUtils.user_viewable_trees:
                        # For tree tables, we need to get the max parent row of the root and add one
                        root_query = f'SELECT {get_headers(db_table)[2]} FROM "{db_table}" WHERE {get_headers(db_table)[1]} IS NULL ORDER BY {get_headers(db_table)[2]} DESC LIMIT 1'
                        if not query.prepare(root_query):
                            logger_setup.get_logger().critical(f'Error importing {table} "{item_name}"')
                            logger_setup.get_logger().debug(f'Failed to prepare query to find maximum parent row for root')
                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                            rollback_savepoint('before_import_items')
                            return False
                        if not query.exec():
                            logger_setup.get_logger().critical(f'Could not search for existing {table} in database')
                            logger_setup.get_logger().debug(f'Failed to get the maximum parent row for root in {table}')
                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                            logger_setup.get_logger().debug(f'Bound values: {query.boundValues()}')
                            rollback_savepoint('before_import_items')
                            return False
                        if query.next():
                            parent_row = query.value(0) + 1
                        else:
                            parent_row = 0
                        input_values[get_headers(db_table)[1]] = 'NULL'
                        input_values[get_headers(db_table)[2]] = parent_row
                    if db_table in SQLUtils.trigger_tables:
                        if 'GPSFormatID' in input_values.keys():
                            error, header = validate_insert(db_table, list(input_values.keys()), list(input_values.values()), input_values['GPSFormatID'])
                        else:
                            error, header = validate_insert(db_table, list(input_values.keys()), list(input_values.values()), None)
                        if error:
                            logger_setup.get_logger().error(f'Error inserting new {db_table} for {name}: {error}')
                            logger_setup.get_logger().debug(f'Error validating data to insert for table {table}')
                            rollback_savepoint('before_import_items')
                            return False
                    insert_query = f'INSERT INTO "{db_table}" ('
                    for column, value in input_values.items():
                        insert_query += f'"{column}", '
                    insert_query = insert_query[:-2]  # Remove the last ", "
                    insert_query += ') VALUES ('
                    for column, value in input_values.items():
                        column = f':{column.replace('/', '').replace('*', '').replace(' ', '_')}'
                        insert_query += f'{column}, '
                    insert_query = insert_query[:-2]  # Remove the last ", "
                    insert_query += ')'
                    if not query.prepare(insert_query):
                        logger_setup.get_logger().critical(f'Error importing {table} "{item_name}"')
                        logger_setup.get_logger().debug(f'Failed to prepare query to insert new item')
                        logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                        logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                        rollback_savepoint('before_import_items')
                        return False
                    for column, value in input_values.items():
                        if value in ['NULL', '', None]:
                            value = QVariant()
                        column = f':{column.replace('/', '').replace('*', '').replace(' ', '_')}'
                        query.bindValue(column, value)
                    if not query.exec():
                    # if not query.exec() and 'UNIQUE constraint failed' not in query.lastError().text():
                        logger_setup.get_logger().critical(f'Error importing {table} "{item_name}"')
                        logger_setup.get_logger().debug(f'Failed to insert values into {table}')
                        logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                        logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                        logger_setup.get_logger().debug(f'Bound values: {query.boundValues()}')
                        rollback_savepoint('before_import_items')
                        return False
                    item_id = query.lastInsertId()
                    for sheet in self.item_ids[table].keys():
                        for column, header in self.item_ids[table][sheet].items():
                            header_name = list(header.keys())[0]
                            if item_name in self.item_ids[table][sheet][column][header_name].keys():
                                self.item_ids[table][sheet][column][header_name][item_name] = item_id
        release_savepoint('before_import_items')
        return True



    def import_many_to_many_to_db(self, many_tables):
        """
        Method to import many-to-many relationships for database tables using the dictionary made during import.
        """
        create_savepoint('before_import_many_to_many')
        linking_progress_dialog = QProgressDialog(
            "Linking tables...", "Cancel", 0, len(many_tables), self
        )
        linking_count = 0
        for table in many_tables:
            linking_count += 1
            linking_progress_dialog.setValue(linking_count)
            # Let the event loop process the dialog's updates
            QApplication.processEvents()
            # If the user clicked "Cancel", we can break out
            if linking_progress_dialog.wasCanceled():
                logger_setup.get_logger().info('Canceled linking tables')
                rollback_savepoint('before_import_many_to_many')
                return False
            table1 = table.split('_')[0]
            table2 = table.split('_')[1]
            if table1 not in self.item_ids.keys() or table2 not in self.item_ids.keys():
                # One of the tables is not being imported, so skip it
                continue
            for sheet in self.item_ids[table1].keys():
                if sheet not in self.item_ids[table2].keys():
                    # The two tables are not on the same sheet, so skip it
                    continue
                disabled_rows = self.right_tables[sheet].model().rows_for_status('disabled')
                table1_id = None
                table2_id = None
                table1_name_header, table1_name_columns = self.find_name_header(table1, sheet)
                table2_name_header, table2_name_columns = self.find_name_header(table2, sheet)
                if not table1_name_header or not table2_name_header:
                    logger_setup.get_logger().critical(f'Could not find names for table {table} in database')
                if not table1_name_columns or not table2_name_columns:
                    logger_setup.get_logger().critical(f'Could not find name columns for table {table} in database')
                    continue
                for row in range(self.right_tables[sheet].model().rowCount()):
                    if row in disabled_rows:
                        continue
                    for table1_name_column in table1_name_columns:
                        table1_name = self.right_tables[sheet].model().index(row, table1_name_column).data()
                        for column1, header1 in self.item_ids[table1][sheet].items():
                            header1_name = list(header1.keys())[0]
                            if table1_name in self.item_ids[table1][sheet][column1][header1_name].keys():
                                table1_id = self.item_ids[table1][sheet][column1][header1_name][table1_name]
                                if not table1_id:
                                    logger_setup.get_logger().critical(
                                        f'Could not link {table1} "{table1_name}" in sheet {sheet} with data from other sheets')
                                    continue
                                else:
                                    for table2_name_column in table2_name_columns:
                                        table2_name = self.right_tables[sheet].model().index(row, table2_name_column).data()
                                        if table1_name in ['NULL', '', None] or table2_name in ['NULL', '', None]:
                                            continue
                                        for column2, header2 in self.item_ids[table2][sheet].items():
                                                header2_name = list(header2.keys())[0]
                                                if table2_name in self.item_ids[table2][sheet][column2][header2_name].keys():
                                                    table2_id = self.item_ids[table2][sheet][column2][header2_name][table2_name]
                                                    if not table2_id:
                                                        logger_setup.get_logger().critical(f'Could not link {table2} "{table2_name}" in sheet {sheet} with data from other sheets')
                                                        continue
                                                    else:
                                                        if table1_id in self.skipped_conflict_ids.get(table1, []) or table2_id in self.skipped_conflict_ids.get(table2, []):
                                                            # One of the items was skipped due to conflict resolution, so skip this relationship
                                                            continue
                                                        # Check if this relationship already exists in the database
                                                        query = QSqlQuery()
                                                        table1_id_field = get_headers(table1)[0]
                                                        table2_id_field = get_headers(table2)[0]
                                                        if not query.prepare(f'SELECT {table1_id_field} FROM {table} WHERE {table1_id_field} = :table1_id AND {table2_id_field} = :table2_id'):
                                                            logger_setup.get_logger().critical(f'Error linking {table1} "{table1_name}" and {table2} "{table2_name}"')
                                                            logger_setup.get_logger().debug(f'Failed to prepare query to find existing relationship')
                                                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                                                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                                                            rollback_savepoint('before_import_many_to_many')
                                                            return False
                                                        query.bindValue(':table1_id', table1_id)
                                                        query.bindValue(':table2_id', table2_id)
                                                        if not query.exec():
                                                            logger_setup.get_logger().critical(f'Error linking {table1} "{table1_name}" and {table2} "{table2_name}"')
                                                            logger_setup.get_logger().debug(f'Failed to query the {table} values')
                                                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                                                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                                                            logger_setup.get_logger().debug(f'Bound values: {query.boundValues()}')
                                                            rollback_savepoint('before_import_many_to_many')
                                                            return False
                                                        if query.next():
                                                            # The relationship already exists, so skip it
                                                            continue
                                                        # Insert the new relationship
                                                        if not query.prepare(f'INSERT INTO {table} ({table1_id_field}, {table2_id_field}) VALUES (:table1_id, :table2_id)'):
                                                            logger_setup.get_logger().critical(f'Error linking {table1} "{table1_name}" and {table2} "{table2_name}"')
                                                            logger_setup.get_logger().debug(f'Failed to prepare query to insert new relationship')
                                                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                                                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                                                            rollback_savepoint('before_import_many_to_many')
                                                            return False
                                                        query.bindValue(':table1_id', table1_id)
                                                        query.bindValue(':table2_id', table2_id)
                                                        if not query.exec():
                                                        # if not query.exec() and 'UNIQUE constraint failed' not in query.lastError().text():
                                                            logger_setup.get_logger().critical(f'Error linking {table1} "{table1_name}" and {table2} "{table2_name}"')
                                                            logger_setup.get_logger().debug(f'Failed to insert values into {table}')
                                                            logger_setup.get_logger().debug(f'Error: {query.lastError().text()}')
                                                            logger_setup.get_logger().debug(f'SQL query: {query.lastQuery()}')
                                                            logger_setup.get_logger().debug(f'Bound values: {query.boundValues()}')
                                                            rollback_savepoint('before_import_many_to_many')
                                                            return False
                                                        logger_setup.get_logger().info(f'Linked {table1} "{table1_name}" (ID {table1_id}) with {table2} "{table2_name}" (ID {table2_id})')
        release_savepoint('before_import_many_to_many')
        return True


    def find_name_header(self, table: str, sheet: str):
        """
        Find the name header for a given table in a given sheet.
        :param table: table to find the name header for
        :param sheet: sheet to search in
        :return: name header and name column if found, otherwise None
        """

        if not table:
            return None, None

        name_header = None
        name_columns = []

        if table not in ['SampleAges', 'References', 'SampleGPSLocations', 'ColumnGPSLocations']:
            name_header = get_headers(table)[get_name_column(table)]
        elif table == 'References':
            if search_dictionary(self.item_ids[table][sheet], 'ReferenceDisplay'):
                name_header = 'ReferenceDisplay'
            else:
                name_header = None
        if name_header:
            for column, item_header in self.item_ids[table][sheet].items():
                item_header = list(item_header.keys())[0]
                if item_header == name_header:
                    name_columns.append(column)
        else:
            parent_table = None
            if self.item_parent_dict[table]:
                # For tables with no name column, we need to find the parent name column
                parent_table = self.item_parent_dict[table]
            if not parent_table:
                name_columns = []
            parent_name_header = get_headers(parent_table)[get_name_column(parent_table)]
            for column, item_header in self.item_ids[parent_table][sheet].items():
                item_header = list(item_header.keys())[0]
                if item_header == parent_name_header:
                    name_header = parent_name_header
                    name_columns.append(column)
            if not name_header or not name_columns:
                logger_setup.get_logger().critical(f"Unknown mapping for table {table} items")
                logger_setup.get_logger().debug(
                    f"Could not find name header {name_header} or  parent column {parent_name_header} in sheet {sheet}")
                return None, None
        return name_header, name_columns


    def find_matching_id(self, table, field_name, value) -> int :
        """

        :param table: table in the database to query from
        :param field_name: field in the table to search for
        :param value: value to find within the field
        :return: a table's primary key that matches value if found, otherwise none
        :rtype int
        """
        query = QSqlQuery()
        id_field = f'{table.strip('s')}ID'
        query.prepare(f'SELECT {id_field} FROM {table} WHERE {field_name}=:value COLLATE NOCASE')
        query.bindValue(':value', value)
        if query.exec():
            if query.next():
                return query.value(0)
        return None

    def determine_gps_format(self, import_row):
        """
        Determine the GPS format based on which fields are populated in the import_row
        :param import_row: dict of {header: value} for the current row being imported
        :return: GPSFormatID if a matching format is found, otherwise None
        :rtype: int or None
        """
        gps_format_name = None
        cardinal = False
        seconds = False
        minutes = False
        degrees = False
        for header, value in import_row.items():
            if 'UTM' in header and value not in (None, '', 'NULL'):
                # GPS format is UTM
                gps_format_name = 'UTM'
                break
            if 'Dir' in header and value not in (None, '', 'NULL'):
                # Directional indicators present, so cardinal format
                cardinal = True
            if 'Sec' in header and value not in (None, '', 'NULL'):
                seconds = True
            if 'Min' in header and value not in (None, '', 'NULL'):
                minutes = True
            if 'Deg' in header and value not in (None, '', 'NULL'):
                degrees = True
        if gps_format_name is None:
            if seconds:
                if cardinal:
                    gps_format_name = 'DMS NSEW'
                else:
                    gps_format_name = 'DMS +/-'
            elif minutes:
                if cardinal:
                    gps_format_name = 'DDM NSEW'
                else:
                    gps_format_name = 'DDM +/-'
            elif degrees:
                if cardinal:
                    gps_format_name = 'DD NSEW'
                else:
                    gps_format_name = 'DD +/-'
            else:
                # No recognizable GPS fields found
                return None
        for format_id in range(0, len(SQLUtils.gps_formats)):
            if SQLUtils.gps_formats[format_id][1] == gps_format_name:
                return format_id+1  # +1 because GPSFormatID starts at 1
        return None

    def close(self):
        self.saveWindowState()
        return super().close()

    def saveWindowState(self):
        settings.setValue("ui/ImportWizard/pos", self.pos())
        settings.setValue("ui/ImportWizard/size", self.size())

    def loadWindowState(self):
        self.move(settings.value("ui/ImportWizard/pos", defaultValue=QPoint(410, 241)))
        self.resize(settings.value("ui/ImportWizard/size", defaultValue=QSize(810, 569)))
