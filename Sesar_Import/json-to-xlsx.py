import json
from openpyxl import Workbook

# ---------------------------------------------------
# file paths
# ---------------------------------------------------
json_file_path = "geocork_import_ready.json"
excel_file_path = "geocork_output.xlsx"

# ---------------------------------------------------
# load json
# ---------------------------------------------------
with open(json_file_path, "r", encoding="utf-8") as f:
    data = json.load(f)

# ---------------------------------------------------
# create workbook
# ---------------------------------------------------
wb = Workbook()

# remove default sheet
wb.remove(wb.active)

# ---------------------------------------------------
# helper function
# ---------------------------------------------------
def write_table_to_sheet(table_name, rows):
    """
    Writes a list of dictionaries to a sheet.
    Each dictionary becomes a row.
    Keys become column headers.
    """
    ws = wb.create_sheet(title=table_name[:31])  # Excel sheet name limit

    # empty table
    if not rows:
        ws.cell(row=1, column=1, value="(empty)")
        return

    # collect all possible keys across rows
    all_keys = set()
    for row in rows:
        all_keys.update(row.keys())
    headers = sorted(all_keys)

    # write headers
    for col, key in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=key)

    # write rows
    for row_index, row in enumerate(rows, start=2):
        for col_index, key in enumerate(headers, start=1):
            value = row.get(key)

            # convert lists/dicts into readable text
            if isinstance(value, list):
                value = ", ".join(map(str, value))
            elif isinstance(value, dict):
                value = json.dumps(value)

            ws.cell(row=row_index, column=col_index, value=value)


# ---------------------------------------------------
# write each table into its own sheet
# ---------------------------------------------------
for table_name, table_rows in data.items():
    if isinstance(table_rows, list):
        write_table_to_sheet(table_name, table_rows)

# ---------------------------------------------------
# save
# ---------------------------------------------------
wb.save(excel_file_path)

print("Excel file created successfully!")
